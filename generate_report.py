#!/usr/bin/env python3
"""
언니가이드 월간 트렌드 리포트 생성기 v2
- 공통 트렌드 리포트 (전체 병원 공유) → unniguide_report_YYYYMM.html
- 병원별 개별 리포트 → hospitals/병원명_YYYYMM.html (월별 누적 포함)
"""

import pandas as pd
import json
import os
import re
from datetime import datetime
from collections import defaultdict

# ============================================================
# 1. 설정
# ============================================================
import sys
import glob

# Excel 파일: 인자로 받거나, Downloads에서 최신 파일 자동 탐색
if len(sys.argv) > 1:
    EXCEL_PATH = os.path.expanduser(sys.argv[1])
else:
    # "언니가이드 운영 트렌드 데이터_*.xlsx" 중 가장 최신 파일
    pattern = os.path.expanduser('~/Downloads/언니가이드 운영 트렌드 데이터_*.xlsx')
    candidates = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    if candidates:
        EXCEL_PATH = candidates[0]
    else:
        print("❌ Excel 파일을 찾을 수 없습니다.")
        print("   사용법: python3 generate_report.py [엑셀파일경로]")
        sys.exit(1)

OUTPUT_DIR = os.path.expanduser('~/Documents/unniguide-report')
HOSPITAL_DIR = os.path.join(OUTPUT_DIR, 'hospitals')

# REPORT_MONTH: 인자로 받거나, 데이터에서 자동 감지 (아래에서 설정)
REPORT_MONTH = sys.argv[2] if len(sys.argv) > 2 else None  # 나중에 자동 설정
REPORT_MONTH_KR = None  # 나중에 자동 설정

print(f"📂 Excel: {EXCEL_PATH}")

os.makedirs(HOSPITAL_DIR, exist_ok=True)

# ============================================================
# 브랜드 컬러 (BI Guidelines)
# ============================================================
BRAND = {
    'orange': '#FF6A3B',       # MAIN ORANGE
    'orange_dark': '#E8551F',  # darker for hover/active
    'orange_light': '#FFF0EB', # tinted bg
    'ivory': '#FBF9F1',       # IVORY
    'plum': '#330C2E',        # PLUM
    'plum_light': '#5C2D54',  # lighter plum for text
}

# 병원 마스터 리스트
HOSPITAL_MASTER = {
    1860: {'name': '우리성형외과의원', 'manager': 'Yun', 'country': '태국 제외, 중국 제외', 'area': '강남', 'type': '시수술'},
    3234: {'name': '사적인아름다움지유의원', 'manager': 'Yun', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    660:  {'name': '루호성형외과의원', 'manager': 'Yun', 'country': '모든 국가', 'area': '강남', 'type': '수술'},
    46:   {'name': '플라덴성형외과의원', 'manager': 'Yun', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    5857: {'name': '우유빛의원', 'manager': 'Lyn', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    7493: {'name': '테이아의원', 'manager': 'Yun', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    116:  {'name': '티에스성형외과', 'manager': 'Hardy', 'country': '모든 국가', 'area': '강남', 'type': '수술'},
    2067: {'name': '아크로한의원', 'manager': 'Runa', 'country': '모든 국가', 'area': '강남', 'type': '한방'},
    6890: {'name': '톡스앤필의원-신논현점', 'manager': 'Runa', 'country': '중국 제외', 'area': '강남', 'type': '시술'},
    7826: {'name': '제이필의원-홍대점', 'manager': 'Su', 'country': '모든 국가', 'area': '홍대', 'type': '시술'},
    2681: {'name': '메이필의원', 'manager': 'Runa', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    580:  {'name': '허쉬성형외과의원', 'manager': 'Hardy', 'country': '모든 국가', 'area': '강남', 'type': '수술'},
    3090: {'name': '플래너성형외과의원', 'manager': 'Winnie', 'country': '모든 국가', 'area': '강남', 'type': '수술'},
    5524: {'name': '오앤의원', 'manager': 'Runa', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    6681: {'name': '홍대셀레나의원', 'manager': 'Su', 'country': '모든 국가', 'area': '홍대', 'type': '시술'},
    6518: {'name': '네스트의원', 'manager': 'Runa', 'country': '중국 제외, 대만 제외', 'area': '강남', 'type': '시술'},
    1889: {'name': '제이필의원-강남점', 'manager': 'Su', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    2991: {'name': '플레저성형외과의원', 'manager': 'Hardy', 'country': '모든 국가', 'area': '강남', 'type': '수술'},
    3505: {'name': '히트성형외과의원', 'manager': 'Runa', 'country': '모든 국가', 'area': '강남', 'type': '수술'},
    802:  {'name': '모즈의원', 'manager': 'Runa', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    1111: {'name': '우아성형외과의원', 'manager': '병원마케팅팀', 'country': '태국만', 'area': '강남', 'type': '수술'},
    3608: {'name': '청담디어의원', 'manager': 'Belle', 'country': '대만 제외, 중국 제외', 'area': '강남', 'type': '시술'},
    5826: {'name': '유픽의원-홍대점', 'manager': 'Yun', 'country': '모든 국가', 'area': '홍대', 'type': '시술'},
    5663: {'name': '유픽의원-강남점', 'manager': 'Runa', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    392:  {'name': '라인앤뷰의원', 'manager': 'Runa', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    4575: {'name': '세가지소원의원-명동점', 'manager': 'Runa', 'country': '모든 국가', 'area': '명동', 'type': '시술'},
    4213: {'name': '올리팅성형외과의원', 'manager': 'Runa', 'country': '중국 제외, 대만 제외', 'area': '강남', 'type': '수술'},
    5451: {'name': '강남셀리팅의원', 'manager': 'Lyn', 'country': '모든 국가', 'area': '강남', 'type': '수술'},
    4414: {'name': '리디아여성의원', 'manager': 'Hardy', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    5192: {'name': '플로리아의원', 'manager': 'Yun', 'country': '모든 국가', 'area': '부산', 'type': '시술'},
    3673: {'name': '브이에스라인의원-압구정점', 'manager': 'Hardy', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    6010: {'name': '엔디어트의원', 'manager': 'Hardy', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    4848: {'name': '메종드엠의원', 'manager': 'Lyn', 'country': '태국만', 'area': '강남', 'type': '시술'},
    9999: {'name': '릴리브의원', 'manager': '병원마케팅팀', 'country': '대만 제외', 'area': '강남', 'type': '시술'},
}

NAME_NORMALIZE = {
    '사적인아름다운지유의원': '사적인아름다움지유의원',
    '루호성형외과': '루호성형외과의원',
    '우리성형외과': '우리성형외과의원',
    '테이아 의원': '테이아의원',
    '티에스성형외과의원': '티에스성형외과',
    '톡스앤필-시논현': '톡스앤필의원-신논현점',
    '톡스앤필 - 신논': '톡스앤필의원-신논현점',
    '톡스앤필 - 신논현': '톡스앤필의원-신논현점',
    '제이필 - 홍대': '제이필의원-홍대점',
    '제이필 - 강남': '제이필의원-강남점',
    '플래저성형외과': '플레저성형외과의원',
    '플래너성형외과': '플래너성형외과의원',
    '유픽의원 홍대': '유픽의원-홍대점',
    '유픽의원-홍대': '유픽의원-홍대점',
    '유픽의원-강남': '유픽의원-강남점',
    '디어 청담의원': '청담디어의원',
    '홍대셀레나': '홍대셀레나의원',
    '히트성형외과': '히트성형외과의원',
}

def normalize_hospital(name):
    if not name or str(name).strip() == '':
        return None
    name = str(name).strip()
    return NAME_NORMALIZE.get(name, name)

COUNTRY_FLAG = {
    '태국': '🇹🇭', '대만': '🇹🇼', '중국': '🇨🇳', '미국': '🇺🇸',
    '호주': '🇦🇺', '일본': '🇯🇵', '홍콩': '🇭🇰', '싱가포르': '🇸🇬',
    '베트남': '🇻🇳', '필리핀': '🇵🇭', '말레이시아': '🇲🇾', '인도네시아': '🇮🇩',
    '영국': '🇬🇧', '프랑스': '🇫🇷', '독일': '🇩🇪', '캐나다': '🇨🇦',
    '인도': '🇮🇳', '러시아': '🇷🇺', '스페인': '🇪🇸', '포르투갈': '🇵🇹',
    '뉴질랜드': '🇳🇿', '노르웨이': '🇳🇴', '폴란드': '🇵🇱', '칠레': '🇨🇱',
    '캄보디아': '🇰🇭', '몽골': '🇲🇳', '사우디 아라비아': '🇸🇦',
    '우즈베키스탄': '🇺🇿', '키르기스스탄': '🇰🇬', '키르기스탄': '🇰🇬',
    '도미니카 공화국': '🇩🇴', '아일랜드': '🇮🇪', '로마니아': '🇷🇴', '부탄': '🇧🇹',
}

def format_krw(amount):
    if amount >= 100000000:
        return f"{amount/100000000:.1f}억원"
    elif amount >= 10000:
        return f"{amount/10000:,.0f}만원"
    else:
        return f"{amount:,.0f}원"

def format_number(n):
    return f"{n:,}"

def safe_filename(name):
    return re.sub(r'[^\w가-힣\-]', '_', name)

# ============================================================
# 2. 데이터 읽기
# ============================================================
print("📊 데이터 로딩 중...")

df_res = pd.read_excel(EXCEL_PATH, sheet_name='언니가이드 예약확정 시트 데일리', header=1)
df_res.columns = [
    'NO', '채팅접수일자', '예약확정일', '담당자', '고객명', '그룹여부',
    '고객국적', '사용언어', '예약상태', '통역서비스요청', '종류',
    '시술수술명', '추천클리닉', '예약클리닉', '내원일', '시간',
    '예상금액', '실제금액', '금액확인', '설문발송여부',
    '후기작성여부', '캐시백지급대상자', '캐시백지급여부', '캐시백금액',
    '캐시백지급일자', 'Remark', '시술수술확정항목'
] + [f'extra_{i}' for i in range(max(0, len(df_res.columns) - 27))]

df_res['병원명'] = df_res['예약클리닉'].apply(normalize_hospital)
df_res['내원일'] = pd.to_datetime(df_res['내원일'], errors='coerce')
df_res['월'] = df_res['내원일'].dt.to_period('M').astype(str)
df_res['실제금액'] = pd.to_numeric(df_res['실제금액'], errors='coerce').fillna(0)
df_completed = df_res[df_res['예약상태'] == '시/수술 완료'].copy()

# 정산 데이터
import openpyxl
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws2 = wb['병원별 정산 및 언니가이드 매출 데이터']

settlement_records = []
current_month = None
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=False):
    vals = [c.value for c in row]
    a_val = str(vals[0]).strip() if vals[0] else ''
    if '정산 내역' in a_val:
        parts = a_val.replace('년', '-').replace('월', '').replace('정산 내역', '').strip()
        try:
            year, month = parts.split('-')[:2]
            current_month = f"{year.strip()}-{int(month.strip()):02d}"
        except:
            pass
        continue
    if a_val in ('NO', '', 'None', '재무팀 정산 요청 내역') or '정산 요청일' in a_val:
        continue
    if current_month and vals[1]:
        hospital = str(vals[1]).strip()
        if hospital in ('병원명', ''):
            continue
        try:
            settlement_records.append({
                '정산월': current_month,
                '병원명': normalize_hospital(hospital),
                '고객명': str(vals[2]).strip() if vals[2] else '',
                '국적': str(vals[4]).strip() if vals[4] else '',
                '구분': str(vals[6]).strip() if vals[6] else '',
                '시술금액': float(vals[7]) if vals[7] else 0,
                '수수료금액': float(vals[8]) if vals[8] else 0,
            })
        except (ValueError, TypeError):
            pass
wb.close()

df_settle = pd.DataFrame(settlement_records)
print(f"  예약 데이터: {len(df_res)}건 (완료 {len(df_completed)}건)")
print(f"  정산 데이터: {len(df_settle)}건")

# ============================================================
# 자동 월 감지
# ============================================================
all_months_raw = sorted(df_completed['월'].dropna().unique())

if REPORT_MONTH is None:
    # 정산 데이터의 최신 월 또는 예약 데이터의 최신 월
    settle_months = sorted(df_settle['정산월'].unique()) if len(df_settle) else []
    if settle_months:
        REPORT_MONTH = settle_months[-1]
    elif all_months_raw:
        REPORT_MONTH = str(all_months_raw[-1])
    else:
        print("❌ 데이터에서 월 정보를 찾을 수 없습니다.")
        sys.exit(1)

# 전월 자동 계산
from dateutil.relativedelta import relativedelta
_rm = datetime.strptime(REPORT_MONTH + '-01', '%Y-%m-%d')
PREV_MONTH = (_rm - relativedelta(months=1)).strftime('%Y-%m')

# 한글 월 표시
REPORT_MONTH_KR = f"{_rm.year}년 {_rm.month}월"

print(f"\n📅 리포트 기준월: {REPORT_MONTH_KR} (전월: {PREV_MONTH})")

# ============================================================
# 3. 지표 산출
# ============================================================
print("📈 지표 산출 중...")

all_months = sorted(df_completed['월'].dropna().unique())
all_canonical = set(h['name'] for h in HOSPITAL_MASTER.values())

# 3a. 플랫폼 월별 지표
mar_completed = df_completed[df_completed['월'] == REPORT_MONTH]
feb_completed = df_completed[df_completed['월'] == PREV_MONTH]
mar_all = df_res[df_res['월'] == REPORT_MONTH]

platform = {
    'month': REPORT_MONTH_KR,
    'completed': len(mar_completed),
    'total_revenue': mar_completed['실제금액'].sum(),
    'avg_price': mar_completed['실제금액'].mean() if len(mar_completed) else 0,
    'num_countries': mar_completed['고객국적'].nunique(),
    'num_hospitals': mar_completed['병원명'].nunique(),
    'prev_completed': len(feb_completed),
    'prev_revenue': feb_completed['실제금액'].sum(),
}
platform['completed_growth'] = round((platform['completed'] - platform['prev_completed']) / max(platform['prev_completed'], 1) * 100, 1) if platform['prev_completed'] else None
platform['revenue_growth'] = round((platform['total_revenue'] - platform['prev_revenue']) / max(platform['prev_revenue'], 1) * 100, 1) if platform['prev_revenue'] else None

cumulative = {
    'completed': len(df_completed),
    'revenue': df_completed['실제금액'].sum(),
    'countries': df_completed['고객국적'].nunique(),
    'hospitals': df_completed['병원명'].nunique(),
    'date_range': f"{df_completed['내원일'].min().strftime('%Y.%m.%d')} ~ {df_completed['내원일'].max().strftime('%Y.%m.%d')}",
}

# 3b. 국적별 트렌드
nat_group = mar_completed.groupby('고객국적').agg(건수=('고객명','count'), 총금액=('실제금액','sum')).reset_index()
nat_group['비중'] = round(nat_group['건수'] / nat_group['건수'].sum() * 100, 1)
nat_group['객단가'] = (nat_group['총금액'] / nat_group['건수']).round(0)
nat_group = nat_group.sort_values('건수', ascending=False)
nationality_stats = [
    {'country': r['고객국적'], 'flag': COUNTRY_FLAG.get(r['고객국적'],'🌍'),
     'count': int(r['건수']), 'pct': float(r['비중']),
     'revenue': float(r['총금액']), 'avg_price': float(r['객단가'])}
    for _, r in nat_group.iterrows()
]

# 3c. 국적별 선호 시술
top_countries = [n['country'] for n in nationality_stats[:8]]
nationality_procedures = {}
for country in top_countries:
    procs = mar_completed[mar_completed['고객국적'] == country]['시술수술명'].value_counts().head(5)
    nationality_procedures[country] = [{'procedure': p, 'count': int(c)} for p, c in procs.items()]

# 3d. 시술 비중
type_counts = mar_completed['종류'].value_counts()
procedure_type_stats = {str(t): {'count': int(c), 'pct': round(c/type_counts.sum()*100,1)} for t, c in type_counts.items()}

# 3e. TOP 시술
top_procs = mar_completed.groupby('시술수술명').agg(건수=('고객명','count'), 총금액=('실제금액','sum')).reset_index().sort_values('건수', ascending=False).head(15)
top_procedures = [{'procedure': r['시술수술명'], 'count': int(r['건수']), 'revenue': float(r['총금액'])} for _, r in top_procs.iterrows()]

# 3f. 병원별 월별 누적 데이터
hospital_monthly = defaultdict(lambda: defaultdict(lambda: {'completed':0, 'revenue':0, 'nationalities':{}, 'procedures':{}}))

for _, row in df_completed.iterrows():
    h = row['병원명']
    m = row['월']
    if h not in all_canonical or pd.isna(m):
        continue
    d = hospital_monthly[h][m]
    d['completed'] += 1
    d['revenue'] += row['실제금액']
    nat = row['고객국적']
    if nat:
        if nat not in d['nationalities']:
            d['nationalities'][nat] = {'count':0, 'revenue':0}
        d['nationalities'][nat]['count'] += 1
        d['nationalities'][nat]['revenue'] += row['실제금액']
    proc = row['시술수술명']
    if proc:
        d['procedures'][proc] = d['procedures'].get(proc, 0) + 1

# 정산 월별
settle_monthly = defaultdict(lambda: defaultdict(lambda: {'revenue':0, 'commission':0, 'by_nat':{}}))
for _, row in df_settle.iterrows():
    h = row['병원명']
    m = row['정산월']
    if not h:
        continue
    d = settle_monthly[h][m]
    d['revenue'] += row['시술금액']
    d['commission'] += row['수수료금액']
    nat = row['국적']
    if nat:
        if nat not in d['by_nat']:
            d['by_nat'][nat] = {'revenue':0, 'count':0}
        d['by_nat'][nat]['revenue'] += row['시술금액']
        d['by_nat'][nat]['count'] += 1

# 벤치마크
active_hospitals = {h: hospital_monthly[h][REPORT_MONTH] for h in all_canonical if hospital_monthly[h][REPORT_MONTH]['completed'] > 0}
ranked_by_rev = sorted(active_hospitals.keys(), key=lambda x: active_hospitals[x]['revenue'], reverse=True)
ranked_by_cnt = sorted(active_hospitals.keys(), key=lambda x: active_hospitals[x]['completed'], reverse=True)
total_active = len(active_hospitals)
platform_avg_completed = sum(v['completed'] for v in active_hospitals.values()) / max(total_active,1)
platform_avg_revenue = sum(v['revenue'] for v in active_hospitals.values()) / max(total_active,1)

rank_info = {}
for i, h in enumerate(ranked_by_cnt):
    rank_info.setdefault(h, {})['rank_count'] = i+1
    rank_info[h]['pct_count'] = round((1 - i/max(total_active,1))*100)
for i, h in enumerate(ranked_by_rev):
    rank_info.setdefault(h, {})['rank_revenue'] = i+1
    rank_info[h]['pct_revenue'] = round((1 - i/max(total_active,1))*100)

print(f"  3월 완료: {platform['completed']}건, 매출: {format_krw(platform['total_revenue'])}")
print(f"  활성 병원: {total_active}개, 월: {len(all_months)}개")

# ============================================================
# 4. CSS (공통)
# ============================================================
CSS = """
:root {
  --primary: """ + BRAND['orange'] + """;
  --primary-dark: """ + BRAND['orange_dark'] + """;
  --primary-light: """ + BRAND['orange_light'] + """;
  --plum: """ + BRAND['plum'] + """;
  --plum-light: """ + BRAND['plum_light'] + """;
  --ivory: """ + BRAND['ivory'] + """;
  --accent: #00B894;
  --accent-light: #E8FBF5;
  --bg: #FAFBFC;
  --card: #FFFFFF;
  --text: #2D3436;
  --text-light: #636E72;
  --border: #E9ECEF;
  --shadow: 0 2px 12px rgba(0,0,0,0.06);
  --radius: 16px;
}
* { margin:0; padding:0; box-sizing:border-box; }
body {
  font-family: -apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,'Noto Sans KR',sans-serif;
  background: var(--bg); color: var(--text); line-height:1.6;
}
.container { max-width:1100px; margin:0 auto; padding:24px 20px; }

/* Header */
.header {
  background: linear-gradient(135deg, var(--primary) 0%, var(--primary-dark) 100%);
  color: white; padding:40px 0 32px; margin-bottom:32px;
  border-radius: 0 0 32px 32px;
}
.header .container { padding-top:0; padding-bottom:0; }
.header-top { display:flex; justify-content:space-between; align-items:center; margin-bottom:8px; }
.logo { font-size:16px; font-weight:800; letter-spacing:1px; opacity:0.95; }
.logo span { font-weight:400; }
.report-date { font-size:13px; opacity:0.8; }
.header h1 { font-size:26px; font-weight:800; margin-bottom:4px; }
.header .subtitle { font-size:14px; opacity:0.85; }

/* Section */
.section { margin-bottom:32px; }
.section-title {
  font-size:19px; font-weight:700; margin-bottom:16px;
  display:flex; align-items:center; gap:8px; color:var(--plum);
}
.section-title .icon { font-size:22px; }

/* Cards */
.card {
  background:var(--card); border-radius:var(--radius);
  box-shadow:var(--shadow); padding:24px; margin-bottom:16px;
  border:1px solid var(--border);
}

.metric-grid { display:grid; grid-template-columns:repeat(auto-fit,minmax(200px,1fr)); gap:16px; }
.metric-card {
  background:var(--card); border-radius:12px; padding:20px;
  border:1px solid var(--border); text-align:center;
}
.metric-card .label { font-size:13px; color:var(--text-light); margin-bottom:4px; }
.metric-card .value { font-size:26px; font-weight:800; color:var(--plum); }
.metric-card .sub { font-size:12px; color:var(--text-light); margin-top:4px; }

/* Badges */
.badge { display:inline-block; padding:2px 10px; border-radius:12px; font-size:12px; font-weight:600; }
.badge-up { background:#E8FBF5; color:#00B894; }
.badge-down { background:#FFE8E8; color:#E74C3C; }
.badge-neutral { background:#F1F2F6; color:#636E72; }
.badge-primary { background:var(--primary-light); color:var(--primary-dark); }

/* Tables */
.data-table { width:100%; border-collapse:collapse; font-size:14px; }
.data-table th {
  background:var(--ivory); padding:12px 16px; text-align:left;
  font-weight:600; color:var(--plum); font-size:12px;
  text-transform:uppercase; letter-spacing:0.5px;
  border-bottom:2px solid var(--border);
}
.data-table td { padding:12px 16px; border-bottom:1px solid var(--border); }
.data-table tr:hover { background:var(--ivory); }
.data-table .num { text-align:right; font-variant-numeric:tabular-nums; }
.data-table .bold { font-weight:700; }

/* Charts */
.chart-container { position:relative; height:320px; margin:16px 0; }
.chart-half { display:grid; grid-template-columns:1fr 1fr; gap:24px; }
.chart-half .chart-container { height:280px; }

/* Insight */
.insight-box {
  background:linear-gradient(135deg,#FFF9E6 0%,#FFF3CC 100%);
  border-left:4px solid #F39C12; border-radius:0 12px 12px 0;
  padding:16px 20px; margin:16px 0; font-size:14px; line-height:1.7;
}
.opportunity-box {
  background:linear-gradient(135deg,var(--accent-light) 0%,#D5F5ED 100%);
  border-left:4px solid var(--accent); border-radius:0 12px 12px 0;
  padding:16px 20px; margin:16px 0;
}
.opportunity-box h4 { color:#00896A; margin-bottom:8px; }
.opportunity-box ul { margin-left:20px; }
.opportunity-box li { margin-bottom:4px; font-size:14px; }

/* Benchmark */
.benchmark-bar { height:8px; background:#E9ECEF; border-radius:4px; margin:8px 0; overflow:hidden; }
.benchmark-bar .fill {
  height:100%; border-radius:4px;
  background:linear-gradient(90deg,var(--primary) 0%,var(--primary-dark) 100%);
  transition:width 0.8s ease;
}
.benchmark-label { display:flex; justify-content:space-between; font-size:13px; color:var(--text-light); }

/* FAQ */
.faq-item { margin-bottom:16px; }
.faq-item h4 { font-size:14px; color:var(--primary-dark); margin-bottom:4px; }
.faq-item p { font-size:14px; color:var(--text-light); }

/* Nav */
.back-link {
  display:inline-flex; align-items:center; gap:6px;
  color:var(--primary); font-weight:600; text-decoration:none;
  margin-bottom:16px; font-size:14px;
}
.back-link:hover { text-decoration:underline; }

/* Monthly trend table */
.trend-table th, .trend-table td { text-align:center; padding:10px 12px; }
.trend-table th { background:var(--ivory); color:var(--plum); }
.trend-table td.highlight { background:var(--primary-light); font-weight:700; }

/* Footer */
.footer {
  text-align:center; padding:32px 0; color:var(--text-light);
  font-size:13px; border-top:1px solid var(--border); margin-top:40px;
}

/* Hospital list */
.hospital-grid { display:grid; grid-template-columns:repeat(auto-fill,minmax(280px,1fr)); gap:12px; }
.hospital-link {
  display:block; padding:16px 20px; background:var(--card);
  border:1px solid var(--border); border-radius:12px;
  text-decoration:none; color:var(--text); transition:all 0.2s;
}
.hospital-link:hover { border-color:var(--primary); box-shadow:0 4px 16px rgba(255,106,59,0.12); transform:translateY(-1px); }
.hospital-link .h-name { font-weight:700; font-size:15px; margin-bottom:4px; color:var(--plum); }
.hospital-link .h-stats { font-size:13px; color:var(--text-light); }
.hospital-link .h-stats strong { color:var(--primary); }

@media (max-width:768px) {
  .header h1 { font-size:20px; }
  .metric-grid { grid-template-columns:repeat(2,1fr); }
  .chart-half { grid-template-columns:1fr; }
  .hospital-grid { grid-template-columns:1fr; }
}
@media print {
  .back-link { display:none; }
  .card { break-inside:avoid; }
}
"""

CHART_COLORS = "['#FF6A3B','#330C2E','#00B894','#FDCB6E','#0984E3','#E17055','#00CEC9','#A29BFE','#FD79A8','#55A3E8','#F39C12','#2ECC71','#E74C3C','#9B59B6','#1ABC9C']"

FOOTER = f"""
<div class="footer">
  <div class="container">
    <p><strong>UNNI GUIDE</strong> | 강남언니 언니가이드 서비스</p>
    <p style="margin-top:4px;">본 리포트는 파트너 병원 전용 자료입니다. 무단 배포를 삼가해 주세요.</p>
    <p style="margin-top:4px;opacity:0.7;">생성일: {datetime.now().strftime('%Y년 %m월 %d일')}</p>
  </div>
</div>
"""

HEADER_HTML = lambda title, subtitle: f"""
<div class="header">
  <div class="container">
    <div class="header-top">
      <div class="logo">UNNI <span>GUIDE</span></div>
      <div class="report-date">{REPORT_MONTH_KR} 리포트</div>
    </div>
    <h1>{title}</h1>
    <div class="subtitle">{subtitle}</div>
  </div>
</div>
"""

# ============================================================
# 5. 공통 트렌드 리포트 HTML
# ============================================================
print("\n🎨 공통 트렌드 리포트 생성 중...")

def growth_badge(g):
    if g is None: return '<span class="badge badge-neutral">N/A</span>'
    if g > 0: return f'<span class="badge badge-up">▲ {g}%</span>'
    if g < 0: return f'<span class="badge badge-down">▼ {abs(g)}%</span>'
    return '<span class="badge badge-neutral">→ 0%</span>'

# 국적 테이블 rows
nat_rows = ""
for n in nationality_stats:
    nat_rows += f'<tr><td>{n["flag"]} {n["country"]}</td><td>{n["count"]}건</td><td>{n["pct"]}%</td><td class="num">{format_krw(n["revenue"])}</td><td class="num bold">{format_krw(n["avg_price"])}</td></tr>\n'

# 국적별 시술 sections
nat_proc_html = ""
for country in top_countries:
    flag = COUNTRY_FLAG.get(country, '🌍')
    procs = nationality_procedures.get(country, [])
    if procs:
        rows = "".join(f'<tr><td>{p["procedure"]}</td><td class="num">{p["count"]}건</td></tr>' for p in procs)
        nat_proc_html += f'<div class="card"><h3 style="font-size:16px;margin-bottom:12px;">{flag} {country}</h3><table class="data-table"><thead><tr><th>시술/수술명</th><th class="num">건수</th></tr></thead><tbody>{rows}</tbody></table></div>\n'

# 병원 리스트 (카드형 링크)
hospital_cards = ""
for h in ranked_by_rev:
    d = active_hospitals[h]
    fname = safe_filename(h)
    hospital_cards += f'''<a class="hospital-link" href="hospitals/{fname}_{REPORT_MONTH.replace("-","")}.html">
  <div class="h-name">{h}</div>
  <div class="h-stats">완료 <strong>{d["completed"]}건</strong> · 매출 <strong>{format_krw(d["revenue"])}</strong></div>
</a>\n'''

chart_data = json.dumps({
    'nat_labels': [n['country'] for n in nationality_stats[:10]],
    'nat_counts': [n['count'] for n in nationality_stats[:10]],
    'nat_pcts': [n['pct'] for n in nationality_stats[:10]],
    'nat_prices': [n['avg_price'] for n in nationality_stats[:10]],
    'type_labels': list(procedure_type_stats.keys()),
    'type_counts': [v['count'] for v in procedure_type_stats.values()],
    'proc_names': [p['procedure'][:20] for p in top_procedures[:10]],
    'proc_counts': [p['count'] for p in top_procedures[:10]],
}, ensure_ascii=False)

common_html = f"""<!DOCTYPE html>
<html lang="ko"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>언니가이드 월간 트렌드 리포트 | {REPORT_MONTH_KR}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>{CSS}</style>
</head><body>

{HEADER_HTML('언니가이드 월간 트렌드 리포트', '파트너 병원을 위한 시장 인사이트 & 성과 분석')}

<div class="container">

<!-- 전체 요약 -->
<div class="section">
  <div class="section-title"><span class="icon">📊</span> {REPORT_MONTH_KR} 전체 요약</div>
  <div class="metric-grid">
    <div class="metric-card"><div class="label">시/수술 완료</div><div class="value">{platform['completed']}건</div><div class="sub">{growth_badge(platform['completed_growth'])}</div></div>
    <div class="metric-card"><div class="label">총 병원 매출</div><div class="value">{format_krw(platform['total_revenue'])}</div><div class="sub">{growth_badge(platform['revenue_growth'])}</div></div>
    <div class="metric-card"><div class="label">평균 객단가</div><div class="value">{format_krw(platform['avg_price'])}</div><div class="sub">1인 평균</div></div>
    <div class="metric-card"><div class="label">참여 국적</div><div class="value">{platform['num_countries']}개국</div><div class="sub">{platform['num_hospitals']}개 병원</div></div>
  </div>
</div>

<!-- 누적 -->
<div class="section">
  <div class="section-title"><span class="icon">🏆</span> 누적 성과 ({cumulative['date_range']})</div>
  <div class="metric-grid">
    <div class="metric-card"><div class="label">총 완료</div><div class="value">{format_number(cumulative['completed'])}건</div></div>
    <div class="metric-card"><div class="label">총 매출</div><div class="value">{format_krw(cumulative['revenue'])}</div></div>
    <div class="metric-card"><div class="label">참여 국적</div><div class="value">{cumulative['countries']}개국</div></div>
    <div class="metric-card"><div class="label">파트너 병원</div><div class="value">{cumulative['hospitals']}개</div></div>
  </div>
</div>

<!-- 국적별 -->
<div class="section">
  <div class="section-title"><span class="icon">🌍</span> 국적별 고객 데이터</div>
  <div class="card">
    <div class="chart-half">
      <div><h3 style="font-size:15px;margin-bottom:8px;">국적별 예약 비중</h3><div class="chart-container"><canvas id="natPie"></canvas></div></div>
      <div><h3 style="font-size:15px;margin-bottom:8px;">국적별 인당 평균 객단가</h3><div class="chart-container"><canvas id="natBar"></canvas></div></div>
    </div>
  </div>
  <div class="card">
    <table class="data-table">
      <thead><tr><th>국적</th><th>건수</th><th>비중</th><th class="num">총 매출</th><th class="num">인당 객단가</th></tr></thead>
      <tbody>{nat_rows}</tbody>
    </table>
  </div>
</div>

<!-- 국적별 시술 -->
<div class="section">
  <div class="section-title"><span class="icon">🔍</span> 국적별 선호 시술 TOP 5</div>
  {nat_proc_html}
</div>

<!-- 시술 트렌드 -->
<div class="section">
  <div class="section-title"><span class="icon">💉</span> 시술 트렌드</div>
  <div class="card">
    <div class="chart-half">
      <div><h3 style="font-size:15px;margin-bottom:8px;">시술 vs 수술 비중</h3><div class="chart-container"><canvas id="typeChart"></canvas></div></div>
      <div><h3 style="font-size:15px;margin-bottom:8px;">인기 시술 TOP 10</h3><div class="chart-container"><canvas id="procChart"></canvas></div></div>
    </div>
  </div>
</div>

<!-- 병원별 리포트 링크 -->
<div class="section">
  <div class="section-title"><span class="icon">🏥</span> 병원별 상세 리포트</div>
  <div class="hospital-grid">{hospital_cards}</div>
</div>

<!-- FAQ -->
<div class="section">
  <div class="section-title"><span class="icon">💡</span> 언니가이드 더 잘 활용하는 법</div>
  <div class="card">
    <div class="faq-item"><h4>Q1. 수수료 구조는?</h4><p>시술: 총 모객 금액의 <strong>10%</strong> / 수술: 총 모객 금액의 <strong>20%</strong></p></div>
    <div class="faq-item"><h4>Q2. 예약을 더 많이 받으려면?</h4><p>언니가이드에서 예약이 활발한 병원들의 공통점은 다음과 같습니다.</p>
      <ul style="margin:8px 0 0 20px;font-size:14px;color:var(--text-light);line-height:1.8;">
        <li>병원에서 보유한 <strong>전체 장비 리스트</strong>를 전달</li>
        <li>병원 보유 <strong>전체 수가를 투명하게 전달</strong>하여 상담사들이 빠르게 고객 안내 가능</li>
        <li>오프라인 센터의 경우, 당일·익일 예약이 가장 많아 원내 <strong>언어 응대자 보유</strong>가 큰 메리트</li>
        <li>고객 예약 및 수가 확인 요청 인입 시 <strong>1시간 이내 답변</strong></li>
        <li>언니가이드 컨설턴트분들과의 <strong>적극적인 소통 및 상호 피드백</strong> 교환</li>
      </ul>
    </div>
    <div class="faq-item"><h4>Q3. 언니가이드의 차별점은?</h4><p>한국인과 동일 수가 보장 (Equal Price), 4M+ 리뷰 기반 검증, 전담 컨설턴트 1:1 매칭.</p></div>
  </div>
</div>

</div>
{FOOTER}

<script>
const D = {chart_data};
const C = {CHART_COLORS};
new Chart(document.getElementById('natPie'),{{type:'doughnut',data:{{labels:D.nat_labels,datasets:[{{data:D.nat_counts,backgroundColor:C,borderWidth:2,borderColor:'#fff'}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'right',labels:{{font:{{size:12}}}}}}}}}}}});
new Chart(document.getElementById('natBar'),{{type:'bar',data:{{labels:D.nat_labels,datasets:[{{data:D.nat_prices,backgroundColor:C.map(c=>c+'88'),borderColor:C,borderWidth:1,borderRadius:6}}]}},options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{callback:v=>(v/10000).toFixed(0)+'만'}}}},y:{{grid:{{display:false}}}}}}}}}});
new Chart(document.getElementById('typeChart'),{{type:'doughnut',data:{{labels:D.type_labels,datasets:[{{data:D.type_counts,backgroundColor:['#FF6A3B','#330C2E','#00B894'],borderWidth:2,borderColor:'#fff'}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'bottom'}}}}}}}});
new Chart(document.getElementById('procChart'),{{type:'bar',data:{{labels:D.proc_names,datasets:[{{data:D.proc_counts,backgroundColor:'#FF6A3B88',borderColor:'#FF6A3B',borderWidth:1,borderRadius:6}}]}},options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',plugins:{{legend:{{display:false}}}},scales:{{x:{{grid:{{color:'#F1F2F6'}}}},y:{{grid:{{display:false}},ticks:{{font:{{size:11}}}}}}}}}}}});
</script>
</body></html>"""

common_path = os.path.join(OUTPUT_DIR, f'unniguide_report_{REPORT_MONTH.replace("-","")}.html')
with open(common_path, 'w', encoding='utf-8') as f:
    f.write(common_html)
print(f"  ✅ {common_path}")

# ============================================================
# 6. 병원별 개별 HTML 생성
# ============================================================
print("\n🏥 병원별 개별 리포트 생성 중...")

# 플랫폼 TOP 시술 세트
platform_top_proc_set = set(p['procedure'] for p in top_procedures[:10])

generated = 0
for h_name in all_canonical:
    # 데이터가 있는 병원만
    h_months = hospital_monthly[h_name]
    s_months = settle_monthly[h_name]
    all_h_months = sorted(set(list(h_months.keys()) + list(s_months.keys())))

    if not all_h_months:
        continue

    # 현재 월 데이터
    mar = h_months.get(REPORT_MONTH, {'completed':0,'revenue':0,'nationalities':{},'procedures':{}})
    mar_settle = s_months.get(REPORT_MONTH, {'revenue':0,'commission':0,'by_nat':{}})
    feb = h_months.get(PREV_MONTH, {'completed':0,'revenue':0})

    # 누적 합산
    cum_completed = sum(h_months[m]['completed'] for m in h_months)
    cum_revenue = sum(h_months[m]['revenue'] for m in h_months)
    cum_settle_rev = sum(s_months[m]['revenue'] for m in s_months)
    cum_settle_comm = sum(s_months[m]['commission'] for m in s_months)

    # 성장률
    count_growth = round((mar['completed'] - feb['completed']) / max(feb['completed'],1) * 100, 1) if feb['completed'] else None

    # 벤치마크
    ri = rank_info.get(h_name, {})

    # 국적 차트 데이터
    nat_sorted = sorted(mar['nationalities'].items(), key=lambda x: x[1]['count'], reverse=True)
    # 정산 국적
    settle_nat_sorted = sorted(mar_settle['by_nat'].items(), key=lambda x: x[1]['revenue'], reverse=True)

    # 시술 정렬
    proc_sorted = sorted(mar['procedures'].items(), key=lambda x: x[1], reverse=True)[:8]

    # 성장기회
    hosp_procs = set(mar['procedures'].keys())
    growth_opps = list(platform_top_proc_set - hosp_procs)[:5]

    # --- 월별 추이 테이블 ---
    monthly_data = []
    running_completed = 0
    running_revenue = 0
    for m in sorted(set(list(h_months.keys()) + list(s_months.keys()))):
        hd = h_months.get(m, {'completed':0,'revenue':0})
        sd = s_months.get(m, {'revenue':0,'commission':0})
        running_completed += hd['completed']
        running_revenue += hd['revenue']
        monthly_data.append({
            'month': m,
            'completed': hd['completed'],
            'revenue': hd['revenue'],
            'settle_rev': sd['revenue'],
            'settle_comm': sd['commission'],
            'cum_completed': running_completed,
            'cum_revenue': running_revenue,
        })

    # 월별 추이 테이블 HTML
    monthly_headers = "".join(f'<th>{"" if md["month"] != REPORT_MONTH else ""}{md["month"][-2:]}월</th>' for md in monthly_data)
    monthly_completed_cells = "".join(f'<td class="{"highlight" if md["month"]==REPORT_MONTH else ""}">{md["completed"]}건</td>' for md in monthly_data)
    monthly_settle_cells = "".join(f'<td class="{"highlight" if md["month"]==REPORT_MONTH else ""}">{format_krw(md["settle_rev"])}</td>' for md in monthly_data)
    monthly_cum_cells = "".join(f'<td class="{"highlight" if md["month"]==REPORT_MONTH else ""}">{md["cum_completed"]}건</td>' for md in monthly_data)

    # 국적 테이블
    nat_table_rows = ""
    for nat, nd in nat_sorted:
        flag = COUNTRY_FLAG.get(nat, '🌍')
        avg_p = nd['revenue'] / nd['count'] if nd['count'] else 0
        nat_table_rows += f'<tr><td>{flag} {nat}</td><td class="num">{nd["count"]}건</td><td class="num">{format_krw(nd["revenue"])}</td><td class="num bold">{format_krw(avg_p)}</td></tr>'

    # 정산 국적 테이블
    settle_nat_rows = ""
    for nat, sd in settle_nat_sorted:
        flag = COUNTRY_FLAG.get(nat, '🌍')
        avg_p = sd['revenue'] / sd['count'] if sd['count'] else 0
        settle_nat_rows += f'<tr><td>{flag} {nat}</td><td class="num">{sd["count"]}건</td><td class="num">{format_krw(sd["revenue"])}</td><td class="num bold">{format_krw(avg_p)}</td></tr>'

    # 시술 테이블
    proc_rows = "".join(f'<tr><td>{p[0]}</td><td class="num">{p[1]}건</td></tr>' for p in proc_sorted)

    # 성장기회 HTML
    growth_html = ""
    if growth_opps:
        items = "".join(f'<li>{g}</li>' for g in growth_opps)
        growth_html = f'<div class="opportunity-box"><h4>플랫폼에서 인기 있지만 귀원에서 아직 예약이 없는 시술</h4><p style="font-size:13px;color:var(--text-light);margin-bottom:8px;">아래 시술을 추가 등록하시면 더 많은 고객 매칭이 가능합니다.</p><ul>{items}</ul></div>'

    # 차트 데이터
    h_chart = json.dumps({
        'nat_labels': [n[0] for n in nat_sorted],
        'nat_counts': [n[1]['count'] for n in nat_sorted],
        'nat_prices': [round(n[1]['revenue']/n[1]['count']) if n[1]['count'] else 0 for n in nat_sorted],
        'monthly_labels': [md['month'][-2:]+'월' for md in monthly_data],
        'monthly_completed': [md['completed'] for md in monthly_data],
        'monthly_revenue': [md['revenue'] for md in monthly_data],
        'monthly_cum': [md['cum_completed'] for md in monthly_data],
    }, ensure_ascii=False)

    # 벤치마크 HTML
    pct_c = ri.get('pct_count', 0)
    pct_r = ri.get('pct_revenue', 0)
    top_pct_c = max(1, 100 - pct_c + round(100/max(total_active,1)))
    top_pct_r = max(1, 100 - pct_r + round(100/max(total_active,1)))

    hosp_html = f"""<!DOCTYPE html>
<html lang="ko"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{h_name} | 언니가이드 리포트 {REPORT_MONTH_KR}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>{CSS}</style>
</head><body>

{HEADER_HTML(h_name, f'{REPORT_MONTH_KR} 파트너 병원 성과 리포트')}

<div class="container">

<a class="back-link" href="../unniguide_report_{REPORT_MONTH.replace('-','')}.html">← 전체 트렌드 리포트로 돌아가기</a>

<!-- 당월 성과 -->
<div class="section">
  <div class="section-title"><span class="icon">📊</span> {REPORT_MONTH_KR} 성과</div>
  <div class="metric-grid">
    <div class="metric-card"><div class="label">시/수술 완료</div><div class="value">{mar['completed']}건</div><div class="sub">{growth_badge(count_growth)} 전월 {feb['completed']}건</div></div>
    <div class="metric-card"><div class="label">정산 매출</div><div class="value">{format_krw(mar_settle['revenue'])}</div><div class="sub">수수료 {format_krw(mar_settle['commission'])}</div></div>
    <div class="metric-card"><div class="label">인당 객단가</div><div class="value">{format_krw(mar['revenue']/mar['completed']) if mar['completed'] else '-'}</div><div class="sub">실 결제 기준</div></div>
    <div class="metric-card"><div class="label">누적 완료</div><div class="value">{cum_completed}건</div><div class="sub">총 {format_krw(cum_revenue)}</div></div>
  </div>
</div>

<!-- 월별 추이 -->
<div class="section">
  <div class="section-title"><span class="icon">📈</span> 월별 추이</div>
  <div class="card">
    <div class="chart-container" style="height:260px;"><canvas id="monthlyChart"></canvas></div>
    <table class="data-table trend-table" style="margin-top:16px;">
      <thead><tr><th>구분</th>{monthly_headers}</tr></thead>
      <tbody>
        <tr><td class="bold">완료 건수</td>{monthly_completed_cells}</tr>
        <tr><td class="bold">정산 금액</td>{monthly_settle_cells}</tr>
        <tr><td class="bold">누적 완료</td>{monthly_cum_cells}</tr>
      </tbody>
    </table>
  </div>
</div>

<!-- 벤치마크 -->
<div class="section">
  <div class="section-title"><span class="icon">🎯</span> 플랫폼 벤치마크</div>
  <div class="card">
    <p style="font-size:14px;color:var(--text-light);margin-bottom:16px;">언니가이드 파트너 병원 {total_active}개 중 귀원의 위치</p>
    <div style="margin-bottom:16px;">
      <div class="benchmark-label"><span>예약 건수 기준</span><span class="badge badge-primary">상위 {top_pct_c}%</span></div>
      <div class="benchmark-bar"><div class="fill" style="width:{pct_c}%"></div></div>
      <div style="font-size:12px;color:var(--text-light);">귀원 {mar['completed']}건 / 플랫폼 평균 {platform_avg_completed:.1f}건</div>
    </div>
    <div>
      <div class="benchmark-label"><span>매출 기준</span><span class="badge badge-primary">상위 {top_pct_r}%</span></div>
      <div class="benchmark-bar"><div class="fill" style="width:{pct_r}%"></div></div>
      <div style="font-size:12px;color:var(--text-light);">귀원 {format_krw(mar['revenue'])} / 플랫폼 평균 {format_krw(platform_avg_revenue)}</div>
    </div>
  </div>
</div>

<!-- 국적별 -->
<div class="section">
  <div class="section-title"><span class="icon">🌏</span> 귀원 고객 국적 분포</div>
  <div class="card">
    <div class="chart-half">
      <div><h3 style="font-size:15px;margin-bottom:8px;">국적 비중</h3><div class="chart-container"><canvas id="natPie"></canvas></div></div>
      <div><h3 style="font-size:15px;margin-bottom:8px;">국적별 객단가</h3><div class="chart-container"><canvas id="natBar"></canvas></div></div>
    </div>
  </div>
  {"<div class='card'><h3 style='font-size:15px;margin-bottom:12px;'>국적별 예약 내역 (실 결제 기준)</h3><table class='data-table'><thead><tr><th>국적</th><th class='num'>건수</th><th class='num'>매출</th><th class='num'>인당 객단가</th></tr></thead><tbody>" + nat_table_rows + "</tbody></table></div>" if nat_table_rows else ""}
  {"<div class='card'><h3 style='font-size:15px;margin-bottom:12px;'>국적별 정산 내역</h3><table class='data-table'><thead><tr><th>국적</th><th class='num'>건수</th><th class='num'>시수술 금액</th><th class='num'>건당 객단가</th></tr></thead><tbody>" + settle_nat_rows + "</tbody></table></div>" if settle_nat_rows else ""}
</div>

<!-- 시술 -->
{"<div class='section'><div class='section-title'><span class='icon'>💉</span> 귀원 인기 시술</div><div class='card'><table class='data-table'><thead><tr><th>시술/수술명</th><th class='num'>건수</th></tr></thead><tbody>" + proc_rows + "</tbody></table></div></div>" if proc_rows else ""}

<!-- 성장기회 -->
{"<div class='section'><div class='section-title'><span class='icon'>🚀</span> 성장 기회</div>" + growth_html + "</div>" if growth_html else ""}

<!-- FAQ -->
<div class="section">
  <div class="section-title"><span class="icon">💡</span> 언니가이드 활용 가이드</div>
  <div class="card">
    <div class="faq-item"><h4>Q1. 수수료 구조는?</h4><p>시술: 총 모객 금액의 <strong>10%</strong> / 수술: 총 모객 금액의 <strong>20%</strong></p></div>
    <div class="faq-item"><h4>Q2. 예약을 더 많이 받으려면?</h4><p>언니가이드에서 예약이 활발한 병원들의 공통점은 다음과 같습니다.</p>
      <ul style="margin:8px 0 0 20px;font-size:14px;color:var(--text-light);line-height:1.8;">
        <li>병원에서 보유한 <strong>전체 장비 리스트</strong>를 전달</li>
        <li>병원 보유 <strong>전체 수가를 투명하게 전달</strong>하여 상담사들이 빠르게 고객 안내 가능</li>
        <li>오프라인 센터의 경우, 당일·익일 예약이 가장 많아 원내 <strong>언어 응대자 보유</strong>가 큰 메리트</li>
        <li>고객 예약 및 수가 확인 요청 인입 시 <strong>1시간 이내 답변</strong></li>
        <li>언니가이드 컨설턴트분들과의 <strong>적극적인 소통 및 상호 피드백</strong> 교환</li>
      </ul>
    </div>
    <div class="faq-item"><h4>Q3. 언니가이드의 차별점은?</h4><p>한국인과 동일 수가 보장 (Equal Price), 4M+ 리뷰 기반 검증, 전담 컨설턴트 1:1 매칭.</p></div>
  </div>
</div>

</div>
{FOOTER}

<script>
const D = {h_chart};
const C = {CHART_COLORS};

// 월별 추이 차트
new Chart(document.getElementById('monthlyChart'),{{
  type:'bar',
  data:{{
    labels:D.monthly_labels,
    datasets:[
      {{label:'완료 건수',data:D.monthly_completed,backgroundColor:'#FF6A3B88',borderColor:'#FF6A3B',borderWidth:1,borderRadius:6,yAxisID:'y'}},
      {{label:'누적 완료',data:D.monthly_cum,type:'line',borderColor:'#330C2E',backgroundColor:'#330C2E22',pointRadius:4,tension:0.3,yAxisID:'y2'}}
    ]
  }},
  options:{{
    responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'top'}}}},
    scales:{{
      y:{{position:'left',title:{{display:true,text:'월별 건수'}},grid:{{color:'#F1F2F6'}}}},
      y2:{{position:'right',title:{{display:true,text:'누적 건수'}},grid:{{display:false}}}}
    }}
  }}
}});

// 국적 차트
if(D.nat_labels.length>0){{
  new Chart(document.getElementById('natPie'),{{type:'doughnut',data:{{labels:D.nat_labels,datasets:[{{data:D.nat_counts,backgroundColor:C,borderWidth:2,borderColor:'#fff'}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'right',labels:{{font:{{size:12}}}}}}}}}}}});
  new Chart(document.getElementById('natBar'),{{type:'bar',data:{{labels:D.nat_labels,datasets:[{{data:D.nat_prices,backgroundColor:C.map(c=>c+'88'),borderColor:C,borderWidth:1,borderRadius:6}}]}},options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{callback:v=>(v/10000).toFixed(0)+'만'}}}},y:{{grid:{{display:false}}}}}}}}}});
}}
</script>
</body></html>"""

    fname = safe_filename(h_name)
    fpath = os.path.join(HOSPITAL_DIR, f'{fname}_{REPORT_MONTH.replace("-","")}.html')
    with open(fpath, 'w', encoding='utf-8') as f:
        f.write(hosp_html)
    generated += 1

print(f"  ✅ 병원별 리포트 {generated}개 생성 완료 → {HOSPITAL_DIR}/")
print(f"\n🎉 전체 완료!")
print(f"  📄 공통 리포트: {common_path}")
print(f"  🏥 병원별 리포트: {HOSPITAL_DIR}/ ({generated}개)")
