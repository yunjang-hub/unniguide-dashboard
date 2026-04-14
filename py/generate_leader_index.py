#!/usr/bin/env python3
"""
리더용 비교 대시보드: 상위 20% 병원 vs 그 외 병원 인덱스
- 각 병원 리포트로 바로 이동하는 카드형 그리드
- 상위 20% / 그 외 섹션으로 구분
"""
import pandas as pd
import openpyxl
import os, re
import sys
from collections import defaultdict
from datetime import datetime

REPORT_MONTH = sys.argv[2] if len(sys.argv) > 2 else '2026-03'
EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser('~/Downloads/언니가이드 운영 트렌드 데이터_26.04.xlsx')
OUTPUT_DIR = os.path.expanduser('~/Documents/unniguide-report')
month_str = REPORT_MONTH.replace('-', '')
report_month_kr = f"{REPORT_MONTH.split('-')[0]}년 {int(REPORT_MONTH.split('-')[1])}월"

NAME_NORMALIZE = {
    '사적인아름다운지유의원': '사적인아름다움지유의원', '루호성형외과': '루호성형외과의원',
    '테이아 의원': '테이아의원', '톡스앤필 - 신논': '톡스앤필의원-신논현점',
    '톡스앤필 - 신논현': '톡스앤필의원-신논현점', '제이필 - 홍대': '제이필의원-홍대점',
    '제이필 - 강남': '제이필의원-강남점', '유픽의원 홍대': '유픽의원-홍대점',
    '유픽의원-홍대': '유픽의원-홍대점', '유픽의원-강남': '유픽의원-강남점',
    '홍대셀레나': '홍대셀레나의원', '히트성형외과': '히트성형외과의원',
    '플래너성형외과': '플래너성형외과의원', '플래저성형외과': '플레저성형외과의원',
}

def normalize_hospital(name):
    if not name or str(name).strip() in ('', 'nan', 'None'):
        return None
    name = str(name).strip()
    return NAME_NORMALIZE.get(name, name)

def format_krw(amount):
    if amount >= 100_000_000: return f"{amount/100_000_000:.1f}억"
    elif amount >= 10_000: return f"{amount/10_000:,.0f}만원"
    else: return f"{amount:,.0f}원"

def safe_filename(name):
    return re.sub(r'[^\w가-힣\-]', '_', name)

# ============================================================
# 정산 데이터 수집 (해당 월)
# ============================================================
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
settle_sheet = None
for name in wb.sheetnames:
    if '정산' in name:
        settle_sheet = name
        break
ws2 = wb[settle_sheet]

settle_records = []
current_month = None
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=False):
    vals = [c.value for c in row]
    a_val = str(vals[0]).strip() if vals[0] else ''
    if '정산 내역' in a_val:
        parts = a_val.replace('년', '-').replace('월', '').replace('정산 내역', '').strip()
        try:
            y, m = parts.split('-')[:2]
            current_month = f"{y.strip()}-{int(m.strip()):02d}"
        except: pass
        continue
    if a_val in ('NO', 'NO.', '', 'None', '재무팀 정산 요청 내역') or '정산 요청일' in a_val:
        continue
    if current_month and vals[1]:
        hospital = str(vals[1]).strip()
        if hospital in ('병원명', ''): continue
        try:
            settle_records.append({
                '정산월': current_month,
                '병원명': normalize_hospital(hospital),
                '시술금액': float(vals[7]) if vals[7] else 0,
                '수수료금액': float(vals[8]) if vals[8] else 0,
            })
        except: pass
wb.close()

df_settle = pd.DataFrame(settle_records)
df_target = df_settle[df_settle['정산월'] == REPORT_MONTH]

# 병원별 집계
h_agg = df_target.groupby('병원명').agg(
    건수=('시술금액', 'count'),
    매출=('시술금액', 'sum'),
    수수료=('수수료금액', 'sum'),
).sort_values('매출', ascending=False).reset_index()

h_agg['순위'] = range(1, len(h_agg) + 1)
total = len(h_agg)
top_20_threshold = max(1, round(total * 0.2))

# ============================================================
# 카드 HTML 생성
# ============================================================
def hospital_card(row, is_top=False):
    fname = safe_filename(row['병원명'])
    href = f"hospitals/{fname}_{month_str}.html"
    rank_badge = f'<span class="rank-badge top">🏆 {row["순위"]}위</span>' if is_top else f'<span class="rank-badge">{row["순위"]}위</span>'

    return f'''
    <a class="hospital-card {'top' if is_top else ''}" href="{href}" target="_blank">
      <div class="card-top">
        {rank_badge}
        <div class="h-name">{row['병원명']}</div>
      </div>
      <div class="card-stats">
        <div class="stat"><span class="label">매출</span><span class="value">{format_krw(row['매출'])}</span></div>
        <div class="stat"><span class="label">건수</span><span class="value">{int(row['건수'])}건</span></div>
        <div class="stat"><span class="label">수수료</span><span class="value">{format_krw(row['수수료'])}</span></div>
      </div>
      <div class="card-foot">리포트 보기 →</div>
    </a>'''

top_cards = ""
rest_cards = ""
for _, r in h_agg.iterrows():
    if r['순위'] <= top_20_threshold:
        top_cards += hospital_card(r, is_top=True)
    else:
        rest_cards += hospital_card(r, is_top=False)

top_revenue_sum = h_agg[h_agg['순위'] <= top_20_threshold]['매출'].sum()
total_revenue = h_agg['매출'].sum()
top_share = top_revenue_sum / total_revenue * 100 if total_revenue > 0 else 0

# ============================================================
# HTML
# ============================================================
html = f"""<!DOCTYPE html>
<html lang="ko"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>언니가이드 병원 리포트 인덱스 | {report_month_kr}</title>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans KR',sans-serif; color:#2D3436; line-height:1.5; background:#FAFBFC; }}
.container {{ max-width:1200px; margin:0 auto; padding:24px 20px; }}

/* Sticky Nav */
.top-nav {{ position:sticky; top:0; z-index:100; background:rgba(255,255,255,0.95); backdrop-filter:blur(10px); border-bottom:1px solid #E9ECEF; padding:10px 0; }}
.top-nav-inner {{ max-width:1200px; margin:0 auto; padding:0 20px; display:flex; gap:16px; align-items:center; }}
.top-nav a {{ text-decoration:none; color:#636E72; font-size:13px; font-weight:600; padding:6px 12px; border-radius:8px; transition:all 0.15s; }}
.top-nav a:hover {{ background:#FFF0EB; color:#FF6A3B; }}
.top-nav a.active {{ background:#FF6A3B; color:white; }}
.top-nav .spacer {{ flex:1; }}
.top-nav .search {{ padding:6px 12px; border:1px solid #E9ECEF; border-radius:8px; font-size:13px; width:200px; outline:none; transition:border 0.15s; }}
.top-nav .search:focus {{ border-color:#FF6A3B; }}

.header {{ background:linear-gradient(135deg,#FF6A3B 0%,#E8551F 100%); color:white; padding:40px 0 32px; margin-bottom:32px; border-radius:0 0 32px 32px; }}
.header-inner {{ max-width:1200px; margin:0 auto; padding:0 20px; }}
.header .logo {{ font-size:16px; font-weight:800; letter-spacing:1px; opacity:0.95; margin-bottom:8px; }}
.header h1 {{ font-size:28px; font-weight:800; margin-bottom:6px; }}
.header .sub {{ font-size:14px; opacity:0.9; }}

.summary-strip {{ display:grid; grid-template-columns:repeat(4,1fr); gap:12px; margin-bottom:32px; }}
.stat-card {{ background:white; border:1px solid #E9ECEF; border-radius:12px; padding:16px 20px; box-shadow:0 2px 8px rgba(0,0,0,0.04); }}
.stat-card .label {{ font-size:12px; color:#636E72; margin-bottom:4px; }}
.stat-card .value {{ font-size:22px; font-weight:800; color:#330C2E; }}
.stat-card .sub {{ font-size:11px; color:#FF6A3B; font-weight:600; margin-top:2px; }}

.section {{ margin-bottom:40px; }}
.section-header {{ display:flex; justify-content:space-between; align-items:flex-end; margin-bottom:16px; padding-bottom:12px; border-bottom:2px solid #FF6A3B33; }}
.section-title {{ font-size:20px; font-weight:800; color:#330C2E; display:flex; align-items:center; gap:10px; }}
.section-title .badge {{ background:#FFF0EB; color:#FF6A3B; font-size:13px; font-weight:700; padding:4px 12px; border-radius:16px; }}
.section-desc {{ font-size:13px; color:#636E72; }}

.grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(280px,1fr)); gap:14px; }}

.hospital-card {{
  display:block; padding:18px; background:white; border:1px solid #E9ECEF;
  border-radius:14px; text-decoration:none; color:inherit;
  transition:all 0.2s; box-shadow:0 2px 8px rgba(0,0,0,0.04);
}}
.hospital-card:hover {{ border-color:#FF6A3B; box-shadow:0 6px 20px rgba(255,106,59,0.15); transform:translateY(-2px); }}
.hospital-card.top {{ border-left:4px solid #FF6A3B; background:linear-gradient(135deg,#FFFFFF 0%,#FFF8F5 100%); }}

.card-top {{ display:flex; flex-direction:column; gap:6px; margin-bottom:12px; }}
.rank-badge {{ display:inline-block; background:#F1F2F6; color:#636E72; font-size:11px; font-weight:700; padding:3px 10px; border-radius:10px; width:fit-content; }}
.rank-badge.top {{ background:#FFF0EB; color:#FF6A3B; }}
.h-name {{ font-size:16px; font-weight:700; color:#330C2E; }}

.card-stats {{ display:grid; grid-template-columns:repeat(3,1fr); gap:8px; padding:10px 0; border-top:1px solid #F1F2F6; border-bottom:1px solid #F1F2F6; }}
.card-stats .stat {{ text-align:center; }}
.card-stats .label {{ display:block; font-size:10px; color:#636E72; margin-bottom:2px; }}
.card-stats .value {{ display:block; font-size:13px; font-weight:700; color:#330C2E; }}

.card-foot {{ font-size:12px; color:#FF6A3B; font-weight:600; margin-top:10px; text-align:right; }}

.common-card {{
  display:block; padding:24px; background:linear-gradient(135deg,#330C2E 0%,#5C2D54 100%); color:white;
  border-radius:14px; text-decoration:none; margin-bottom:32px; box-shadow:0 4px 16px rgba(51,12,46,0.2);
  transition:all 0.2s;
}}
.common-card:hover {{ transform:translateY(-2px); box-shadow:0 8px 24px rgba(51,12,46,0.3); }}
.common-card h2 {{ font-size:22px; font-weight:800; margin-bottom:6px; }}
.common-card p {{ font-size:14px; opacity:0.85; }}
.common-card .arrow {{ float:right; font-size:24px; opacity:0.6; }}

.footer {{ text-align:center; padding:40px 0 20px; color:#636E72; font-size:12px; border-top:1px solid #E9ECEF; margin-top:40px; }}
.hidden {{ display:none !important; }}

@media (max-width:768px) {{
  .summary-strip {{ grid-template-columns:repeat(2,1fr); }}
  .grid {{ grid-template-columns:1fr; }}
}}
</style>
</head><body>

<!-- Sticky Nav -->
<div class="top-nav">
  <div class="top-nav-inner">
    <a href="index.html" class="active">🏠 Index</a>
    <a href="unniguide_report_{month_str}.html">📈 전체 트렌드</a>
    <a href="#hospitals-section">🏥 병원별 리포트</a>
    <div class="spacer"></div>
    <input type="text" class="search" id="hospSearch" placeholder="병원명 검색..." onkeyup="filterHospitals(this.value)" />
  </div>
</div>

<div class="header">
  <div class="header-inner">
    <div class="logo">UNNI GUIDE</div>
    <h1>📊 병원 리포트 인덱스</h1>
    <div class="sub">{report_month_kr} · 정산 병원 {total}개 성과 한눈에 보기</div>
  </div>
</div>

<div class="container">

  <!-- 요약 -->
  <div class="summary-strip">
    <div class="stat-card"><div class="label">총 정산 병원</div><div class="value">{total}개</div><div class="sub">정산 기준</div></div>
    <div class="stat-card"><div class="label">총 정산 매출</div><div class="value">{format_krw(total_revenue)}</div><div class="sub">{report_month_kr}</div></div>
    <div class="stat-card"><div class="label">상위 20% 병원</div><div class="value">{top_20_threshold}개</div><div class="sub">TOP tier</div></div>
    <div class="stat-card"><div class="label">상위 20% 매출 비중</div><div class="value">{top_share:.1f}%</div><div class="sub">전체 매출 중</div></div>
  </div>

  <!-- 공통 리포트 -->
  <a class="common-card" href="unniguide_report_{month_str}.html" target="_blank">
    <div class="arrow">→</div>
    <h2>📈 언니가이드 서비스 전체 트렌드 리포트</h2>
    <p>{report_month_kr} 전체 플랫폼 시장 인사이트 · 국적별 트렌드 · 인기 시술 · 병원 전체 현황</p>
  </a>

  <!-- 상위 20% -->
  <div class="section" id="hospitals-section">
    <div class="section-header">
      <div class="section-title">
        🏆 상위 20% 병원 <span class="badge">TOP {top_20_threshold}</span>
      </div>
      <div class="section-desc">"언니가이드 서비스 내 순위" 섹션 노출 · 긍정 시그널 중심</div>
    </div>
    <div class="grid">{top_cards}</div>
  </div>

  <!-- 그 외 -->
  <div class="section">
    <div class="section-header">
      <div class="section-title">
        🏥 그 외 병원 <span class="badge">{total - top_20_threshold}개</span>
      </div>
      <div class="section-desc">성장 기회 · 1위 병원 비교 중심 · 액션 인사이트 강조</div>
    </div>
    <div class="grid">{rest_cards}</div>
  </div>

</div>

<script>
function filterHospitals(query) {{
  const q = query.toLowerCase().trim();
  document.querySelectorAll('.hospital-card').forEach(card => {{
    const name = card.querySelector('.h-name').textContent.toLowerCase();
    card.classList.toggle('hidden', q && !name.includes(q));
  }});
}}
</script>

<div class="footer">
  <strong>UNNI GUIDE</strong> · Index | 생성일: {datetime.now().strftime('%Y.%m.%d')}
  <br>본 페이지는 내부 자료입니다.
</div>

</body></html>"""

out = os.path.join(OUTPUT_DIR, f'index.html')
with open(out, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"✅ Index 생성: {out}")
print(f"   상위 20% ({top_20_threshold}개) 매출 비중: {top_share:.1f}%")
