#!/usr/bin/env python3
"""
언니가이드 월간 트렌드 리포트 생성기 v2
- 공통 트렌드 리포트 (전체 병원 공유) → unniguide_report_YYYYMM.html
- 병원별 개별 리포트 → hospitals/병원명_YYYYMM.html (월별 누적 포함)
"""

import pandas as pd
import json
import os
import re
from datetime import datetime
from collections import defaultdict

# ============================================================
# 1. 설정
# ============================================================
import sys
import glob

# Excel 파일: 인자로 받거나, Downloads에서 최신 파일 자동 탐색
if len(sys.argv) > 1:
    EXCEL_PATH = os.path.expanduser(sys.argv[1])
else:
    # "언니가이드 운영 트렌드 데이터_*.xlsx" 중 가장 최신 파일
    pattern = os.path.expanduser('~/Downloads/언니가이드 운영 트렌드 데이터_*.xlsx')
    candidates = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    if candidates:
        EXCEL_PATH = candidates[0]
    else:
        print("❌ Excel 파일을 찾을 수 없습니다.")
        print("   사용법: python3 generate_report.py [엑셀파일경로]")
        sys.exit(1)

OUTPUT_DIR = os.path.expanduser('~/Documents/Unniguide/unniguide-report')
HOSPITAL_DIR = os.path.join(OUTPUT_DIR, 'hospitals')

# REPORT_MONTH: 인자로 받거나, 데이터에서 자동 감지 (아래에서 설정)
REPORT_MONTH = sys.argv[2] if len(sys.argv) > 2 else None  # 나중에 자동 설정
REPORT_MONTH_KR = None  # 나중에 자동 설정

print(f"📂 Excel: {EXCEL_PATH}")

os.makedirs(HOSPITAL_DIR, exist_ok=True)

# ============================================================
# 브랜드 컬러 (BI Guidelines)
# ============================================================
BRAND = {
    'orange': '#FF6A3B',       # MAIN ORANGE
    'orange_dark': '#E8551F',  # darker for hover/active
    'orange_light': '#FFF0EB', # tinted bg
    'ivory': '#FBF9F1',       # IVORY
    'plum': '#330C2E',        # PLUM
    'plum_light': '#5C2D54',  # lighter plum for text
}

# 병원 마스터 리스트
HOSPITAL_MASTER = {
    1860: {'name': '우리성형외과의원', 'manager': 'Yun', 'country': '태국 제외, 중국 제외', 'area': '강남', 'type': '시수술'},
    3234: {'name': '사적인아름다움지유의원', 'manager': 'Yun', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    660:  {'name': '루호성형외과의원', 'manager': 'Yun', 'country': '모든 국가', 'area': '강남', 'type': '수술'},
    46:   {'name': '플라덴성형외과의원', 'manager': 'Yun', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    5857: {'name': '우유빛의원', 'manager': 'Lyn', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    7493: {'name': '테이아의원', 'manager': 'Yun', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    116:  {'name': '티에스성형외과', 'manager': 'Hardy', 'country': '모든 국가', 'area': '강남', 'type': '수술'},
    2067: {'name': '아크로한의원', 'manager': 'Runa', 'country': '모든 국가', 'area': '강남', 'type': '한방'},
    6890: {'name': '톡스앤필의원-신논현점', 'manager': 'Runa', 'country': '중국 제외', 'area': '강남', 'type': '시술'},
    7826: {'name': '제이필의원-홍대점', 'manager': 'Su', 'country': '모든 국가', 'area': '홍대', 'type': '시술'},
    2681: {'name': '메이필의원', 'manager': 'Runa', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    580:  {'name': '허쉬성형외과의원', 'manager': 'Hardy', 'country': '모든 국가', 'area': '강남', 'type': '수술'},
    3090: {'name': '플래너성형외과의원', 'manager': 'Winnie', 'country': '모든 국가', 'area': '강남', 'type': '수술'},
    5524: {'name': '오앤의원', 'manager': 'Runa', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    6681: {'name': '홍대셀레나의원', 'manager': 'Su', 'country': '모든 국가', 'area': '홍대', 'type': '시술'},
    6518: {'name': '네스트의원', 'manager': 'Runa', 'country': '중국 제외, 대만 제외', 'area': '강남', 'type': '시술'},
    1889: {'name': '제이필의원-강남점', 'manager': 'Su', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    2991: {'name': '플레저성형외과의원', 'manager': 'Hardy', 'country': '모든 국가', 'area': '강남', 'type': '수술'},
    3505: {'name': '히트성형외과의원', 'manager': 'Runa', 'country': '모든 국가', 'area': '강남', 'type': '수술'},
    802:  {'name': '모즈의원', 'manager': 'Runa', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    1111: {'name': '우아성형외과의원', 'manager': '병원마케팅팀', 'country': '태국만', 'area': '강남', 'type': '수술'},
    3608: {'name': '청담디어의원', 'manager': 'Belle', 'country': '대만 제외, 중국 제외', 'area': '강남', 'type': '시술'},
    5826: {'name': '유픽의원-홍대점', 'manager': 'Yun', 'country': '모든 국가', 'area': '홍대', 'type': '시술'},
    5663: {'name': '유픽의원-강남점', 'manager': 'Runa', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    392:  {'name': '라인앤뷰의원', 'manager': 'Runa', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    4575: {'name': '세가지소원의원-명동점', 'manager': 'Runa', 'country': '모든 국가', 'area': '명동', 'type': '시술'},
    4213: {'name': '올리팅성형외과의원', 'manager': 'Runa', 'country': '중국 제외, 대만 제외', 'area': '강남', 'type': '수술'},
    5451: {'name': '강남셀리팅의원', 'manager': 'Lyn', 'country': '모든 국가', 'area': '강남', 'type': '수술'},
    4414: {'name': '리디아여성의원', 'manager': 'Hardy', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    5192: {'name': '플로리아의원', 'manager': 'Yun', 'country': '모든 국가', 'area': '부산', 'type': '시술'},
    3673: {'name': '브이에스라인의원-압구정점', 'manager': 'Hardy', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    6010: {'name': '엔디어트의원', 'manager': 'Hardy', 'country': '모든 국가', 'area': '강남', 'type': '시술'},
    4848: {'name': '메종드엠의원', 'manager': 'Lyn', 'country': '태국만', 'area': '강남', 'type': '시술'},
    9999: {'name': '릴리브의원', 'manager': '병원마케팅팀', 'country': '대만 제외', 'area': '강남', 'type': '시술'},
}

NAME_NORMALIZE = {
    '사적인아름다운지유의원': '사적인아름다움지유의원',
    '루호성형외과': '루호성형외과의원',
    '우리성형외과': '우리성형외과의원',
    '테이아 의원': '테이아의원',
    '티에스성형외과의원': '티에스성형외과',
    '톡스앤필-시논현': '톡스앤필의원-신논현점',
    '톡스앤필 - 신논': '톡스앤필의원-신논현점',
    '톡스앤필 - 신논현': '톡스앤필의원-신논현점',
    '제이필 - 홍대': '제이필의원-홍대점',
    '제이필 - 강남': '제이필의원-강남점',
    '플래저성형외과': '플레저성형외과의원',
    '플래너성형외과': '플래너성형외과의원',
    '유픽의원 홍대': '유픽의원-홍대점',
    '유픽의원-홍대': '유픽의원-홍대점',
    '유픽의원-강남': '유픽의원-강남점',
    '디어 청담의원': '청담디어의원',
    '홍대셀레나': '홍대셀레나의원',
    '히트성형외과': '히트성형외과의원',
}

def normalize_hospital(name):
    if not name or str(name).strip() == '':
        return None
    name = str(name).strip()
    return NAME_NORMALIZE.get(name, name)

COUNTRY_FLAG = {
    '태국': '🇹🇭', '대만': '🇹🇼', '중국': '🇨🇳', '미국': '🇺🇸',
    '호주': '🇦🇺', '일본': '🇯🇵', '홍콩': '🇭🇰', '싱가포르': '🇸🇬',
    '베트남': '🇻🇳', '필리핀': '🇵🇭', '말레이시아': '🇲🇾', '인도네시아': '🇮🇩',
    '영국': '🇬🇧', '프랑스': '🇫🇷', '독일': '🇩🇪', '캐나다': '🇨🇦',
    '인도': '🇮🇳', '러시아': '🇷🇺', '스페인': '🇪🇸', '포르투갈': '🇵🇹',
    '뉴질랜드': '🇳🇿', '노르웨이': '🇳🇴', '폴란드': '🇵🇱', '칠레': '🇨🇱',
    '캄보디아': '🇰🇭', '몽골': '🇲🇳', '사우디 아라비아': '🇸🇦',
    '우즈베키스탄': '🇺🇿', '키르기스스탄': '🇰🇬', '키르기스탄': '🇰🇬',
    '도미니카 공화국': '🇩🇴', '아일랜드': '🇮🇪', '로마니아': '🇷🇴', '부탄': '🇧🇹',
}

def format_krw(amount):
    if amount >= 100000000:
        return f"{amount/100000000:.1f}억원"
    elif amount >= 10000:
        return f"{amount/10000:,.0f}만원"
    else:
        return f"{amount:,.0f}원"

def format_number(n):
    return f"{n:,}"

def safe_filename(name):
    return re.sub(r'[^\w가-힣\-]', '_', name)

# ============================================================
# 시술명 표준화 & 분할 로직
# ============================================================
# 표준 시술 카테고리 + 매칭 키워드 (소문자 비교)
PROCEDURE_KEYWORDS = [
    ('울쎄라', ['울쎄라', '울쎼라', 'ulthera', 'ulthera prime', '울쎄라피', '울쎄라 프라임']),
    ('써마지', ['써마지', 'thermage', 'flx', '써마지flx']),
    ('슈링크', ['슈링크', 'shrink', '유니버스']),
    ('소프웨이브', ['소프웨이브', 'sofwave']),
    ('포텐자', ['포텐자', 'potenza']),
    ('올리지오', ['올리지오', 'oligio']),
    ('온다', ['온다', 'onda']),
    ('인모드', ['인모드', 'inmode']),
    ('덴서티', ['덴서티', 'density']),
    ('티타늄', ['티타늄']),
    ('볼뉴머', ['볼뉴머']),
    ('리프팅레이저', ['리프팅 레이저', '리프팅레이저']),
    ('BBL레이저', ['bbl', 'b.b.l']),
    ('피코레이저', ['피코 레이저', '피코레이저', '피코 토닝', '피코토닝', '피코프락셀', '피코 프락셀']),
    ('토닝레이저', ['토닝']),
    ('트리플레이저', ['트리플 레이저', 'triple laser']),
    ('제네시스', ['제네시스', 'genesis']),
    ('라라필', ['라라필']),
    ('아쿠아필', ['아쿠아필', 'aquapeel', '아쿠아 필']),
    ('LDM', ['ldm']),
    ('CO2레이저', ['co2', 'co₂']),
    ('보톡스', ['보톡스', 'botox', '보톡', '코어톡스', '더마톡신', '엘러간', '제오민', '주름보톡스', '이마 보톡스', '미간 보톡스', '눈가 보톡스', '턱끝 보톡스', '사각턱', '승모근 보톡스', '종아리 보톡스']),
    ('스킨보톡스', ['스킨보톡스', '스킨 보톡스', '코어톡스 스킨', '스킨바이브']),
    ('필러', ['필러', 'filler', '팔자 필러', '팔자필러', '입술필러', '입술 필러', '턱끝 필러', '턱끝필러', '이마 필러', '이마필러', '아띠에르', '잼버실', '스컬트라', '레디어스', '쥬베더', 'juvederm', 'ha 필러', 'hyaluronic acid', 'hyaluronic']),
    ('리쥬란', ['리쥬란', 'rejuran']),
    ('쥬베룩', ['쥬베룩', 'juvelook']),
    ('물광주사', ['물광', '더마샤인', 'skinbooster', '스킨부스터']),
    ('엑소좀', ['엑소좀', 'exosome']),
    ('셀르디엠', ['셀르디엠', 'cellrdm', '셀르']),
    ('모델링', ['모델링', '특수필']),
    ('아그네스', ['아그네스']),
    ('아크로웨이브', ['아크로웨이브', '아크로']),
    ('메타뷰', ['메타뷰', 'sylfirm']),
    ('실리프팅', ['실리프팅', '실 리프팅', '코그', '실리어']),
    ('리쥬비놀', ['리쥬비놀']),
    ('지방분해주사', ['지방분해', '지방분해주사', '윤곽주사', '커팅주사', '팻큐', 'fat dissolv']),
    ('수액', ['수액', '신데렐라', '비콤']),
    ('여드름치료', ['여드름', '염증주사']),
    ('모공치료', ['모공']),
    ('색소치료', ['색소']),
    ('눈밑치료', ['눈밑', '다크서클']),
    ('리투오', ['리투오']),
    ('울블랑', ['울블랑']),
    ('스마일리프팅', ['스마일 리프팅', '스마일리프팅']),
    ('눈성형', ['쌍꺼풀', '상안검', '하안검', '눈매교정', '눈 재수술', '눈 수술', '눈성형', '눈밑 지방', '안검']),
    ('코성형', ['코수술', '코 수술', '코끝', '콧대', '코 재수술', '짧은 코', '짧은코', '연부조직 제거']),
    ('안면거상', ['거상', '안면거상', '목거상', '미니거상']),
    ('지방이식', ['지방이식']),
    ('양악수술', ['양악']),
    ('윤곽수술', ['윤곽수술', '사각턱 수술']),
    ('가슴성형', ['가슴', '유방']),
    ('리프팅시술', ['리프팅']),  # 기타 리프팅 - 마지막
]

# 시술로 카운트하지 않을 패턴 (상담, 결제, 일정 등)
EXCLUDE_PATTERNS = [
    '상담', '결제', '예약금', '귀국', '일정변경', '일정 변경', '한국일정', '취소',
    '방문', '예약', '문의', '확인', '컨펌', 'confirm',
]


def split_procedures(text):
    """자유 텍스트에서 시술명을 분할 + 매칭. (matched_keywords, unmatched_pieces) 반환"""
    if not text or str(text).strip().lower() in ('nan', 'none', ''):
        return [], []
    text = str(text).strip()

    # 먼저 전체 텍스트가 제외 패턴이면 스킵
    if any(excl in text and len(text) < 30 for excl in EXCLUDE_PATTERNS):
        return [], []

    # 구분자로 분할
    pieces = re.split(r'[,+·&/;／、]| 및 | 와 | 과 | 그리고 ', text)

    matched = set()
    unmatched = []

    for piece in pieces:
        piece_clean = piece.strip()
        if not piece_clean or len(piece_clean) < 2:
            continue

        piece_lower = piece_clean.lower()

        # 제외 패턴 체크
        if any(excl in piece_clean for excl in EXCLUDE_PATTERNS):
            continue

        # 키워드 매칭
        hit = None
        for category, keywords in PROCEDURE_KEYWORDS:
            for kw in keywords:
                if kw.lower() in piece_lower:
                    hit = category
                    break
            if hit:
                break

        if hit:
            matched.add(hit)
        else:
            # 숫자/단위/주 주석 등은 노이즈로 처리
            if re.fullmatch(r'[\d\s\-.,()×xX샷shot회cc만원]+', piece_clean):
                continue
            if len(piece_clean) > 60:  # 너무 긴 설명은 무시
                continue
            unmatched.append(piece_clean)

    return list(matched), unmatched


# 매칭 안된 시술명 수집용 (전역)
UNMATCHED_PROCEDURES = []

# ============================================================
# 2. 데이터 읽기
# ============================================================
print("📊 데이터 로딩 중...")

df_res = pd.read_excel(EXCEL_PATH, sheet_name='언니가이드 예약확정 시트 데일리', header=1)
df_res.columns = [
    'NO', '채팅접수일자', '예약확정일', '담당자', '고객명', '그룹여부',
    '고객국적', '사용언어', '예약상태', '통역서비스요청', '종류',
    '시술수술명', '추천클리닉', '예약클리닉', '내원일', '시간',
    '예상금액', '실제금액', '금액확인', '설문발송여부',
    '후기작성여부', '캐시백지급대상자', '캐시백지급여부', '캐시백금액',
    '캐시백지급일자', 'Remark', '시술수술확정항목'
] + [f'extra_{i}' for i in range(max(0, len(df_res.columns) - 27))]

df_res['병원명'] = df_res['예약클리닉'].apply(normalize_hospital)
df_res['내원일'] = pd.to_datetime(df_res['내원일'], errors='coerce')
df_res['월'] = df_res['내원일'].dt.to_period('M').astype(str)
df_res['실제금액'] = pd.to_numeric(df_res['실제금액'], errors='coerce').fillna(0)
df_completed = df_res[df_res['예약상태'] == '시/수술 완료'].copy()

# 정산 데이터
import openpyxl
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws2 = wb['병원별 정산 및 언니가이드 매출 데이터']

settlement_records = []
current_month = None
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=False):
    vals = [c.value for c in row]
    a_val = str(vals[0]).strip() if vals[0] else ''
    if '정산 내역' in a_val:
        parts = a_val.replace('년', '-').replace('월', '').replace('정산 내역', '').strip()
        try:
            year, month = parts.split('-')[:2]
            current_month = f"{year.strip()}-{int(month.strip()):02d}"
        except:
            pass
        continue
    if a_val in ('NO', '', 'None', '재무팀 정산 요청 내역') or '정산 요청일' in a_val:
        continue
    if current_month and vals[1]:
        hospital = str(vals[1]).strip()
        if hospital in ('병원명', ''):
            continue
        try:
            settlement_records.append({
                '정산월': current_month,
                '병원명': normalize_hospital(hospital),
                '고객명': str(vals[2]).strip() if vals[2] else '',
                '국적': str(vals[4]).strip() if vals[4] else '',
                '구분': str(vals[6]).strip() if vals[6] else '',
                '시술금액': float(vals[7]) if vals[7] else 0,
                '수수료금액': float(vals[8]) if vals[8] else 0,
            })
        except (ValueError, TypeError):
            pass
wb.close()

df_settle = pd.DataFrame(settlement_records)
print(f"  예약 데이터: {len(df_res)}건 (완료 {len(df_completed)}건)")
print(f"  정산 데이터: {len(df_settle)}건")

# ============================================================
# 자동 월 감지
# ============================================================
all_months_raw = sorted(df_completed['월'].dropna().unique())

if REPORT_MONTH is None:
    # 정산 데이터의 최신 월 또는 예약 데이터의 최신 월
    settle_months = sorted(df_settle['정산월'].unique()) if len(df_settle) else []
    if settle_months:
        REPORT_MONTH = settle_months[-1]
    elif all_months_raw:
        REPORT_MONTH = str(all_months_raw[-1])
    else:
        print("❌ 데이터에서 월 정보를 찾을 수 없습니다.")
        sys.exit(1)

# 전월 자동 계산
from dateutil.relativedelta import relativedelta
_rm = datetime.strptime(REPORT_MONTH + '-01', '%Y-%m-%d')
PREV_MONTH = (_rm - relativedelta(months=1)).strftime('%Y-%m')

# 한글 월 표시
REPORT_MONTH_KR = f"{_rm.year}년 {_rm.month}월"

print(f"\n📅 리포트 기준월: {REPORT_MONTH_KR} (전월: {PREV_MONTH})")

# ============================================================
# 3. 지표 산출
# ============================================================
print("📈 지표 산출 중...")

all_months = sorted(df_completed['월'].dropna().unique())
all_canonical = set(h['name'] for h in HOSPITAL_MASTER.values())

# 3a. 플랫폼 월별 지표
mar_completed = df_completed[df_completed['월'] == REPORT_MONTH]
feb_completed = df_completed[df_completed['월'] == PREV_MONTH]
mar_all = df_res[df_res['월'] == REPORT_MONTH]

platform = {
    'month': REPORT_MONTH_KR,
    'completed': len(mar_completed),
    'total_revenue': mar_completed['실제금액'].sum(),
    'avg_price': mar_completed['실제금액'].mean() if len(mar_completed) else 0,
    'num_countries': mar_completed['고객국적'].nunique(),
    'num_hospitals': mar_completed['병원명'].nunique(),
    'prev_completed': len(feb_completed),
    'prev_revenue': feb_completed['실제금액'].sum(),
}
platform['completed_growth'] = round((platform['completed'] - platform['prev_completed']) / max(platform['prev_completed'], 1) * 100, 1) if platform['prev_completed'] else None
platform['revenue_growth'] = round((platform['total_revenue'] - platform['prev_revenue']) / max(platform['prev_revenue'], 1) * 100, 1) if platform['prev_revenue'] else None

cumulative = {
    'completed': len(df_completed),
    'revenue': df_completed['실제금액'].sum(),
    'countries': df_completed['고객국적'].nunique(),
    'hospitals': df_completed['병원명'].nunique(),
    'date_range': f"{df_completed['내원일'].min().strftime('%Y.%m.%d')} ~ {df_completed['내원일'].max().strftime('%Y.%m.%d')}",
}

# 3b. 국적별 트렌드
nat_group = mar_completed.groupby('고객국적').agg(건수=('고객명','count'), 총금액=('실제금액','sum')).reset_index()
nat_group['비중'] = round(nat_group['건수'] / nat_group['건수'].sum() * 100, 1)
nat_group['객단가'] = (nat_group['총금액'] / nat_group['건수']).round(0)
nat_group = nat_group.sort_values('건수', ascending=False)
nationality_stats = [
    {'country': r['고객국적'], 'flag': COUNTRY_FLAG.get(r['고객국적'],'🌍'),
     'count': int(r['건수']), 'pct': float(r['비중']),
     'revenue': float(r['총금액']), 'avg_price': float(r['객단가'])}
    for _, r in nat_group.iterrows()
]

# 3c. 국적별 선호 시술
top_countries = [n['country'] for n in nationality_stats[:8]]
nationality_procedures = {}
for country in top_countries:
    _c_counter = {}
    for txt in mar_completed[mar_completed['고객국적'] == country]['시술수술명']:
        matched, _ = split_procedures(txt)
        for p in matched:
            _c_counter[p] = _c_counter.get(p, 0) + 1
    sorted_c = sorted(_c_counter.items(), key=lambda x: x[1], reverse=True)[:5]
    nationality_procedures[country] = [{'procedure': p, 'count': c} for p, c in sorted_c]

# 3d. 시술 비중
type_counts = mar_completed['종류'].value_counts()
procedure_type_stats = {str(t): {'count': int(c), 'pct': round(c/type_counts.sum()*100,1)} for t, c in type_counts.items()}

# 3e. TOP 시술
# 플랫폼 TOP 시술 - 키워드 기반 분할 집계
_proc_counter = {}
_proc_revenue = {}
for _, r in mar_completed.iterrows():
    matched, _ = split_procedures(r['시술수술명'])
    for p in matched:
        _proc_counter[p] = _proc_counter.get(p, 0) + 1
        _proc_revenue[p] = _proc_revenue.get(p, 0) + r['실제금액']
top_procedures = sorted(
    [{'procedure': k, 'count': v, 'revenue': _proc_revenue.get(k, 0)} for k, v in _proc_counter.items()],
    key=lambda x: x['count'], reverse=True,
)[:15]

# 3f. 병원별 월별 누적 데이터
hospital_monthly = defaultdict(lambda: defaultdict(lambda: {'completed':0, 'revenue':0, 'nationalities':{}, 'procedures':{}}))

for _, row in df_completed.iterrows():
    h = row['병원명']
    m = row['월']
    if h not in all_canonical or pd.isna(m):
        continue
    d = hospital_monthly[h][m]
    d['completed'] += 1
    d['revenue'] += row['실제금액']
    nat = row['고객국적']
    if nat:
        if nat not in d['nationalities']:
            d['nationalities'][nat] = {'count':0, 'revenue':0}
        d['nationalities'][nat]['count'] += 1
        d['nationalities'][nat]['revenue'] += row['실제금액']
    proc = row['시술수술명']
    if proc:
        matched, unmatched = split_procedures(proc)
        for p in matched:
            d['procedures'][p] = d['procedures'].get(p, 0) + 1
        # 매칭 안 된 것 수집 (병원명 + 원문과 함께)
        for u in unmatched:
            UNMATCHED_PROCEDURES.append({'병원': h, '월': m, '원문_일부': u, '원문_전체': proc})

# 정산 월별
settle_monthly = defaultdict(lambda: defaultdict(lambda: {'revenue':0, 'commission':0, 'by_nat':{}}))
for _, row in df_settle.iterrows():
    h = row['병원명']
    m = row['정산월']
    if not h:
        continue
    d = settle_monthly[h][m]
    d['revenue'] += row['시술금액']
    d['commission'] += row['수수료금액']
    nat = row['국적']
    if nat:
        if nat not in d['by_nat']:
            d['by_nat'][nat] = {'revenue':0, 'count':0}
        d['by_nat'][nat]['revenue'] += row['시술금액']
        d['by_nat'][nat]['count'] += 1

# 벤치마크
active_hospitals = {h: hospital_monthly[h][REPORT_MONTH] for h in all_canonical if hospital_monthly[h][REPORT_MONTH]['completed'] > 0}
ranked_by_rev = sorted(active_hospitals.keys(), key=lambda x: active_hospitals[x]['revenue'], reverse=True)
ranked_by_cnt = sorted(active_hospitals.keys(), key=lambda x: active_hospitals[x]['completed'], reverse=True)
total_active = len(active_hospitals)
platform_avg_completed = sum(v['completed'] for v in active_hospitals.values()) / max(total_active,1)
platform_avg_revenue = sum(v['revenue'] for v in active_hospitals.values()) / max(total_active,1)

rank_info = {}
for i, h in enumerate(ranked_by_cnt):
    rank_info.setdefault(h, {})['rank_count'] = i+1
    rank_info[h]['pct_count'] = round((1 - i/max(total_active,1))*100)
for i, h in enumerate(ranked_by_rev):
    rank_info.setdefault(h, {})['rank_revenue'] = i+1
    rank_info[h]['pct_revenue'] = round((1 - i/max(total_active,1))*100)

print(f"  3월 완료: {platform['completed']}건, 매출: {format_krw(platform['total_revenue'])}")
print(f"  활성 병원: {total_active}개, 월: {len(all_months)}개")

# ============================================================
# 4. CSS (공통)
# ============================================================
CSS = """
:root {
  --primary: """ + BRAND['orange'] + """;
  --primary-dark: """ + BRAND['orange_dark'] + """;
  --primary-light: """ + BRAND['orange_light'] + """;
  --plum: """ + BRAND['plum'] + """;
  --plum-light: """ + BRAND['plum_light'] + """;
  --ivory: """ + BRAND['ivory'] + """;
  --accent: #00B894;
  --accent-light: #E8FBF5;
  --bg: #FAFBFC;
  --card: #FFFFFF;
  --text: #2D3436;
  --text-light: #636E72;
  --border: #E9ECEF;
  --shadow: 0 2px 12px rgba(0,0,0,0.06);
  --radius: 16px;
}
* { margin:0; padding:0; box-sizing:border-box; }
body {
  font-family: -apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,'Noto Sans KR',sans-serif;
  background: var(--bg); color: var(--text); line-height:1.6;
}
.container { max-width:1100px; margin:0 auto; padding:24px 20px; }

/* Header */
.header {
  background: linear-gradient(135deg, var(--primary) 0%, var(--primary-dark) 100%);
  color: white; padding:40px 0 32px; margin-bottom:32px;
  border-radius: 0 0 32px 32px;
}
.header .container { padding-top:0; padding-bottom:0; }
.header-top { display:flex; justify-content:space-between; align-items:center; margin-bottom:8px; }
.logo { font-size:16px; font-weight:800; letter-spacing:1px; opacity:0.95; }
.logo span { font-weight:400; }
.report-date { font-size:13px; opacity:0.8; }
.header h1 { font-size:26px; font-weight:800; margin-bottom:4px; }
.header .subtitle { font-size:14px; opacity:0.85; }

/* Section */
.section { margin-bottom:32px; }
.section-title {
  font-size:19px; font-weight:700; margin-bottom:16px;
  display:flex; align-items:center; gap:8px; color:var(--plum);
}
.section-title .icon { font-size:22px; }

/* Cards */
.card {
  background:var(--card); border-radius:var(--radius);
  box-shadow:var(--shadow); padding:24px; margin-bottom:16px;
  border:1px solid var(--border);
}

.metric-grid { display:grid; grid-template-columns:repeat(auto-fit,minmax(200px,1fr)); gap:16px; }
.metric-card {
  background:var(--card); border-radius:12px; padding:20px;
  border:1px solid var(--border); text-align:center;
}
.metric-card .label { font-size:13px; color:var(--text-light); margin-bottom:4px; }
.metric-card .value { font-size:26px; font-weight:800; color:var(--plum); }
.metric-card .sub { font-size:12px; color:var(--text-light); margin-top:4px; }

/* Badges */
.badge { display:inline-block; padding:2px 10px; border-radius:12px; font-size:12px; font-weight:600; }
.badge-up { background:#E8FBF5; color:#00B894; }
.badge-down { background:#FFE8E8; color:#E74C3C; }
.badge-neutral { background:#F1F2F6; color:#636E72; }
.badge-primary { background:var(--primary-light); color:var(--primary-dark); }

/* Tables */
.data-table { width:100%; border-collapse:collapse; font-size:14px; }
.data-table th {
  background:var(--ivory); padding:12px 16px; text-align:left;
  font-weight:600; color:var(--plum); font-size:12px;
  text-transform:uppercase; letter-spacing:0.5px;
  border-bottom:2px solid var(--border);
}
.data-table td { padding:12px 16px; border-bottom:1px solid var(--border); }
.data-table tr:hover { background:var(--ivory); }
.data-table .num { text-align:right; font-variant-numeric:tabular-nums; }
.data-table .bold { font-weight:700; }

/* Charts */
.chart-container { position:relative; height:320px; margin:16px 0; }
.chart-half { display:grid; grid-template-columns:1fr 1fr; gap:24px; }
.chart-half .chart-container { height:280px; }

/* Insight */
.insight-box {
  background:linear-gradient(135deg,#FFF9E6 0%,#FFF3CC 100%);
  border-left:4px solid #F39C12; border-radius:0 12px 12px 0;
  padding:16px 20px; margin:16px 0; font-size:14px; line-height:1.7;
}
.opportunity-box {
  background:linear-gradient(135deg,var(--accent-light) 0%,#D5F5ED 100%);
  border-left:4px solid var(--accent); border-radius:0 12px 12px 0;
  padding:16px 20px; margin:16px 0;
}
.opportunity-box h4 { color:#00896A; margin-bottom:8px; }
.opportunity-box ul { margin-left:20px; }
.opportunity-box li { margin-bottom:4px; font-size:14px; }

/* Benchmark */
.benchmark-bar { height:8px; background:#E9ECEF; border-radius:4px; margin:8px 0; overflow:hidden; }
.benchmark-bar .fill {
  height:100%; border-radius:4px;
  background:linear-gradient(90deg,var(--primary) 0%,var(--primary-dark) 100%);
  transition:width 0.8s ease;
}
.benchmark-label { display:flex; justify-content:space-between; font-size:13px; color:var(--text-light); }

/* FAQ */
.faq-item { margin-bottom:16px; }
.faq-item h4 { font-size:14px; color:var(--primary-dark); margin-bottom:4px; }
.faq-item p { font-size:14px; color:var(--text-light); }

/* Nav */
.back-link {
  display:inline-flex; align-items:center; gap:6px;
  color:var(--primary); font-weight:600; text-decoration:none;
  margin-bottom:16px; font-size:14px;
}
.back-link:hover { text-decoration:underline; }

/* Monthly trend table */
.trend-table th, .trend-table td { text-align:center; padding:10px 12px; }
.trend-table th { background:var(--ivory); color:var(--plum); }
.trend-table td.highlight { background:var(--primary-light); font-weight:700; }

/* Footer */
.footer {
  text-align:center; padding:32px 0; color:var(--text-light);
  font-size:13px; border-top:1px solid var(--border); margin-top:40px;
}

/* Hospital list */
.hospital-grid { display:grid; grid-template-columns:repeat(auto-fill,minmax(280px,1fr)); gap:12px; }
.hospital-link {
  display:block; padding:16px 20px; background:var(--card);
  border:1px solid var(--border); border-radius:12px;
  text-decoration:none; color:var(--text); transition:all 0.2s;
}
.hospital-link:hover { border-color:var(--primary); box-shadow:0 4px 16px rgba(255,106,59,0.12); transform:translateY(-1px); }
.hospital-link .h-name { font-weight:700; font-size:15px; margin-bottom:4px; color:var(--plum); }
.hospital-link .h-stats { font-size:13px; color:var(--text-light); }
.hospital-link .h-stats strong { color:var(--primary); }

/* Sticky Nav */
.top-nav { position:sticky; top:0; z-index:100; background:rgba(255,255,255,0.95); backdrop-filter:blur(10px); border-bottom:1px solid var(--border); padding:10px 0; }
.top-nav-inner { max-width:1100px; margin:0 auto; padding:0 20px; display:flex; gap:12px; align-items:center; flex-wrap:wrap; }
.top-nav a { text-decoration:none; color:var(--text-light); font-size:13px; font-weight:600; padding:6px 12px; border-radius:8px; transition:all 0.15s; }
.top-nav a:hover { background:var(--primary-light); color:var(--primary-dark); }
.top-nav a.active { background:var(--primary); color:white; }

@media (max-width:768px) {
  .header h1 { font-size:20px; }
  .metric-grid { grid-template-columns:repeat(2,1fr); }
  .chart-half { grid-template-columns:1fr; }
  .hospital-grid { grid-template-columns:1fr; }
}
@media print {
  .back-link { display:none; }
  .card { break-inside:avoid; }
}
"""

CHART_COLORS = "['#FF6A3B','#330C2E','#00B894','#FDCB6E','#0984E3','#E17055','#00CEC9','#A29BFE','#FD79A8','#55A3E8','#F39C12','#2ECC71','#E74C3C','#9B59B6','#1ABC9C']"

FOOTER = f"""
<div class="footer">
  <div class="container">
    <p><strong>UNNI GUIDE</strong> | 강남언니 언니가이드 서비스</p>
    <p style="margin-top:4px;">본 리포트는 파트너 병원 전용 자료입니다. 무단 배포를 삼가해 주세요.</p>
    <p style="margin-top:4px;opacity:0.7;">생성일: {datetime.now().strftime('%Y년 %m월 %d일')}</p>
  </div>
</div>
"""

TOP_NAV_HTML = lambda current, hospital_path_prefix='': f"""
<div class="top-nav">
  <div class="top-nav-inner">
    <a href="{hospital_path_prefix}index.html" class="{'active' if current == 'index' else ''}">🏠 Index</a>
    <a href="{hospital_path_prefix}unniguide_report_{REPORT_MONTH.replace('-','')}.html" class="{'active' if current == 'common' else ''}">📈 전체 트렌드</a>
    <a href="{hospital_path_prefix}index.html#hospitals-section" class="{'active' if current == 'hospital' else ''}">🏥 병원별 리포트</a>
  </div>
</div>
"""

HEADER_HTML = lambda title, subtitle: f"""
<div class="header">
  <div class="container">
    <div class="header-top">
      <div class="logo">UNNI <span>GUIDE</span></div>
      <div class="report-date">{REPORT_MONTH_KR} 리포트</div>
    </div>
    <h1>{title}</h1>
    <div class="subtitle">{subtitle}</div>
  </div>
</div>
"""

# ============================================================
# 5. 공통 트렌드 리포트 HTML
# ============================================================
print("\n🎨 공통 트렌드 리포트 생성 중...")

def growth_badge(g):
    if g is None: return '<span class="badge badge-neutral">N/A</span>'
    if g > 0: return f'<span class="badge badge-up">▲ {g}%</span>'
    if g < 0: return f'<span class="badge badge-down">▼ {abs(g)}%</span>'
    return '<span class="badge badge-neutral">→ 0%</span>'

# 국적 테이블 rows
nat_rows = ""
for n in nationality_stats:
    nat_rows += f'<tr><td>{n["flag"]} {n["country"]}</td><td>{n["count"]}건</td><td>{n["pct"]}%</td><td class="num">{format_krw(n["revenue"])}</td><td class="num bold">{format_krw(n["avg_price"])}</td></tr>\n'

# 국적별 시술 sections
nat_proc_html = ""
for country in top_countries:
    flag = COUNTRY_FLAG.get(country, '🌍')
    procs = nationality_procedures.get(country, [])
    if procs:
        rows = "".join(f'<tr><td>{p["procedure"]}</td><td class="num">{p["count"]}건</td></tr>' for p in procs)
        nat_proc_html += f'<div class="card"><h3 style="font-size:16px;margin-bottom:12px;">{flag} {country}</h3><table class="data-table"><thead><tr><th>시술/수술명</th><th class="num">건수</th></tr></thead><tbody>{rows}</tbody></table></div>\n'

# 병원 리스트 (카드형 링크)
hospital_cards = ""
for h in ranked_by_rev:
    d = active_hospitals[h]
    fname = safe_filename(h)
    hospital_cards += f'''<a class="hospital-link" href="hospitals/{fname}_{REPORT_MONTH.replace("-","")}.html">
  <div class="h-name">{h}</div>
  <div class="h-stats">완료 <strong>{d["completed"]}건</strong> · 매출 <strong>{format_krw(d["revenue"])}</strong></div>
</a>\n'''

chart_data = json.dumps({
    'nat_labels': [n['country'] for n in nationality_stats[:10]],
    'nat_counts': [n['count'] for n in nationality_stats[:10]],
    'nat_pcts': [n['pct'] for n in nationality_stats[:10]],
    'nat_prices': [n['avg_price'] for n in nationality_stats[:10]],
    'type_labels': list(procedure_type_stats.keys()),
    'type_counts': [v['count'] for v in procedure_type_stats.values()],
    'proc_names': [p['procedure'][:20] for p in top_procedures[:10]],
    'proc_counts': [p['count'] for p in top_procedures[:10]],
}, ensure_ascii=False)

common_html = f"""<!DOCTYPE html>
<html lang="ko"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>언니가이드 월간 트렌드 리포트 | {REPORT_MONTH_KR}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>{CSS}</style>
</head><body>

{TOP_NAV_HTML('common')}
{HEADER_HTML('언니가이드 월간 트렌드 리포트', '파트너 병원을 위한 시장 인사이트 & 성과 분석')}

<div class="container">

<!-- 전체 요약 -->
<div class="section">
  <div class="section-title"><span class="icon">📊</span> {REPORT_MONTH_KR} 전체 요약</div>
  <div class="metric-grid">
    <div class="metric-card"><div class="label">시/수술 완료</div><div class="value">{platform['completed']}건</div><div class="sub">{growth_badge(platform['completed_growth'])}</div></div>
    <div class="metric-card"><div class="label">총 병원 매출</div><div class="value">{format_krw(platform['total_revenue'])}</div><div class="sub">{growth_badge(platform['revenue_growth'])}</div></div>
    <div class="metric-card"><div class="label">평균 객단가</div><div class="value">{format_krw(platform['avg_price'])}</div><div class="sub">1인 평균</div></div>
    <div class="metric-card"><div class="label">참여 국적</div><div class="value">{platform['num_countries']}개국</div><div class="sub">{platform['num_hospitals']}개 병원</div></div>
  </div>
</div>

<!-- 누적 -->
<div class="section">
  <div class="section-title"><span class="icon">🏆</span> 누적 성과 ({cumulative['date_range']})</div>
  <div class="metric-grid">
    <div class="metric-card"><div class="label">총 완료</div><div class="value">{format_number(cumulative['completed'])}건</div></div>
    <div class="metric-card"><div class="label">총 매출</div><div class="value">{format_krw(cumulative['revenue'])}</div></div>
    <div class="metric-card"><div class="label">참여 국적</div><div class="value">{cumulative['countries']}개국</div></div>
    <div class="metric-card"><div class="label">파트너 병원</div><div class="value">{cumulative['hospitals']}개</div></div>
  </div>
</div>

<!-- 국적별 -->
<div class="section">
  <div class="section-title"><span class="icon">🌍</span> 국적별 고객 데이터</div>
  <div class="card">
    <div class="chart-half">
      <div><h3 style="font-size:15px;margin-bottom:8px;">국적별 예약 비중</h3><div class="chart-container"><canvas id="natPie"></canvas></div></div>
      <div><h3 style="font-size:15px;margin-bottom:8px;">국적별 인당 평균 객단가</h3><div class="chart-container"><canvas id="natBar"></canvas></div></div>
    </div>
  </div>
  <div class="card">
    <table class="data-table">
      <thead><tr><th>국적</th><th>건수</th><th>비중</th><th class="num">총 매출</th><th class="num">인당 객단가</th></tr></thead>
      <tbody>{nat_rows}</tbody>
    </table>
  </div>
</div>

<!-- 국적별 시술 -->
<div class="section">
  <div class="section-title"><span class="icon">🔍</span> 국적별 선호 시술 TOP 5</div>
  {nat_proc_html}
</div>

<!-- 시술 트렌드 -->
<div class="section">
  <div class="section-title"><span class="icon">💉</span> 시술 트렌드</div>
  <div class="card">
    <div class="chart-half">
      <div><h3 style="font-size:15px;margin-bottom:8px;">시술 vs 수술 비중</h3><div class="chart-container"><canvas id="typeChart"></canvas></div></div>
      <div><h3 style="font-size:15px;margin-bottom:8px;">인기 시술 TOP 10</h3><div class="chart-container"><canvas id="procChart"></canvas></div></div>
    </div>
  </div>
</div>

<!-- 병원별 리포트 링크 -->
<div class="section">
  <div class="section-title"><span class="icon">🏥</span> 병원별 상세 리포트</div>
  <div class="hospital-grid">{hospital_cards}</div>
</div>

<!-- 예약 성공 공식 -->
<div class="section">
  <div class="section-title"><span class="icon">💡</span> 예약을 더 많이 받는 5가지 성공 공식</div>
  <p style="font-size:13px;color:var(--text-light);margin-bottom:14px;">언니가이드에서 예약이 활발한 병원들의 공통점을 정리했습니다.</p>
  <div class="hospital-grid">
    <div class="card" style="border-left:4px solid var(--primary);background:var(--primary-light);">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;"><span style="background:var(--primary);color:white;font-weight:800;font-size:13px;width:24px;height:24px;border-radius:50%;display:flex;align-items:center;justify-content:center;">1</span><strong style="color:var(--plum);font-size:14px;">장비 리스트 공유</strong></div>
      <div style="font-size:13px;color:var(--text);">병원에서 보유한 <strong>전체 장비 리스트</strong>를 투명하게 전달</div>
    </div>
    <div class="card" style="border-left:4px solid var(--primary);background:var(--primary-light);">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;"><span style="background:var(--primary);color:white;font-weight:800;font-size:13px;width:24px;height:24px;border-radius:50%;display:flex;align-items:center;justify-content:center;">2</span><strong style="color:var(--plum);font-size:14px;">수가 투명 공개</strong></div>
      <div style="font-size:13px;color:var(--text);">병원 보유 <strong>전체 수가</strong>를 공유하여 상담사가 빠르게 고객 안내 가능</div>
    </div>
    <div class="card" style="border-left:4px solid var(--primary);background:var(--primary-light);">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;"><span style="background:var(--primary);color:white;font-weight:800;font-size:13px;width:24px;height:24px;border-radius:50%;display:flex;align-items:center;justify-content:center;">3</span><strong style="color:var(--plum);font-size:14px;">원내 언어 응대자</strong></div>
      <div style="font-size:13px;color:var(--text);">오프라인 센터는 당일·익일 예약이 많아 <strong>원내 언어 응대자 보유</strong>가 큰 메리트</div>
    </div>
    <div class="card" style="border-left:4px solid var(--primary);background:var(--primary-light);">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;"><span style="background:var(--primary);color:white;font-weight:800;font-size:13px;width:24px;height:24px;border-radius:50%;display:flex;align-items:center;justify-content:center;">4</span><strong style="color:var(--plum);font-size:14px;">1시간 이내 답변</strong></div>
      <div style="font-size:13px;color:var(--text);">고객 예약·수가 확인 요청 인입 시 <strong>1시간 이내 답변</strong></div>
    </div>
    <div class="card" style="border-left:4px solid var(--primary);background:var(--primary-light);">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;"><span style="background:var(--primary);color:white;font-weight:800;font-size:13px;width:24px;height:24px;border-radius:50%;display:flex;align-items:center;justify-content:center;">5</span><strong style="color:var(--plum);font-size:14px;">컨설턴트와 적극 소통</strong></div>
      <div style="font-size:13px;color:var(--text);">언니가이드 컨설턴트와의 <strong>적극적 소통·상호 피드백</strong> 교환</div>
    </div>
  </div>
</div>

</div>
{FOOTER}

<script>
const D = {chart_data};
const C = {CHART_COLORS};
new Chart(document.getElementById('natPie'),{{type:'doughnut',data:{{labels:D.nat_labels,datasets:[{{data:D.nat_counts,backgroundColor:C,borderWidth:2,borderColor:'#fff'}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'right',labels:{{font:{{size:12}}}}}}}}}}}});
new Chart(document.getElementById('natBar'),{{type:'bar',data:{{labels:D.nat_labels,datasets:[{{data:D.nat_prices,backgroundColor:C.map(c=>c+'88'),borderColor:C,borderWidth:1,borderRadius:6}}]}},options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{callback:v=>(v/10000).toFixed(0)+'만'}}}},y:{{grid:{{display:false}}}}}}}}}});
new Chart(document.getElementById('typeChart'),{{type:'doughnut',data:{{labels:D.type_labels,datasets:[{{data:D.type_counts,backgroundColor:['#FF6A3B','#330C2E','#00B894'],borderWidth:2,borderColor:'#fff'}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'bottom'}}}}}}}});
new Chart(document.getElementById('procChart'),{{type:'bar',data:{{labels:D.proc_names,datasets:[{{data:D.proc_counts,backgroundColor:'#FF6A3B88',borderColor:'#FF6A3B',borderWidth:1,borderRadius:6}}]}},options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',plugins:{{legend:{{display:false}}}},scales:{{x:{{grid:{{color:'#F1F2F6'}}}},y:{{grid:{{display:false}},ticks:{{font:{{size:11}}}}}}}}}}}});
</script>
</body></html>"""

common_path = os.path.join(OUTPUT_DIR, f'unniguide_report_{REPORT_MONTH.replace("-","")}.html')
with open(common_path, 'w', encoding='utf-8') as f:
    f.write(common_html)
print(f"  ✅ {common_path}")

# ============================================================
# 6. 병원별 개별 HTML 생성
# ============================================================
print("\n🏥 병원별 개별 리포트 생성 중...")

# 플랫폼 TOP 시술 세트
platform_top_proc_set = set(p['procedure'] for p in top_procedures[:10])

# 1등 병원 (매출 기준)의 인기 시술 TOP 5 추출 - 익명화
top1_hospital = ranked_by_rev[0] if ranked_by_rev else None
top1_procedures = []
if top1_hospital:
    top1_data = hospital_monthly[top1_hospital][REPORT_MONTH]
    top1_proc_sorted = sorted([(k, v) for k, v in top1_data['procedures'].items() if k and str(k).strip().lower() not in ('nan', 'none', '')], key=lambda x: x[1], reverse=True)[:5]
    top1_procedures = [p[0] for p in top1_proc_sorted]

generated = 0
for h_name in all_canonical:
    # 데이터가 있는 병원만
    h_months = hospital_monthly[h_name]
    s_months = settle_monthly[h_name]
    all_h_months = sorted(set(list(h_months.keys()) + list(s_months.keys())))

    if not all_h_months:
        continue

    # 현재 월 데이터
    mar = h_months.get(REPORT_MONTH, {'completed':0,'revenue':0,'nationalities':{},'procedures':{}})
    mar_settle = s_months.get(REPORT_MONTH, {'revenue':0,'commission':0,'by_nat':{}})
    feb = h_months.get(PREV_MONTH, {'completed':0,'revenue':0})

    # 누적 합산
    cum_completed = sum(h_months[m]['completed'] for m in h_months)
    cum_revenue = sum(h_months[m]['revenue'] for m in h_months)
    cum_settle_rev = sum(s_months[m]['revenue'] for m in s_months)
    cum_settle_comm = sum(s_months[m]['commission'] for m in s_months)

    # 성장률
    count_growth = round((mar['completed'] - feb['completed']) / max(feb['completed'],1) * 100, 1) if feb['completed'] else None

    # 벤치마크
    ri = rank_info.get(h_name, {})

    # 국적 차트 데이터
    nat_sorted = sorted(mar['nationalities'].items(), key=lambda x: x[1]['count'], reverse=True)
    # 정산 국적
    settle_nat_sorted = sorted(mar_settle['by_nat'].items(), key=lambda x: x[1]['revenue'], reverse=True)

    # 시술 정렬
    proc_sorted = sorted([(k, v) for k, v in mar['procedures'].items() if k and str(k).strip().lower() not in ('nan', 'none', '')], key=lambda x: x[1], reverse=True)[:8]

    # 성장기회
    hosp_procs = set(mar['procedures'].keys())
    growth_opps = list(platform_top_proc_set - hosp_procs)[:5]

    # --- 월별 추이 테이블 ---
    monthly_data = []
    running_completed = 0
    running_revenue = 0
    for m in sorted(set(list(h_months.keys()) + list(s_months.keys()))):
        hd = h_months.get(m, {'completed':0,'revenue':0})
        sd = s_months.get(m, {'revenue':0,'commission':0})
        running_completed += hd['completed']
        running_revenue += hd['revenue']
        monthly_data.append({
            'month': m,
            'completed': hd['completed'],
            'revenue': hd['revenue'],
            'settle_rev': sd['revenue'],
            'settle_comm': sd['commission'],
            'cum_completed': running_completed,
            'cum_revenue': running_revenue,
        })

    # 월별 추이 테이블 HTML
    monthly_headers = "".join(f'<th>{"" if md["month"] != REPORT_MONTH else ""}{md["month"][-2:]}월</th>' for md in monthly_data)
    monthly_completed_cells = "".join(f'<td class="{"highlight" if md["month"]==REPORT_MONTH else ""}">{md["completed"]}건</td>' for md in monthly_data)
    monthly_settle_cells = "".join(f'<td class="{"highlight" if md["month"]==REPORT_MONTH else ""}">{format_krw(md["settle_rev"])}</td>' for md in monthly_data)
    monthly_cum_cells = "".join(f'<td class="{"highlight" if md["month"]==REPORT_MONTH else ""}">{md["cum_completed"]}건</td>' for md in monthly_data)

    # MoM 성장률 계산
    def mom_cell(curr, prev, is_highlight):
        if prev == 0 or prev is None:
            return f'<td class="{"highlight" if is_highlight else ""}">-</td>'
        pct = (curr - prev) / prev * 100
        color = '#00B894' if pct > 0 else ('#E74C3C' if pct < 0 else '#636E72')
        arrow = '▲' if pct > 0 else ('▼' if pct < 0 else '→')
        return f'<td class="{"highlight" if is_highlight else ""}" style="color:{color};font-weight:600;">{arrow} {abs(pct):.1f}%</td>'

    mom_count_cells = ""
    mom_rev_cells = ""
    for i, md in enumerate(monthly_data):
        is_hl = md['month'] == REPORT_MONTH
        if i == 0:
            mom_count_cells += f'<td class="{"highlight" if is_hl else ""}">-</td>'
            mom_rev_cells += f'<td class="{"highlight" if is_hl else ""}">-</td>'
        else:
            prev = monthly_data[i - 1]
            mom_count_cells += mom_cell(md['completed'], prev['completed'], is_hl)
            mom_rev_cells += mom_cell(md['settle_rev'], prev['settle_rev'], is_hl)

    # 국적 테이블
    nat_table_rows = ""
    for nat, nd in nat_sorted:
        flag = COUNTRY_FLAG.get(nat, '🌍')
        avg_p = nd['revenue'] / nd['count'] if nd['count'] else 0
        nat_table_rows += f'<tr><td>{flag} {nat}</td><td class="num">{nd["count"]}건</td><td class="num">{format_krw(nd["revenue"])}</td><td class="num bold">{format_krw(avg_p)}</td></tr>'

    # 정산 국적 테이블 (+ 인기 시술)
    # 국적별 시술 집계 (예약완료 기반, 이미 split 된 데이터)
    hosp_month_df = mar_completed[mar_completed['병원명'] == h_name]
    nat_proc_map = {}
    for _, r in hosp_month_df.iterrows():
        matched, _ = split_procedures(r['시술수술명'])
        nat = r['고객국적']
        if not nat:
            continue
        if nat not in nat_proc_map:
            nat_proc_map[nat] = {}
        for p in matched:
            nat_proc_map[nat][p] = nat_proc_map[nat].get(p, 0) + 1

    settle_nat_rows = ""
    for nat, sd in settle_nat_sorted:
        flag = COUNTRY_FLAG.get(nat, '🌍')
        avg_p = sd['revenue'] / sd['count'] if sd['count'] else 0
        top_procs_for_nat = sorted(nat_proc_map.get(nat, {}).items(), key=lambda x: x[1], reverse=True)[:3]
        proc_label = ', '.join(p[0] for p in top_procs_for_nat) if top_procs_for_nat else '<span style="color:#999;">-</span>'
        settle_nat_rows += f'<tr><td>{flag} {nat}</td><td class="num">{sd["count"]}건</td><td class="num">{format_krw(sd["revenue"])}</td><td class="num bold">{format_krw(avg_p)}</td><td style="font-size:12px;">{proc_label}</td></tr>'

    # 응대 언어 인사이트 (귀원 국적 기준 TOP 3)
    top_nats_for_lang = [n[0] for n in nat_sorted[:3]]
    language_insight_html = ""
    if top_nats_for_lang:
        # 국적 → 주요 언어 매핑
        lang_map = {
            '태국': '태국어', '대만': '중국어(번체)', '중국': '중국어(간체)', '홍콩': '중국어(광둥)/영어',
            '싱가포르': '영어/중국어', '말레이시아': '영어/말레이어', '인도네시아': '인도네시아어',
            '미국': '영어', '캐나다': '영어', '호주': '영어', '영국': '영어', '아일랜드': '영어',
            '일본': '일본어', '몽골': '몽골어', '베트남': '베트남어', '필리핀': '영어',
            '프랑스': '프랑스어', '독일': '독일어', '러시아': '러시아어', '인도': '영어/힌디',
            '키르기스스탄': '러시아어', '키르기스탄': '러시아어', '캄보디아': '크메르어',
            '폴란드': '폴란드어', '스페인': '스페인어', '뉴질랜드': '영어',
        }
        lang_list = []
        seen = set()
        for nat in top_nats_for_lang:
            l = lang_map.get(nat)
            if l and l not in seen:
                lang_list.append(f'<strong>{l}</strong> ({nat})')
                seen.add(l)
        lang_text = ' · '.join(lang_list) if lang_list else ''
        if lang_text:
            language_insight_html = f'''
<div class="insight-box" style="margin-top:16px;">
  💡 <strong>응대 언어 인사이트:</strong> 귀원 주요 고객 국적 기준, 원내에 {lang_text} 응대 가능 인력이 있다면 상담 전환율과 만족도를 크게 끌어올릴 수 있습니다. 오프라인 센터 데이터상 당일·익일 예약 비중이 높아 실시간 언어 응대가 핵심 성공 요인으로 확인되었습니다.
</div>'''

    # 언니가이드 서비스 전체 국적 트렌드 rows
    platform_nat_rows = ""
    for n in nationality_stats[:10]:
        country = n['country']
        top3 = nationality_procedures.get(country, [])[:3]
        proc_names = ', '.join(p['procedure'] for p in top3) if top3 else '<span style="color:#999;">-</span>'
        platform_nat_rows += f'<tr><td>{n["flag"]} {country}</td><td>{n["pct"]}%</td><td class="num bold">{format_krw(n["avg_price"])}</td><td style="font-size:12px;">{proc_names}</td></tr>'
    if len(nationality_stats) > 10:
        platform_nat_rows += f'<tr><td colspan="4" style="text-align:center;color:#999;font-size:12px;">외 {len(nationality_stats)-10}개국</td></tr>'

    # 시술 테이블
    proc_rows = "".join(f'<tr><td>{p[0]}</td><td class="num">{p[1]}건</td></tr>' for p in proc_sorted)

    # 성장기회 HTML
    growth_html = ""
    if growth_opps:
        items = "".join(f'<li>{g}</li>' for g in growth_opps)
        growth_html = f'<div class="opportunity-box"><h4>언니가이드 서비스에서 인기 있지만 귀원에서 아직 예약이 없는 시술</h4><p style="font-size:13px;color:var(--text-light);margin-bottom:8px;">아래 시술로의 고객 유입 확대가 가능합니다.</p><ul>{items}</ul></div>'

    # 1등 병원 vs 귀원 인기 시술 비교 (1등 병원은 익명, 보유 여부 체크)
    top1_compare_html = ""
    if top1_procedures and h_name != top1_hospital:
        # nan, 빈값 제거
        def clean_proc(p):
            return p and str(p).strip().lower() not in ('nan', 'none', '')

        top1_clean = [p for p in top1_procedures if clean_proc(p)][:5]
        # 1위 병원 시술별 건수
        top1_procs_dict = {p[0]: p[1] for p in sorted([(k, v) for k, v in hospital_monthly[top1_hospital][REPORT_MONTH]['procedures'].items() if clean_proc(k)], key=lambda x: x[1], reverse=True)}
        hospital_procs_dict = {p[0]: p[1] for p in proc_sorted if clean_proc(p[0])}

        if top1_clean:
            # 각 시술에 대해 귀원 보유 여부 판정
            compare_rows = ""
            missing_count = 0
            for i, proc in enumerate(top1_clean, 1):
                top1_cnt = top1_procs_dict.get(proc, 0)
                # 키워드 기반 매칭
                owned = False
                owned_count = 0
                for hp, cnt in hospital_procs_dict.items():
                    key = proc.split(',')[0].split('+')[0].split(' ')[0].strip()
                    if key and (key in hp or hp in proc):
                        owned = True
                        owned_count += cnt
                if owned:
                    badge = f'<span style="background:#E8FBF5;color:#00896A;padding:4px 10px;border-radius:12px;font-size:12px;font-weight:600;">✓ 귀원 {owned_count}건</span>'
                    row_bg = ''
                else:
                    badge = '<span style="background:#FFF3E0;color:#E67E22;padding:4px 10px;border-radius:12px;font-size:12px;font-weight:700;">📈 유입 확대 여지</span>'
                    row_bg = 'background:#FFF8F5;'
                    missing_count += 1
                compare_rows += f'<tr style="{row_bg}"><td style="text-align:center;font-weight:700;color:var(--primary);">{i}</td><td style="font-size:14px;font-weight:500;">{proc}</td><td class="num" style="font-size:13px;color:var(--text-light);">{top1_cnt}건</td><td style="text-align:right;">{badge}</td></tr>'

            summary_text = f'<strong style="color:#E67E22;">{missing_count}개 시술</strong>에서 유입 확대 여지 — 추가 마케팅 시 매출 상승 기회' if missing_count > 0 else '<strong style="color:#00896A;">모든 TOP 시술 보유 중</strong>'

            top1_compare_html = f'''
<div class="card" style="background:linear-gradient(135deg,#FFF9E6 0%,#FFFDF5 100%);border:1px solid #F5E6A8;margin-top:12px;">
  <h3 style="font-size:16px;margin-bottom:8px;color:var(--plum);">🏆 매출 1위 병원 TOP 시술 · 귀원 보유 현황</h3>
  <p style="font-size:13px;color:var(--text-light);margin-bottom:16px;">이달 매출 1위 병원(익명)의 인기 TOP 5 시술을 기준으로 귀원 등록 여부를 비교합니다. {summary_text}</p>
  <table class="data-table" style="background:white;">
    <thead>
      <tr>
        <th style="width:50px;text-align:center;">순위</th>
        <th>1위 병원 TOP 시술</th>
        <th class="num" style="width:70px;">1위 건수</th>
        <th style="text-align:right;width:160px;">귀원 현황</th>
      </tr>
    </thead>
    <tbody>{compare_rows}</tbody>
  </table>
</div>'''

    # 차트 데이터
    # MoM 계산 (매출 기준)
    monthly_mom_rev = []
    for i, md in enumerate(monthly_data):
        if i == 0 or monthly_data[i-1]['settle_rev'] == 0:
            monthly_mom_rev.append(None)
        else:
            pct = (md['settle_rev'] - monthly_data[i-1]['settle_rev']) / monthly_data[i-1]['settle_rev'] * 100
            monthly_mom_rev.append(round(pct, 1))

    h_chart = json.dumps({
        'nat_labels': [n[0] for n in nat_sorted],
        'nat_counts': [n[1]['count'] for n in nat_sorted],
        'nat_prices': [round(n[1]['revenue']/n[1]['count']) if n[1]['count'] else 0 for n in nat_sorted],
        'monthly_labels': [md['month'][-2:]+'월' for md in monthly_data],
        'monthly_completed': [md['completed'] for md in monthly_data],
        'monthly_revenue': [md['settle_rev'] for md in monthly_data],
        'monthly_mom': monthly_mom_rev,
    }, ensure_ascii=False)

    # 벤치마크 HTML (상위 20% 병원에만 노출)
    pct_c = ri.get('pct_count', 0)
    pct_r = ri.get('pct_revenue', 0)
    raw_pct_c = max(1, 100 - pct_c + round(100/max(total_active,1)))
    raw_pct_r = max(1, 100 - pct_r + round(100/max(total_active,1)))
    # 10% 단위 내림 (16% → 10%, 22% → 20%)
    top_pct_c = max(10, (raw_pct_c // 10) * 10)
    top_pct_r = max(10, (raw_pct_r // 10) * 10)

    # 예약 건수 또는 매출 기준 중 하나라도 상위 20% 안에 들면 노출
    show_benchmark = (top_pct_c <= 20) or (top_pct_r <= 20)

    if show_benchmark:
        # 상위 20% 안에 든 기준만 카드 노출
        rank_cards = ""
        if top_pct_c <= 20:
            rank_cards += f'''
      <div style="text-align:center;padding:24px;background:var(--ivory);border-radius:12px;">
        <div style="font-size:13px;color:var(--text-light);margin-bottom:10px;">예약 건수 기준</div>
        <div style="font-size:32px;font-weight:800;color:var(--primary);margin-bottom:8px;">상위 {top_pct_c}%</div>
        <div style="font-size:12px;color:var(--text-light);">귀원 {mar['completed']}건</div>
      </div>'''
        if top_pct_r <= 20:
            rank_cards += f'''
      <div style="text-align:center;padding:24px;background:var(--ivory);border-radius:12px;">
        <div style="font-size:13px;color:var(--text-light);margin-bottom:10px;">매출 기준</div>
        <div style="font-size:32px;font-weight:800;color:var(--primary);margin-bottom:8px;">상위 {top_pct_r}%</div>
        <div style="font-size:12px;color:var(--text-light);">귀원 {format_krw(mar['revenue'])}</div>
      </div>'''

        grid_cols = "1fr 1fr" if (top_pct_c <= 20 and top_pct_r <= 20) else "1fr"
        benchmark_html = f'''
<!-- 벤치마크 -->
<div class="section">
  <div class="section-title"><span class="icon">🏆</span> 언니가이드 서비스 내 순위 (TOP 20%)</div>
  <div class="card">
    <p style="font-size:13px;color:var(--text-light);margin-bottom:16px;">귀원은 언니가이드 파트너 병원 중 <strong>상위 20% 이내</strong>의 성과를 기록하고 있습니다.</p>
    <div style="display:grid;grid-template-columns:{grid_cols};gap:20px;">{rank_cards}
    </div>
  </div>
</div>'''
    else:
        benchmark_html = ""

    hosp_html = f"""<!DOCTYPE html>
<html lang="ko"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>{h_name} | 언니가이드 리포트 {REPORT_MONTH_KR}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>{CSS}</style>
</head><body>

{TOP_NAV_HTML('hospital', hospital_path_prefix='../')}
{HEADER_HTML(h_name, f'{REPORT_MONTH_KR} 파트너 병원 성과 리포트')}

<div class="container">

<a class="back-link" href="../unniguide_report_{REPORT_MONTH.replace('-','')}.html">← 전체 트렌드 리포트로 돌아가기</a>

<!-- 당월 성과 -->
<div class="section">
  <div class="section-title"><span class="icon">📊</span> {REPORT_MONTH_KR} 성과</div>
  <div class="metric-grid">
    <div class="metric-card"><div class="label">시/수술 완료</div><div class="value">{mar['completed']}건</div><div class="sub">{growth_badge(count_growth)} 전월 {feb['completed']}건</div></div>
    <div class="metric-card"><div class="label">정산 매출</div><div class="value">{format_krw(mar_settle['revenue'])}</div><div class="sub">수수료 {format_krw(mar_settle['commission'])}</div></div>
    <div class="metric-card"><div class="label">인당 객단가</div><div class="value">{format_krw(mar['revenue']/mar['completed']) if mar['completed'] else '-'}</div><div class="sub">실 결제 기준</div></div>
    <div class="metric-card"><div class="label">누적 완료</div><div class="value">{cum_completed}건</div><div class="sub">총 {format_krw(cum_revenue)}</div></div>
  </div>
</div>

<!-- 월별 추이 -->
<div class="section">
  <div class="section-title"><span class="icon">📈</span> 월별 추이</div>
  <div class="card">
    <div class="chart-container" style="height:260px;"><canvas id="monthlyChart"></canvas></div>
    <table class="data-table trend-table" style="margin-top:16px;">
      <thead><tr><th>구분</th>{monthly_headers}</tr></thead>
      <tbody>
        <tr><td class="bold">완료 건수</td>{monthly_completed_cells}</tr>
        <tr><td style="color:var(--text-light);font-size:12px;">건수 MoM</td>{mom_count_cells}</tr>
        <tr><td class="bold">정산 금액</td>{monthly_settle_cells}</tr>
        <tr><td style="color:var(--text-light);font-size:12px;">매출 MoM</td>{mom_rev_cells}</tr>
      </tbody>
    </table>
  </div>
</div>

{benchmark_html}

<!-- 국적별 -->
<div class="section">
  <div class="section-title"><span class="icon">🌏</span> 귀원 고객 국적 분포</div>
  <div class="card">
    <div class="chart-half">
      <div><h3 style="font-size:15px;margin-bottom:8px;">국적 비중</h3><div class="chart-container"><canvas id="natPie"></canvas></div></div>
      <div><h3 style="font-size:15px;margin-bottom:8px;">국적별 객단가</h3><div class="chart-container"><canvas id="natBar"></canvas></div></div>
    </div>
  </div>
  {"<div class='card'><h3 style='font-size:15px;margin-bottom:12px;'>국적별 정산 내역 · 인기 시술</h3><table class='data-table'><thead><tr><th>국적</th><th class='num'>건수</th><th class='num'>시수술 금액</th><th class='num'>건당 객단가</th><th>인기 시술 TOP 3</th></tr></thead><tbody>" + settle_nat_rows + "</tbody></table></div>" if settle_nat_rows else ""}
  {language_insight_html}
</div>

<!-- 언니가이드 서비스 전체 국적 트렌드 (귀원 참고용) -->
<div class="section">
  <div class="section-title"><span class="icon">🌐</span> 언니가이드 서비스 전체 · {REPORT_MONTH_KR} 국적 트렌드</div>
  <p style="font-size:13px;color:var(--text-light);margin-bottom:12px;">언니가이드 서비스 전체 외국인 고객의 국적별 비중과 객단가 추이입니다. 귀원의 타겟 고객군 발굴에 참고해주세요.</p>
  <div class="card">
    <table class="data-table">
      <thead><tr><th>국적</th><th>비중</th><th class="num">인당 객단가</th><th>선호 시술 TOP 3</th></tr></thead>
      <tbody>{platform_nat_rows}</tbody>
    </table>
  </div>
</div>

<!-- 시술 -->
{"<div class='section'><div class='section-title'><span class='icon'>💉</span> 귀원 인기 시술</div><div class='card'><table class='data-table'><thead><tr><th>시술/수술명</th><th class='num'>건수</th></tr></thead><tbody>" + proc_rows + "</tbody></table></div></div>" if proc_rows else ""}

<!-- 성장기회 + 1등 병원 비교 -->
{"<div class='section'><div class='section-title'><span class='icon'>🚀</span> 성장 기회</div>" + growth_html + top1_compare_html + "</div>" if (growth_html or top1_compare_html) else ""}

<!-- FAQ: 예약 성공 공식 -->
<div class="section">
  <div class="section-title"><span class="icon">💡</span> 예약을 더 많이 받는 5가지 성공 공식</div>
  <p style="font-size:13px;color:var(--text-light);margin-bottom:14px;">언니가이드에서 예약이 활발한 병원들의 공통점을 정리했습니다.</p>
  <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:12px;">
    <div style="padding:18px;background:linear-gradient(135deg,#FFF0EB 0%,#FFF8F5 100%);border-left:4px solid var(--primary);border-radius:10px;">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">
        <span style="background:var(--primary);color:white;font-weight:800;font-size:14px;width:26px;height:26px;border-radius:50%;display:flex;align-items:center;justify-content:center;">1</span>
        <strong style="color:var(--plum);font-size:14px;">장비 리스트 공유</strong>
      </div>
      <div style="font-size:13px;color:var(--text);line-height:1.5;">병원에서 보유한 <strong>전체 장비 리스트</strong>를 투명하게 전달</div>
    </div>
    <div style="padding:18px;background:linear-gradient(135deg,#FFF0EB 0%,#FFF8F5 100%);border-left:4px solid var(--primary);border-radius:10px;">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">
        <span style="background:var(--primary);color:white;font-weight:800;font-size:14px;width:26px;height:26px;border-radius:50%;display:flex;align-items:center;justify-content:center;">2</span>
        <strong style="color:var(--plum);font-size:14px;">수가 투명 공개</strong>
      </div>
      <div style="font-size:13px;color:var(--text);line-height:1.5;">병원 보유 <strong>전체 수가</strong>를 공유하여 상담사가 빠르게 고객 안내 가능</div>
    </div>
    <div style="padding:18px;background:linear-gradient(135deg,#FFF0EB 0%,#FFF8F5 100%);border-left:4px solid var(--primary);border-radius:10px;">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">
        <span style="background:var(--primary);color:white;font-weight:800;font-size:14px;width:26px;height:26px;border-radius:50%;display:flex;align-items:center;justify-content:center;">3</span>
        <strong style="color:var(--plum);font-size:14px;">원내 언어 응대자</strong>
      </div>
      <div style="font-size:13px;color:var(--text);line-height:1.5;">오프라인 센터는 당일·익일 예약이 많아 <strong>원내 언어 응대자 보유</strong>가 큰 메리트</div>
    </div>
    <div style="padding:18px;background:linear-gradient(135deg,#FFF0EB 0%,#FFF8F5 100%);border-left:4px solid var(--primary);border-radius:10px;">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">
        <span style="background:var(--primary);color:white;font-weight:800;font-size:14px;width:26px;height:26px;border-radius:50%;display:flex;align-items:center;justify-content:center;">4</span>
        <strong style="color:var(--plum);font-size:14px;">1시간 이내 답변</strong>
      </div>
      <div style="font-size:13px;color:var(--text);line-height:1.5;">고객 예약·수가 확인 요청 인입 시 <strong>1시간 이내 답변</strong></div>
    </div>
    <div style="padding:18px;background:linear-gradient(135deg,#FFF0EB 0%,#FFF8F5 100%);border-left:4px solid var(--primary);border-radius:10px;">
      <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px;">
        <span style="background:var(--primary);color:white;font-weight:800;font-size:14px;width:26px;height:26px;border-radius:50%;display:flex;align-items:center;justify-content:center;">5</span>
        <strong style="color:var(--plum);font-size:14px;">컨설턴트와 적극 소통</strong>
      </div>
      <div style="font-size:13px;color:var(--text);line-height:1.5;">언니가이드 컨설턴트와의 <strong>적극적 소통·상호 피드백</strong> 교환</div>
    </div>
  </div>
</div>

</div>
{FOOTER}

<script>
const D = {h_chart};
const C = {CHART_COLORS};

// 월별 추이 차트 (정산 매출 바)
new Chart(document.getElementById('monthlyChart'),{{
  type:'bar',
  data:{{
    labels:D.monthly_labels,
    datasets:[
      {{label:'정산 매출',data:D.monthly_revenue,backgroundColor:'#FF6A3B88',borderColor:'#FF6A3B',borderWidth:1,borderRadius:6}}
    ]
  }},
  options:{{
    responsive:true,maintainAspectRatio:false,
    plugins:{{
      legend:{{display:false}},
      tooltip:{{callbacks:{{label:function(ctx){{return '정산 매출: '+(ctx.parsed.y/10000).toLocaleString()+'만원';}}}}}}
    }},
    scales:{{
      y:{{title:{{display:true,text:'정산 매출 (원)'}},grid:{{color:'#F1F2F6'}},ticks:{{callback:v=>(v/10000).toFixed(0)+'만'}}}}
    }}
  }}
}});

// 국적 차트
if(D.nat_labels.length>0){{
  new Chart(document.getElementById('natPie'),{{type:'doughnut',data:{{labels:D.nat_labels,datasets:[{{data:D.nat_counts,backgroundColor:C,borderWidth:2,borderColor:'#fff'}}]}},options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'right',labels:{{font:{{size:12}}}}}}}}}}}});
  // % 표시 플러그인
  Chart.getChart('natPie').options.plugins.tooltip = {{callbacks:{{label:function(ctx){{const total=ctx.dataset.data.reduce((a,b)=>a+b,0);const pct=(ctx.parsed/total*100).toFixed(1);return ctx.label+': '+pct+'%';}}}}}};
  Chart.getChart('natPie').update();
  new Chart(document.getElementById('natBar'),{{type:'bar',data:{{labels:D.nat_labels,datasets:[{{data:D.nat_prices,backgroundColor:C.map(c=>c+'88'),borderColor:C,borderWidth:1,borderRadius:6}}]}},options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{callback:v=>(v/10000).toFixed(0)+'만'}}}},y:{{grid:{{display:false}}}}}}}}}});
}}
</script>
</body></html>"""

    fname = safe_filename(h_name)
    fpath = os.path.join(HOSPITAL_DIR, f'{fname}_{REPORT_MONTH.replace("-","")}.html')
    with open(fpath, 'w', encoding='utf-8') as f:
        f.write(hosp_html)
    generated += 1

print(f"  ✅ 병원별 리포트 {generated}개 생성 완료 → {HOSPITAL_DIR}/")
print(f"\n🎉 전체 완료!")
print(f"  📄 공통 리포트: {common_path}")
print(f"  🏥 병원별 리포트: {HOSPITAL_DIR}/ ({generated}개)")

# ============================================================
# 매칭 안 된 시술명 저장 (확인 필요)
# ============================================================
if UNMATCHED_PROCEDURES:
    unmatched_df = pd.DataFrame(UNMATCHED_PROCEDURES)
    unmatched_counts = unmatched_df['원문_일부'].value_counts().reset_index()
    unmatched_counts.columns = ['미매칭_시술명', '등장횟수']
    unmatched_path = os.path.join(OUTPUT_DIR, f'unmatched_procedures_{REPORT_MONTH.replace("-","")}.csv')
    unmatched_counts.to_csv(unmatched_path, index=False, encoding='utf-8-sig')

    print(f"\n⚠️  매칭 안 된 시술명 {len(unmatched_counts)}개 발견 (총 {len(UNMATCHED_PROCEDURES)}건 등장)")
    print(f"    → {unmatched_path}")
    print(f"\n  TOP 20 미매칭 시술명:")
    for _, r in unmatched_counts.head(20).iterrows():
        print(f"    {r['등장횟수']:3d}회: {r['미매칭_시술명']}")
