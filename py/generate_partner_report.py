#!/usr/bin/env python3
"""
언니가이드 오프라인 센터 + 서비스 실적 제휴사용 원페이지 리포트
기간: 2026.03.03 (정식 오픈) ~ 2026.04.02 (오픈 첫 1개월)
"""
import pandas as pd
import openpyxl
import os, json
from datetime import datetime

CSV_PATH = os.path.expanduser("~/Downloads/언니가이드_운영의 사본 - Offline - Daily (외국인).csv")
EXCEL_PATH = os.path.expanduser("~/Downloads/언니가이드 운영 트렌드 데이터_26.04.xlsx")
OUTPUT_DIR = os.path.expanduser("~/Documents/Unniguide/unniguide-report")

NAME_NORMALIZE = {
    '사적인아름다운지유의원': '사적인아름다움지유의원', '루호성형외과': '루호성형외과의원',
    '테이아 의원': '테이아의원', '톡스앤필 - 신논': '톡스앤필의원-신논현점',
    '톡스앤필 - 신논현': '톡스앤필의원-신논현점', '제이필 - 홍대': '제이필의원-홍대점',
    '제이필 - 강남': '제이필의원-강남점', '유픽의원 홍대': '유픽의원-홍대점',
    '유픽의원-홍대': '유픽의원-홍대점', '유픽의원-강남': '유픽의원-강남점',
    '홍대셀레나': '홍대셀레나의원', '히트성형외과': '히트성형외과의원',
    '플래너성형외과': '플래너성형외과의원', '플래저성형외과': '플레저성형외과의원',
}
COUNTRY_FLAG = {
    '태국':'🇹🇭','대만':'🇹🇼','중국':'🇨🇳','미국':'🇺🇸','호주':'🇦🇺','싱가포르':'🇸🇬',
    '몽골':'🇲🇳','캐나다':'🇨🇦','말레이시아':'🇲🇾','영국':'🇬🇧','홍콩':'🇭🇰',
    '필리핀':'🇵🇭','프랑스':'🇫🇷','독일':'🇩🇪','인도':'🇮🇳','일본':'🇯🇵',
    '폴란드':'🇵🇱','아일랜드':'🇮🇪','키르기스스탄':'🇰🇬',
}

def normalize_hospital(name):
    if not name or str(name).strip() in ('', 'nan', 'None'): return None
    name = str(name).strip()
    return NAME_NORMALIZE.get(name, name)

def format_krw(amount):
    if amount >= 100_000_000: return f"{amount/100_000_000:.1f}억"
    elif amount >= 10_000: return f"{amount/10_000:,.0f}만원"
    else: return f"{amount:,.0f}원"

# ============================================================
# 1. 오프라인 센터 데이터
# ============================================================
df = pd.read_csv(CSV_PATH, encoding='latin1', header=None)

def fix_enc(s):
    if pd.isna(s): return ''
    try: return str(s).encode('latin1').decode('euc-kr')
    except:
        try: return str(s).encode('latin1').decode('cp949')
        except: return str(s)

dates_row = df.iloc[4].tolist()
target_cols = [i for i, d in enumerate(dates_row) if str(d).strip()[:10] >= '2026-03-03' and str(d).strip()[:10] <= '2026-03-31']
num_days = len(target_cols)

def row_sum(row_idx):
    total = 0
    for c in target_cols:
        v = df.iloc[row_idx, c]
        if pd.isna(v) or str(v).strip() in ('-', '', 'nan'): continue
        try: total += float(str(v).replace(',', ''))
        except: pass
    return int(total)

def daily_vals(row_idx):
    return [(str(dates_row[c])[:10], int(float(str(df.iloc[row_idx, c]).replace(',', ''))) if pd.notna(df.iloc[row_idx, c]) and str(df.iloc[row_idx, c]).strip() not in ('-', '', 'nan') else 0) for c in target_cols]

예약고객수 = row_sum(6)
실제방문자 = row_sum(18)
방한_예약 = row_sum(10)
재한_예약 = row_sum(14)
방한_방문 = row_sum(22)  # 방문 하위
재한_방문 = row_sum(26)
영어 = row_sum(7); 태국어 = row_sum(9); 중국어 = row_sum(8)
영어_방문 = row_sum(19); 태국어_방문 = row_sum(21); 중국어_방문 = row_sum(20)
일평균방문 = 실제방문자 / num_days
일평균예약 = 예약고객수 / num_days

daily_b = daily_vals(6)
daily_v = daily_vals(18)
chart_labels = [d[5:] for d, _ in daily_b]
chart_booking = [v for _, v in daily_b]
chart_visit = [v for _, v in daily_v]

weeks_row = df.iloc[3].tolist()
weekly = {}
for idx, c in enumerate(target_cols):
    w = f"W{str(weeks_row[c]).strip()}"
    if w not in weekly: weekly[w] = {'예약': 0, '방문': 0}
    weekly[w]['예약'] += chart_booking[idx]
    weekly[w]['방문'] += chart_visit[idx]

# 누적 (우상향 추이용)
cum_visit = 0
cum_booking = 0
for w in weekly:
    cum_visit += weekly[w]['방문']
    cum_booking += weekly[w]['예약']
    weekly[w]['누적방문'] = cum_visit
    weekly[w]['누적예약'] = cum_booking

# ============================================================
# 2. 정산 데이터 (3월 국적별/객단가)
# ============================================================
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws2 = wb['병원별 정산 및 언니가이드 매출 데이터']
settle_records = []
current_month = None
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=False):
    vals = [c.value for c in row]
    a = str(vals[0]).strip() if vals[0] else ''
    if '정산 내역' in a:
        parts = a.replace('년','-').replace('월','').replace('정산 내역','').strip()
        try:
            y, m = parts.split('-')[:2]
            current_month = f"{y.strip()}-{int(m.strip()):02d}"
        except: pass
        continue
    if current_month == '2026-03' and vals[1]:
        hospital = str(vals[1]).strip()
        if hospital in ('병원명', ''): continue
        try:
            settle_records.append({
                '국적': str(vals[4]).strip() if vals[4] else '',
                '시술금액': float(vals[7]) if vals[7] else 0,
            })
        except: pass
wb.close()

df_settle = pd.DataFrame(settle_records)
nat_stats = df_settle.groupby('국적').agg(건수=('시술금액','count'), 매출=('시술금액','sum')).sort_values('건수', ascending=False).reset_index()
nat_stats['비중'] = (nat_stats['건수'] / nat_stats['건수'].sum() * 100).round(1)
nat_stats['객단가'] = (nat_stats['매출'] / nat_stats['건수']).round(0)
nat_stats['국기'] = nat_stats['국적'].map(COUNTRY_FLAG).fillna('🌍')

total_건수 = nat_stats['건수'].sum()
total_매출 = nat_stats['매출'].sum()
avg_객단가 = total_매출 / total_건수 if total_건수 > 0 else 0
num_국적 = len(nat_stats)

# 국적별 비중 + 객단가 테이블 HTML
nat_rows_with_price_html = ""
for _, r in nat_stats.head(8).iterrows():
    nat_rows_with_price_html += f'<tr><td>{r["국기"]} {r["국적"]}</td><td class="bold">{r["비중"]}%</td><td class="bold">{format_krw(r["객단가"])}</td></tr>\n'
if len(nat_stats) > 8:
    etc_cnt = nat_stats.iloc[8:]['건수'].sum()
    etc_pct = (etc_cnt / total_건수 * 100)
    nat_rows_with_price_html += f'<tr><td>🌍 기타 {len(nat_stats)-8}개국</td><td>{etc_pct:.1f}%</td><td>-</td></tr>\n'

# 주요 4개국 객단가 차트 데이터 (태국, 대만, 미국, 중국)
focus_countries = ['태국', '대만', '미국', '중국']
focus_data = nat_stats[nat_stats['국적'].isin(focus_countries)].sort_values('객단가', ascending=False)
price_labels = [f"{r['국기']} {r['국적']}" for _, r in focus_data.iterrows()]
price_values = [int(r['객단가']) for _, r in focus_data.iterrows()]

# ============================================================
# 3월 시술 카테고리 TOP 5 분석 (온·오프라인 통합)
# ============================================================
import pandas as _pd
df_proc = _pd.read_excel(EXCEL_PATH, sheet_name=0, header=1)
_cols = ['NO','채팅접수일자','예약확정일','담당자','고객명','그룹여부','고객국적','사용언어',
         '예약상태','통역서비스요청','종류','시술수술명','추천클리닉','예약클리닉','내원일',
         '시간','예상금액','실제금액','금액확인','설문발송여부','후기작성여부','캐시백지급대상자',
         '캐시백지급여부','캐시백금액','캐시백지급일자','Remark','시술수술확정항목']
_extra = [f'extra_{i}' for i in range(max(0, len(df_proc.columns) - len(_cols)))]
df_proc.columns = (_cols + _extra)[:len(df_proc.columns)]
df_proc['내원일'] = _pd.to_datetime(df_proc['내원일'], errors='coerce')
df_proc['월'] = df_proc['내원일'].dt.to_period('M').astype(str)
_mar = df_proc[(df_proc['월'] == '2026-03') & (df_proc['예약상태'].astype(str).str.strip() == '시/수술 완료')]
_total_cust = len(_mar)

_CAT = {
    '리프팅': ['울쎄라','ulthera','울쎼라','써마지','thermage','flx','슈링크','shrink','소프웨이브','sofwave',
             '올리지오','oligio','온다','onda','덴서티','티타늄','인모드','inmode','아크로웨이브','볼뉴머',
             '실리프팅','코그','거상','스마일 리프팅','리프팅'],
    '보톡스': ['보톡스','botox','코어톡스','제오민','엘러간','더마톡신','스킨보톡스'],
    '스킨부스터': ['리쥬란','rejuran','쥬베룩','juvelook','물광','더마샤인','스킨부스터','엑소좀','exosome',
                '셀르디엠','스킨바이브','리쥬비놀'],
    '마이크로니들링': ['포텐자','potenza','아그네스','라라필','메타뷰','sylfirm'],
    '레이저': ['bbl','피코','pico','트리플 레이저','triple laser','제네시스','genesis','토닝','co2','색소','레이저','laser'],
    '필링/아쿠아': ['아쿠아필','aquapeel','블랙필','ldm'],
    '필러': ['필러','filler','아띠에르','잼버실','스컬트라','juvederm','hyaluronic','레디어스'],
}

def _match_cats(text):
    if _pd.isna(text): return set()
    t = str(text).lower()
    return {c for c, kws in _CAT.items() if any(k.lower() in t for k in kws)}

_cat_cnt = {c: 0 for c in _CAT}
for _, r in _mar.iterrows():
    for c in _match_cats(r['시술수술명']):
        _cat_cnt[c] += 1

cat_rank = sorted([(c, v) for c, v in _cat_cnt.items()], key=lambda x: x[1], reverse=True)[:5]
cat_labels = [c for c, _ in cat_rank]
cat_values = [v for _, v in cat_rank]
cat_percentages = [round(v / _total_cust * 100, 1) for v in cat_values]

# TOP 5 카테고리 HTML rows
_emoji = {'리프팅': '✨', '보톡스': '💉', '스킨부스터': '💧', '마이크로니들링': '🔬',
          '레이저': '⚡', '필링/아쿠아': '🌊', '필러': '🎯'}
cat_rows_html = ""
for i, (c, v) in enumerate(cat_rank, 1):
    pct = v / _total_cust * 100
    bar_width = min(100, pct * 2)  # 50% = 100% bar
    emoji = _emoji.get(c, '💎')
    cat_rows_html += f'''
    <div style="margin-bottom:10px;">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px;">
        <div style="font-size:13px;font-weight:700;color:#330C2E;">{i}. {emoji} {c}</div>
        <div style="font-size:12px;color:#636E72;"><strong style="color:#FF6A3B;">{pct:.1f}%</strong></div>
      </div>
      <div style="height:8px;background:#F1F2F6;border-radius:4px;overflow:hidden;">
        <div style="height:100%;width:{bar_width}%;background:linear-gradient(90deg,#FF6A3B,#E8551F);border-radius:4px;"></div>
      </div>
    </div>'''

chart_data = json.dumps({
    'labels': chart_labels, 'booking': chart_booking, 'visit': chart_visit,
    'wk_labels': list(weekly.keys())[:-1],
    'wk_visit': [v['방문'] for v in weekly.values()][:-1],
    'nat_labels': [r['국적'] for _, r in nat_stats.head(6).iterrows()],
    'nat_counts': [int(r['건수']) for _, r in nat_stats.head(6).iterrows()],
    'price_labels': price_labels,
    'price_values': price_values,
}, ensure_ascii=False)

# ============================================================
# HTML
# ============================================================
html = f"""<!DOCTYPE html>
<html lang="ko"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>UNNI GUIDE 센터 리포트 | 오픈 첫 1개월</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
@page {{ size:A4; margin:10mm; }}
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans KR',sans-serif;color:#2D3436;line-height:1.5;background:#fff;font-size:13px;}}
.page{{max-width:190mm;margin:0 auto;padding:0;}}

.header{{background:linear-gradient(135deg,#FF6A3B 0%,#E8551F 100%);color:#fff;padding:20px 26px;border-radius:12px;margin-bottom:14px;}}
.header-top{{display:flex;justify-content:space-between;align-items:center;margin-bottom:4px;}}
.logo{{font-size:15px;font-weight:800;letter-spacing:1px;}}
.header h1{{font-size:21px;font-weight:800;margin:4px 0;}}
.header .sub{{font-size:12px;opacity:0.85;}}

.section{{margin-bottom:14px;}}
.section-title{{font-size:15px;font-weight:700;color:#330C2E;margin-bottom:7px;}}

.kpi-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;}}
.kpi{{background:#FAFBFC;border:1px solid #E9ECEF;border-radius:10px;padding:12px;text-align:center;}}
.kpi .label{{font-size:11px;color:#636E72;margin-bottom:3px;}}
.kpi .value{{font-size:26px;font-weight:800;color:#330C2E;line-height:1.1;}}
.kpi .sub{{font-size:10px;color:#FF6A3B;font-weight:600;margin-top:3px;}}
.kpi.accent .value{{color:#FF6A3B;}}

.row{{display:grid;gap:12px;}}.row-2{{grid-template-columns:1fr 1fr;}}
.card{{background:#FAFBFC;border:1px solid #E9ECEF;border-radius:10px;padding:12px 14px;}}
.card h3{{font-size:13px;color:#330C2E;margin-bottom:6px;}}

.chart-container{{position:relative;height:150px;}}

table{{width:100%;border-collapse:collapse;font-size:11px;}}
th{{background:#FBF9F1;padding:5px 8px;text-align:center;font-weight:600;color:#330C2E;border-bottom:1.5px solid #E9ECEF;font-size:10px;}}
td{{padding:5px 8px;text-align:center;border-bottom:1px solid #E9ECEF;}}
.bold{{font-weight:700;}}

.funnel{{display:flex;flex-direction:column;gap:4px;}}
.funnel-step{{display:flex;align-items:center;gap:10px;}}
.funnel-bar{{height:26px;border-radius:5px;display:flex;align-items:center;padding:0 10px;color:#fff;font-weight:700;font-size:11px;min-width:55px;}}
.funnel-label{{width:80px;font-size:11px;color:#636E72;text-align:right;}}

.insight{{background:linear-gradient(135deg,#FFF9E6 0%,#FFF3CC 100%);border-left:3px solid #F39C12;border-radius:0 8px 8px 0;padding:10px 14px;font-size:11px;line-height:1.65;}}
.note{{font-size:10px;color:#999;margin-top:3px;}}

.footer{{text-align:center;padding:8px 0;color:#636E72;font-size:10px;border-top:1px solid #E9ECEF;margin-top:10px;}}
@media print{{
  html,body{{margin:0;padding:0;-webkit-print-color-adjust:exact;print-color-adjust:exact;}}
  .page{{padding:0;}}
  .section{{break-inside:avoid;page-break-inside:avoid;page-break-after:auto;}}
  .card{{break-inside:avoid;page-break-inside:avoid;}}
  .kpi-grid{{break-inside:avoid;page-break-inside:avoid;}}
  .row{{break-inside:avoid;page-break-inside:avoid;}}
  .insight{{break-inside:avoid;page-break-inside:avoid;}}
}}
</style>
</head><body>
<div class="page">

<div class="header">
  <div class="header-top">
    <div class="logo">UNNI GUIDE</div>
    <div style="font-size:10px;opacity:0.8;">Confidential | 제휴사 공유용</div>
  </div>
  <h1>오프라인 센터 운영 리포트 | 2026년 3월</h1>
  <div class="sub">2026.03.03 ~ 2026.03.31 | 서울 강남 언니가이드 센터</div>
</div>

<!-- 유동인구 배너 -->
<div style="background:linear-gradient(135deg,#FFF3CC 0%,#FFE8B8 100%);border-left:4px solid #F39C12;border-radius:10px;padding:10px 16px;margin-bottom:14px;display:flex;align-items:center;gap:14px;">
  <div style="font-size:26px;">📍</div>
  <div style="flex:1;">
    <div style="font-size:10px;color:#8B7500;font-weight:600;margin-bottom:2px;">LOCATION POWER</div>
    <div style="font-size:14px;font-weight:700;color:#330C2E;">센터 인근 500m 월 평균 유동인구 약 <span style="color:#FF6A3B;font-size:20px;">174만명</span></div>
    <div style="font-size:10px;color:#636E72;margin-top:2px;">2026년 3월 데이터 기준 · 강남 메인 상권 중심부 입지</div>
  </div>
</div>

<!-- KPI -->
<div class="section">
  <div class="section-title">센터 핵심 성과</div>
  <div class="kpi-grid" style="grid-template-columns:repeat(2,1fr);">
    <div class="kpi accent"><div class="label">센터 방문객 수</div><div class="value">{실제방문자:,}</div><div class="sub">일평균 {일평균방문:.0f}명 방문</div></div>
    <div class="kpi"><div class="label">센터 방문객 국적</div><div class="value">20+개국</div><div class="sub">글로벌 고객</div></div>
  </div>
  <div class="note">* 방문객 수는 예약 건 단위로 집계 (1건 예약 = 1명 카운트). 함께 방문한 대기 고객은 트래킹 대상에서 제외.</div>
</div>

<!-- 센터 분포 -->
<div class="section">
  <div class="row row-2">
    <div class="card">
      <h3>언어별 센터 방문 비중</h3>
      <table>
        <thead><tr><th>언어</th><th>방문 고객</th><th>비중</th></tr></thead>
        <tbody>
          <tr><td>🇬🇧 영어권</td><td class="bold">{영어_방문:,}</td><td class="bold">{영어_방문/max(실제방문자,1)*100:.1f}%</td></tr>
          <tr><td>🇹🇭 태국어권</td><td class="bold">{태국어_방문:,}</td><td class="bold">{태국어_방문/max(실제방문자,1)*100:.1f}%</td></tr>
          <tr><td>🇨🇳 중국어권</td><td class="bold">{중국어_방문:,}</td><td class="bold">{중국어_방문/max(실제방문자,1)*100:.1f}%</td></tr>
        </tbody>
      </table>
      <div style="margin-top:10px;">
        <h3>방문 고객 유형</h3>
        <table>
          <thead><tr><th>구분</th><th>건수</th><th>비중</th></tr></thead>
          <tbody>
            <tr><td>방한 외국인 (관광)</td><td class="bold">{방한_방문:,}</td><td>{방한_방문/max(실제방문자,1)*100:.1f}%</td></tr>
            <tr><td>재한 외국인 (거주)</td><td class="bold">{재한_방문:,}</td><td>{재한_방문/max(실제방문자,1)*100:.1f}%</td></tr>
          </tbody>
        </table>
        <div class="note">* 4월 1일부 센터 방문 고객 100% 방한 외국인 전환 완료</div>
      </div>
    </div>
    <div class="card">
      <h3>주간별 센터 방문 추이</h3>
      <div class="chart-container" style="height:240px;"><canvas id="weeklyChart"></canvas></div>
    </div>
  </div>
</div>

<!-- 구분선 -->
<div style="border-top:2px solid #FF6A3B;margin:14px 0 12px;opacity:0.25;"></div>

<!-- 서비스 실적: 국적 + 객단가 -->
<div class="section">
  <div class="section-title">언니가이드 서비스 실적 (온/오프라인 통합)</div>
  <div style="font-size:10px;color:#636E72;margin-bottom:8px;">언니가이드 서비스를 통해 시술·수술을 완료한 외국인 고객 데이터 (2026년 3월 정산 기준)</div>
  <div class="kpi-grid" style="grid-template-columns:repeat(2,1fr);">
    <div class="kpi accent"><div class="label">총 시/수술 금액</div><div class="value">{format_krw(total_매출)}</div><div class="sub">{num_국적}개국 고객 · 3월 단월 기준</div></div>
    <div class="kpi"><div class="label">평균 객단가</div><div class="value">{format_krw(avg_객단가)}</div><div class="sub">1인 평균 시/수술 금액</div></div>
  </div>
</div>

<div class="section">
  <div class="row row-2">
    <div class="card">
      <h3>국적별 고객 비중 · 객단가 (3월)</h3>
      <table>
        <thead><tr><th>국적</th><th>비중</th><th>인당 객단가</th></tr></thead>
        <tbody>{nat_rows_with_price_html}</tbody>
      </table>
    </div>
    <div class="card">
      <h3>주요 국적별 인당 객단가</h3>
      <div class="chart-container" style="height:200px;"><canvas id="priceChart"></canvas></div>
      <div class="note">※ 태국 · 대만 · 미국 · 중국 기준</div>
    </div>
  </div>
</div>

<div class="section">
  <div class="card">
    <h3>3월 인기 시술 카테고리 TOP 5</h3>
    <div class="note" style="margin-bottom:10px;">언니가이드 서비스 3월 시/수술 완료 고객 기준 (중복 포함)</div>
    {cat_rows_html}
  </div>
</div>

<!-- 인사이트 -->
<div class="section">
  <div class="insight">
    <strong>Key Insights</strong><br>
    · 3월 한 달간 <strong>총 {실제방문자:,}명</strong>의 외국인 고객이 센터에 직접 방문 (일평균 <strong>{일평균방문:.0f}명</strong>)<br>
    · 주간별 방문객 수 우상향 추세 — 외국인 대상 센터 인지도 빠르게 확산<br>
    · 언니가이드 서비스를 통한 3월 병원 시/수술 금액 <strong>{format_krw(total_매출)}</strong>, 평균 객단가 <strong>{format_krw(avg_객단가)}</strong><br>
    · <strong>{num_국적}개국</strong> 다국적 고객층 — 태국·대만·미국이 80% 차지, 고단가 고객(미국·중국·말레이시아)도 확인<br>
    · <strong>4월부터 100% 방한 외국인 전환</strong> — 인바운드 관광 타겟 브랜드에 최적 접점<br>
    <br>
    <strong style="color:#FF6A3B;">제휴 브랜드 노출 효과:</strong>
    월 <strong>{실제방문자:,}명+</strong> 외국인이 방문하는 오프라인 공간에서 K-Beauty 브랜드 체험 · 구매 전환까지 원스톱 접점 확보
  </div>
</div>

<div class="footer">
  <p><strong>UNNI GUIDE</strong> | 강남언니 언니가이드 오프라인 센터 · 서울 강남구</p>
  <p>본 리포트는 제휴 브랜드 전용 자료이며 외부 배포를 삼가해 주세요. | {datetime.now().strftime('%Y.%m.%d')}</p>
</div>

</div>
<script>
const D = {chart_data};
const C = ['#FF6A3B','#330C2E','#00B894','#FDCB6E','#0984E3','#E17055','#00CEC9','#A29BFE'];
new Chart(document.getElementById('weeklyChart'),{{
  type:'bar',
  data:{{labels:D.wk_labels,datasets:[
    {{data:D.wk_visit,backgroundColor:'#FF6A3B',borderRadius:6}}
  ]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:{{x:{{grid:{{display:false}},ticks:{{font:{{size:11}}}}}},y:{{grid:{{color:'#F1F2F6'}},ticks:{{font:{{size:10}}}},beginAtZero:true}}}}}}
}});
new Chart(document.getElementById('priceChart'),{{
  type:'bar',
  data:{{labels:D.price_labels,datasets:[{{data:D.price_values,backgroundColor:C.map(c=>c+'88'),borderColor:C,borderWidth:1,borderRadius:4}}]}},
  options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',plugins:{{legend:{{display:false}},tooltip:{{enabled:false}}}},scales:{{x:{{display:false,grid:{{display:false}}}},y:{{grid:{{display:false}},ticks:{{font:{{size:11,weight:'600'}}}}}}}}}}
}});
</script>
</body></html>"""

out_path = os.path.join(OUTPUT_DIR, 'unniguide_center_report_202603.html')
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"✅ PDF용: {out_path}")

# ============================================================
# Web 버전 HTML (크기 여유, 인터랙티브, URL 공유용)
# ============================================================
web_html = f"""<!DOCTYPE html>
<html lang="ko"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>언니가이드 센터 운영 리포트 | 2026년 3월</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans KR',sans-serif;color:#2D3436;line-height:1.6;background:#FAFBFC;font-size:15px;}}
.container{{max-width:1100px;margin:0 auto;padding:24px 20px 60px;}}

.top-nav{{position:sticky;top:0;z-index:100;background:rgba(255,255,255,0.95);backdrop-filter:blur(10px);border-bottom:1px solid #E9ECEF;padding:12px 0;margin-bottom:0;}}
.top-nav-inner{{max-width:1100px;margin:0 auto;padding:0 20px;display:flex;gap:16px;align-items:center;flex-wrap:wrap;}}
.top-nav a{{text-decoration:none;color:#636E72;font-size:14px;font-weight:600;padding:8px 14px;border-radius:8px;transition:all 0.15s;}}
.top-nav a:hover{{background:#FFF0EB;color:#FF6A3B;}}
.top-nav .spacer{{flex:1;}}

.header{{background:linear-gradient(135deg,#FF6A3B 0%,#E8551F 100%);color:white;padding:48px 40px;border-radius:0 0 24px 24px;margin:-24px -20px 32px;}}
.header-top{{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;}}
.logo{{font-size:16px;font-weight:800;letter-spacing:1.5px;}}
.header h1{{font-size:32px;font-weight:800;margin:8px 0;}}
.header .sub{{font-size:15px;opacity:0.9;}}

.location-banner{{background:linear-gradient(135deg,#FFF3CC 0%,#FFE8B8 100%);border-left:6px solid #F39C12;border-radius:12px;padding:20px 28px;margin-bottom:32px;display:flex;align-items:center;gap:20px;}}
.location-banner .icon{{font-size:40px;}}
.location-banner .label{{font-size:12px;color:#8B7500;font-weight:700;letter-spacing:1px;margin-bottom:4px;}}
.location-banner .main{{font-size:20px;font-weight:700;color:#330C2E;}}
.location-banner .main strong{{color:#FF6A3B;font-size:28px;}}
.location-banner .sub{{font-size:13px;color:#636E72;margin-top:4px;}}

.section{{margin-bottom:40px;}}
.section-title{{font-size:22px;font-weight:800;color:#330C2E;margin-bottom:8px;display:flex;align-items:center;gap:10px;}}
.section-desc{{font-size:14px;color:#636E72;margin-bottom:20px;}}

.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:16px;}}
.kpi{{background:white;border:1px solid #E9ECEF;border-radius:14px;padding:24px;text-align:center;box-shadow:0 2px 12px rgba(0,0,0,0.04);transition:transform 0.2s;}}
.kpi:hover{{transform:translateY(-2px);box-shadow:0 8px 24px rgba(0,0,0,0.08);}}
.kpi .label{{font-size:13px;color:#636E72;margin-bottom:8px;}}
.kpi .value{{font-size:38px;font-weight:800;color:#330C2E;line-height:1.1;}}
.kpi .sub{{font-size:13px;color:#FF6A3B;font-weight:600;margin-top:6px;}}
.kpi.accent .value{{color:#FF6A3B;}}

.row{{display:grid;gap:20px;}}.row-2{{grid-template-columns:1fr 1fr;}}
.card{{background:white;border:1px solid #E9ECEF;border-radius:14px;padding:24px 28px;box-shadow:0 2px 12px rgba(0,0,0,0.04);}}
.card h3{{font-size:17px;color:#330C2E;margin-bottom:14px;font-weight:700;}}

.chart-container{{position:relative;height:280px;margin-top:12px;}}

table{{width:100%;border-collapse:collapse;font-size:14px;}}
th{{background:#FBF9F1;padding:12px 14px;text-align:center;font-weight:700;color:#330C2E;border-bottom:2px solid #E9ECEF;font-size:12px;text-transform:uppercase;letter-spacing:0.5px;}}
td{{padding:12px 14px;text-align:center;border-bottom:1px solid #E9ECEF;}}
tbody tr:hover{{background:#FFF8F5;}}
.bold{{font-weight:700;}}

.insight{{background:linear-gradient(135deg,#FFF9E6 0%,#FFF3CC 100%);border-left:4px solid #F39C12;border-radius:0 12px 12px 0;padding:20px 28px;font-size:14px;line-height:1.8;}}
.insight ul{{padding-left:20px;margin-top:8px;}}
.insight li{{margin-bottom:6px;}}
.insight strong.highlight{{color:#FF6A3B;}}
.note{{font-size:12px;color:#999;margin-top:8px;}}

.footer{{text-align:center;padding:32px 0;color:#636E72;font-size:13px;border-top:1px solid #E9ECEF;margin-top:48px;}}

@media (max-width:768px){{
  .row-2{{grid-template-columns:1fr;}}
  .header{{padding:32px 20px;}}
  .header h1{{font-size:24px;}}
  .kpi .value{{font-size:30px;}}
}}
@media print{{
  html,body{{margin:0;padding:0;-webkit-print-color-adjust:exact;print-color-adjust:exact;background:#fff;}}
  .top-nav{{display:none !important;}}
  .container{{padding:12px;}}
  .section{{break-inside:avoid;page-break-inside:avoid;margin-bottom:20px;}}
  .card{{break-inside:avoid;page-break-inside:avoid;}}
  .kpi-grid{{break-inside:avoid;page-break-inside:avoid;}}
  .kpi{{break-inside:avoid;}}
  .row{{break-inside:avoid;page-break-inside:avoid;}}
  .insight{{break-inside:avoid;page-break-inside:avoid;}}
  .location-banner{{break-inside:avoid;page-break-inside:avoid;}}
  .header{{break-after:avoid;}}
}}
</style>
</head><body>

<div class="top-nav">
  <div class="top-nav-inner">
    <a href="#kpi">핵심 성과</a>
    <a href="#language">언어 · 유형</a>
    <a href="#trend">주간 추이</a>
    <a href="#service">서비스 실적</a>
    <a href="#nationality">국적 객단가</a>
    <a href="#categories">인기 시술</a>
    <a href="#insight">Key Insights</a>
    <div class="spacer"></div>
    <span style="font-size:12px;color:#999;">Confidential</span>
  </div>
</div>

<div class="container">

<div class="header">
  <div class="header-top">
    <div class="logo">UNNI GUIDE</div>
    <div style="font-size:13px;opacity:0.8;">제휴사 공유용</div>
  </div>
  <h1>오프라인 센터 운영 리포트</h1>
  <div class="sub">2026.03.03 ~ 2026.03.31 (29일) · 서울 강남 언니가이드 센터</div>
</div>

<!-- 유동인구 배너 -->
<div class="location-banner">
  <div class="icon">📍</div>
  <div style="flex:1;">
    <div class="label">LOCATION POWER</div>
    <div class="main">센터 인근 500m 월 평균 유동인구 약 <strong>174만명</strong></div>
    <div class="sub">2026년 3월 데이터 기준 · 강남 메인 상권 중심부 입지</div>
  </div>
</div>

<!-- KPI -->
<div class="section" id="kpi">
  <div class="section-title">📊 센터 핵심 성과</div>
  <div class="section-desc">외국인 고객 대상 월간 센터 운영 지표입니다.</div>
  <div class="kpi-grid">
    <div class="kpi accent"><div class="label">센터 방문객 수</div><div class="value">{실제방문자:,}</div><div class="sub">일평균 {일평균방문:.0f}명 방문</div></div>
    <div class="kpi"><div class="label">센터 방문 예약 수</div><div class="value">{예약고객수:,}</div><div class="sub">일평균 {일평균예약:.0f}건 예약</div></div>
    <div class="kpi"><div class="label">센터 방문객 국적</div><div class="value">20+개국</div><div class="sub">글로벌 고객</div></div>
  </div>
  <div class="note">* 방문객 수는 예약 건 단위로 집계 (1건 = 1명 카운트). 함께 방문한 대기 고객은 트래킹 대상에서 제외.</div>
</div>

<!-- 언어 & 방문 유형 -->
<div class="section" id="language">
  <div class="section-title">🌐 언어권 · 방문 고객 유형</div>
  <div class="section-desc">방문 고객의 언어권 분포와 방한/재한 비중입니다.</div>
  <div class="row row-2">
    <div class="card">
      <h3>언어별 센터 방문 비중</h3>
      <table>
        <thead><tr><th>언어</th><th>방문 고객</th><th>비중</th></tr></thead>
        <tbody>
          <tr><td>🇬🇧 영어권</td><td class="bold">{영어_방문:,}명</td><td class="bold">{영어_방문/max(실제방문자,1)*100:.1f}%</td></tr>
          <tr><td>🇹🇭 태국어권</td><td class="bold">{태국어_방문:,}명</td><td class="bold">{태국어_방문/max(실제방문자,1)*100:.1f}%</td></tr>
          <tr><td>🇨🇳 중국어권</td><td class="bold">{중국어_방문:,}명</td><td class="bold">{중국어_방문/max(실제방문자,1)*100:.1f}%</td></tr>
        </tbody>
      </table>
    </div>
    <div class="card">
      <h3>방문 고객 유형</h3>
      <table>
        <thead><tr><th>구분</th><th>건수</th><th>비중</th></tr></thead>
        <tbody>
          <tr><td>방한 외국인 (관광)</td><td class="bold">{방한_방문:,}명</td><td>{방한_방문/max(실제방문자,1)*100:.1f}%</td></tr>
          <tr><td>재한 외국인 (거주)</td><td class="bold">{재한_방문:,}명</td><td>{재한_방문/max(실제방문자,1)*100:.1f}%</td></tr>
        </tbody>
      </table>
      <div class="note">* 4월 1일부 센터 방문 고객 100% 방한 외국인 전환 완료</div>
    </div>
  </div>
</div>

<!-- 주간 추이 -->
<div class="section" id="trend">
  <div class="section-title">📈 주간별 센터 방문 추이</div>
  <div class="section-desc">센터 오픈 이후 누적 방문객 수 추이. 지속적인 우상향 확인.</div>
  <div class="card">
    <div class="chart-container" style="height:340px;"><canvas id="weeklyChartWeb"></canvas></div>
  </div>
</div>

<!-- 서비스 실적 -->
<div class="section" id="service">
  <div class="section-title">💼 언니가이드 서비스 실적 (온/오프라인 통합)</div>
  <div class="section-desc">언니가이드 서비스를 통해 시술·수술을 완료한 외국인 고객 데이터 (2026년 3월 정산 기준)</div>
  <div class="kpi-grid" style="grid-template-columns:repeat(2,1fr);">
    <div class="kpi accent"><div class="label">총 시/수술 금액</div><div class="value">{format_krw(total_매출)}</div><div class="sub">{num_국적}개국 고객 · 3월 단월 기준</div></div>
    <div class="kpi"><div class="label">평균 객단가</div><div class="value">{format_krw(avg_객단가)}</div><div class="sub">1인 평균 시/수술 금액</div></div>
  </div>
</div>

<!-- 국적 -->
<div class="section" id="nationality">
  <div class="section-title">🌍 국적별 비중 · 객단가 분포</div>
  <div class="section-desc">3월 시/수술 완료 고객 국적별 분포와 TOP 5 객단가 국적입니다.</div>
  <div class="row row-2">
    <div class="card">
      <h3>국적별 고객 비중 · 객단가</h3>
      <table>
        <thead><tr><th>국적</th><th>비중</th><th>인당 객단가</th></tr></thead>
        <tbody>{nat_rows_with_price_html}</tbody>
      </table>
    </div>
    <div class="card">
      <h3>주요 국적별 인당 객단가</h3>
      <div class="chart-container" style="height:260px;"><canvas id="priceChartWeb"></canvas></div>
      <div class="note">※ 고객 프리미엄 시술 선호도 지표</div>
    </div>
  </div>
</div>

<!-- 인기 시술 카테고리 -->
<div class="section" id="categories">
  <div class="section-title">🔥 3월 인기 시술 카테고리 TOP 5</div>
  <div class="section-desc">언니가이드 서비스 3월 시/수술 완료 고객 기준 (중복 포함)</div>
  <div class="card">
    {cat_rows_html}
  </div>
</div>

<!-- 인사이트 -->
<div class="section" id="insight">
  <div class="section-title">💡 Key Insights</div>
  <div class="insight">
    <ul style="list-style:none;padding-left:0;">
      <li>📍 3월 한 달간 <strong class="highlight">총 {실제방문자:,}명</strong>의 외국인 고객이 센터에 직접 방문 (일평균 <strong>{일평균방문:.0f}명</strong>)</li>
      <li>📊 주간별 방문객 수 <strong>우상향 추세</strong> — 외국인 대상 센터 인지도 빠르게 확산</li>
      <li>💰 언니가이드 서비스 3월 시/수술 금액 <strong class="highlight">{format_krw(total_매출)}</strong>, 평균 객단가 <strong>{format_krw(avg_객단가)}</strong></li>
      <li>🌍 <strong>{num_국적}개국</strong> 다국적 고객층 — 태국·대만·미국이 80% 차지, 고단가 고객(미국·중국·말레이시아)도 고르게 분포</li>
      <li>🎯 <strong>4월부터 100% 방한 외국인 전환</strong> — 인바운드 관광 타겟 브랜드에 최적 접점</li>
    </ul>
    <div style="margin-top:20px;padding-top:16px;border-top:1px solid rgba(243,156,18,0.3);">
      <strong class="highlight" style="font-size:16px;">🏢 제휴 브랜드 노출 효과</strong>
      <div style="margin-top:8px;">
        월 <strong class="highlight">{실제방문자:,}명+</strong> 외국인이 방문하는 오프라인 공간에서 K-Beauty 브랜드 체험 · 구매 전환까지 <strong>원스톱 접점 확보</strong>
      </div>
    </div>
  </div>
</div>

<div class="footer">
  <strong>UNNI GUIDE</strong> | 강남언니 언니가이드 오프라인 센터 · 서울 강남구<br>
  본 리포트는 제휴 브랜드 전용 자료이며 외부 배포를 삼가해 주세요. | {datetime.now().strftime('%Y.%m.%d')}
</div>

</div>

<script>
const D = {chart_data};
const C = ['#FF6A3B','#330C2E','#00B894','#FDCB6E','#0984E3','#E17055','#00CEC9','#A29BFE'];

new Chart(document.getElementById('weeklyChartWeb'),{{
  type:'bar',
  data:{{labels:D.wk_labels,datasets:[
    {{data:D.wk_visit,backgroundColor:'#FF6A3B',borderRadius:8}}
  ]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}},tooltip:{{titleFont:{{size:14}},bodyFont:{{size:14}},padding:12}}}},scales:{{x:{{grid:{{display:false}},ticks:{{font:{{size:13}}}}}},y:{{grid:{{color:'#F1F2F6'}},ticks:{{font:{{size:12}}}},beginAtZero:true}}}}}}
}});

new Chart(document.getElementById('priceChartWeb'),{{
  type:'bar',
  data:{{labels:D.price_labels,datasets:[{{data:D.price_values,backgroundColor:C.map(c=>c+'88'),borderColor:C,borderWidth:2,borderRadius:6}}]}},
  options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',plugins:{{legend:{{display:false}},tooltip:{{enabled:false}}}},scales:{{x:{{display:false,grid:{{display:false}}}},y:{{grid:{{display:false}},ticks:{{font:{{size:14,weight:'600'}}}}}}}}}}
}});
</script>
</body></html>"""

web_path = os.path.join(OUTPUT_DIR, 'unniguide_center_report_web_202603.html')
with open(web_path, 'w', encoding='utf-8') as f:
    f.write(web_html)
print(f"✅ Web용: {web_path}")

print(f"\n=== 센터 ===  방문객: {실제방문자:,} | 예약: {예약고객수:,}")
print(f"=== 서비스 === {total_건수}건 | {format_krw(total_매출)} | 객단가 {format_krw(avg_객단가)} | {num_국적}개국")
