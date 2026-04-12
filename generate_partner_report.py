#!/usr/bin/env python3
"""
언니가이드 오프라인 센터 + 서비스 실적 제휴사용 원페이지 리포트
기간: 2026.03.03 (정식 오픈) ~ 2026.04.02 (오픈 첫 1개월)
"""
import pandas as pd
import openpyxl
import os, json
from datetime import datetime

CSV_PATH = os.path.expanduser("~/Downloads/언니가이드_운영의 사본 - Offline - Daily (외국인).csv")
EXCEL_PATH = os.path.expanduser("~/Downloads/언니가이드 운영 트렌드 데이터_26.04.xlsx")
OUTPUT_DIR = os.path.expanduser("~/Documents/unniguide-report")

NAME_NORMALIZE = {
    '사적인아름다운지유의원': '사적인아름다움지유의원', '루호성형외과': '루호성형외과의원',
    '테이아 의원': '테이아의원', '톡스앤필 - 신논': '톡스앤필의원-신논현점',
    '톡스앤필 - 신논현': '톡스앤필의원-신논현점', '제이필 - 홍대': '제이필의원-홍대점',
    '제이필 - 강남': '제이필의원-강남점', '유픽의원 홍대': '유픽의원-홍대점',
    '유픽의원-홍대': '유픽의원-홍대점', '유픽의원-강남': '유픽의원-강남점',
    '홍대셀레나': '홍대셀레나의원', '히트성형외과': '히트성형외과의원',
    '플래너성형외과': '플래너성형외과의원', '플래저성형외과': '플레저성형외과의원',
}
COUNTRY_FLAG = {
    '태국':'🇹🇭','대만':'🇹🇼','중국':'🇨🇳','미국':'🇺🇸','호주':'🇦🇺','싱가포르':'🇸🇬',
    '몽골':'🇲🇳','캐나다':'🇨🇦','말레이시아':'🇲🇾','영국':'🇬🇧','홍콩':'🇭🇰',
    '필리핀':'🇵🇭','프랑스':'🇫🇷','독일':'🇩🇪','인도':'🇮🇳','일본':'🇯🇵',
    '폴란드':'🇵🇱','아일랜드':'🇮🇪','키르기스스탄':'🇰🇬',
}

def normalize_hospital(name):
    if not name or str(name).strip() in ('', 'nan', 'None'): return None
    name = str(name).strip()
    return NAME_NORMALIZE.get(name, name)

def format_krw(amount):
    if amount >= 100_000_000: return f"{amount/100_000_000:.1f}억"
    elif amount >= 10_000: return f"{amount/10_000:,.0f}만원"
    else: return f"{amount:,.0f}원"

# ============================================================
# 1. 오프라인 센터 데이터
# ============================================================
df = pd.read_csv(CSV_PATH, encoding='latin1', header=None)

def fix_enc(s):
    if pd.isna(s): return ''
    try: return str(s).encode('latin1').decode('euc-kr')
    except:
        try: return str(s).encode('latin1').decode('cp949')
        except: return str(s)

dates_row = df.iloc[4].tolist()
target_cols = [i for i, d in enumerate(dates_row) if str(d).strip()[:10] >= '2026-03-03' and str(d).strip()[:10] <= '2026-04-02']
num_days = len(target_cols)

def row_sum(row_idx):
    total = 0
    for c in target_cols:
        v = df.iloc[row_idx, c]
        if pd.isna(v) or str(v).strip() in ('-', '', 'nan'): continue
        try: total += float(str(v).replace(',', ''))
        except: pass
    return int(total)

def daily_vals(row_idx):
    return [(str(dates_row[c])[:10], int(float(str(df.iloc[row_idx, c]).replace(',', ''))) if pd.notna(df.iloc[row_idx, c]) and str(df.iloc[row_idx, c]).strip() not in ('-', '', 'nan') else 0) for c in target_cols]

예약고객수 = row_sum(6)
실제방문자 = row_sum(18)
방한_예약 = row_sum(10)
재한_예약 = row_sum(14)
영어 = row_sum(7); 태국어 = row_sum(9); 중국어 = row_sum(8)
영어_방문 = row_sum(19); 태국어_방문 = row_sum(21); 중국어_방문 = row_sum(20)
일평균방문 = 실제방문자 / num_days
일평균예약 = 예약고객수 / num_days

daily_b = daily_vals(6)
daily_v = daily_vals(18)
chart_labels = [d[5:] for d, _ in daily_b]
chart_booking = [v for _, v in daily_b]
chart_visit = [v for _, v in daily_v]

weeks_row = df.iloc[3].tolist()
weekly = {}
for idx, c in enumerate(target_cols):
    w = f"W{str(weeks_row[c]).strip()}"
    if w not in weekly: weekly[w] = {'예약': 0, '방문': 0}
    weekly[w]['예약'] += chart_booking[idx]
    weekly[w]['방문'] += chart_visit[idx]

# ============================================================
# 2. 정산 데이터 (3월 국적별/객단가)
# ============================================================
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
ws2 = wb['병원별 정산 및 언니가이드 매출 데이터']
settle_records = []
current_month = None
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=False):
    vals = [c.value for c in row]
    a = str(vals[0]).strip() if vals[0] else ''
    if '정산 내역' in a:
        parts = a.replace('년','-').replace('월','').replace('정산 내역','').strip()
        try:
            y, m = parts.split('-')[:2]
            current_month = f"{y.strip()}-{int(m.strip()):02d}"
        except: pass
        continue
    if current_month == '2026-03' and vals[1]:
        hospital = str(vals[1]).strip()
        if hospital in ('병원명', ''): continue
        try:
            settle_records.append({
                '국적': str(vals[4]).strip() if vals[4] else '',
                '시술금액': float(vals[7]) if vals[7] else 0,
            })
        except: pass
wb.close()

df_settle = pd.DataFrame(settle_records)
nat_stats = df_settle.groupby('국적').agg(건수=('시술금액','count'), 매출=('시술금액','sum')).sort_values('건수', ascending=False).reset_index()
nat_stats['비중'] = (nat_stats['건수'] / nat_stats['건수'].sum() * 100).round(1)
nat_stats['객단가'] = (nat_stats['매출'] / nat_stats['건수']).round(0)
nat_stats['국기'] = nat_stats['국적'].map(COUNTRY_FLAG).fillna('🌍')

total_건수 = nat_stats['건수'].sum()
total_매출 = nat_stats['매출'].sum()
avg_객단가 = total_매출 / total_건수 if total_건수 > 0 else 0
num_국적 = len(nat_stats)

# 국적별 테이블 HTML
nat_rows_html = ""
for _, r in nat_stats.head(8).iterrows():
    nat_rows_html += f'<tr><td>{r["국기"]} {r["국적"]}</td><td class="bold">{int(r["건수"])}건</td><td>{r["비중"]}%</td><td class="bold">{format_krw(r["객단가"])}</td></tr>\n'
if len(nat_stats) > 8:
    etc_cnt = nat_stats.iloc[8:]['건수'].sum()
    etc_pct = (etc_cnt / total_건수 * 100)
    nat_rows_html += f'<tr><td>🌍 기타 {len(nat_stats)-8}개국</td><td>{int(etc_cnt)}건</td><td>{etc_pct:.1f}%</td><td>-</td></tr>\n'

# 객단가 상위 차트 데이터
top_price = nat_stats[nat_stats['건수'] >= 2].sort_values('객단가', ascending=False).head(8)
price_labels = [f"{r['국기']} {r['국적']}" for _, r in top_price.iterrows()]
price_values = [int(r['객단가']) for _, r in top_price.iterrows()]

chart_data = json.dumps({
    'labels': chart_labels, 'booking': chart_booking, 'visit': chart_visit,
    'wk_labels': list(weekly.keys()),
    'wk_booking': [v['예약'] for v in weekly.values()],
    'wk_visit': [v['방문'] for v in weekly.values()],
    'nat_labels': [r['국적'] for _, r in nat_stats.head(6).iterrows()],
    'nat_counts': [int(r['건수']) for _, r in nat_stats.head(6).iterrows()],
    'price_labels': price_labels,
    'price_values': price_values,
}, ensure_ascii=False)

# ============================================================
# HTML
# ============================================================
html = f"""<!DOCTYPE html>
<html lang="ko"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>UNNI GUIDE 센터 리포트 | 오픈 첫 1개월</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
@page {{ size:A4; margin:10mm; }}
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI','Noto Sans KR',sans-serif;color:#2D3436;line-height:1.5;background:#fff;font-size:12px;}}
.page{{max-width:860px;margin:0 auto;padding:12px;}}

.header{{background:linear-gradient(135deg,#FF6A3B 0%,#E8551F 100%);color:#fff;padding:22px 26px;border-radius:14px;margin-bottom:16px;}}
.header-top{{display:flex;justify-content:space-between;align-items:center;margin-bottom:4px;}}
.logo{{font-size:14px;font-weight:800;letter-spacing:1px;}}
.header h1{{font-size:19px;font-weight:800;}}
.header .sub{{font-size:11px;opacity:0.85;margin-top:2px;}}

.section{{margin-bottom:14px;}}
.section-title{{font-size:14px;font-weight:700;color:#330C2E;margin-bottom:6px;}}

.kpi-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:8px;}}
.kpi{{background:#FAFBFC;border:1px solid #E9ECEF;border-radius:10px;padding:10px;text-align:center;}}
.kpi .label{{font-size:9px;color:#636E72;margin-bottom:1px;}}
.kpi .value{{font-size:24px;font-weight:800;color:#330C2E;}}
.kpi .sub{{font-size:9px;color:#FF6A3B;font-weight:600;margin-top:1px;}}
.kpi.accent .value{{color:#FF6A3B;}}

.row{{display:grid;gap:10px;}}.row-2{{grid-template-columns:1fr 1fr;}}
.card{{background:#FAFBFC;border:1px solid #E9ECEF;border-radius:10px;padding:12px;}}
.card h3{{font-size:12px;color:#330C2E;margin-bottom:5px;}}

.chart-container{{position:relative;height:140px;}}

table{{width:100%;border-collapse:collapse;font-size:10px;}}
th{{background:#FBF9F1;padding:5px 6px;text-align:center;font-weight:600;color:#330C2E;border-bottom:2px solid #E9ECEF;font-size:9px;}}
td{{padding:5px 6px;text-align:center;border-bottom:1px solid #E9ECEF;}}
.bold{{font-weight:700;}}

.insight{{background:linear-gradient(135deg,#FFF9E6 0%,#FFF3CC 100%);border-left:3px solid #F39C12;border-radius:0 8px 8px 0;padding:8px 12px;font-size:10px;line-height:1.6;}}
.note{{font-size:9px;color:#999;margin-top:4px;}}

.footer{{text-align:center;padding:10px 0;color:#636E72;font-size:9px;border-top:1px solid #E9ECEF;margin-top:12px;}}
@media print{{.page{{padding:0;}}}}
</style>
</head><body>
<div class="page">

<div class="header">
  <div class="header-top">
    <div class="logo">UNNI GUIDE</div>
    <div style="font-size:10px;opacity:0.8;">Confidential | 제휴사 공유용</div>
  </div>
  <h1>오프라인 센터 운영 리포트 | 오픈 첫 1개월</h1>
  <div class="sub">2026.03.03 (정식 오픈) ~ 2026.04.02 | 서울 강남 언니가이드 센터</div>
</div>

<!-- KPI -->
<div class="section">
  <div class="section-title">센터 핵심 성과</div>
  <div class="kpi-grid">
    <div class="kpi accent"><div class="label">센터 방문객 수</div><div class="value">{실제방문자:,}</div><div class="sub">일평균 {일평균방문:.0f}명 방문</div></div>
    <div class="kpi"><div class="label">센터 방문 예약 수</div><div class="value">{예약고객수:,}</div><div class="sub">일평균 {일평균예약:.0f}건 예약 인입</div></div>
    <div class="kpi"><div class="label">방문 고객 국적</div><div class="value">3개 언어권</div><div class="sub">영어 · 중국어 · 태국어</div></div>
  </div>
  <div class="note">* 방문 예약: 기간 내 접수된 센터 방문 예약 건수 (미래 방문 포함) / 방문객: 기간 내 실제 센터를 방문한 고객 수</div>
</div>

<!-- 센터 분포 -->
<div class="section">
  <div class="row row-2">
    <div class="card">
      <h3>언어별 방문 예약 비중</h3>
      <table>
        <thead><tr><th>언어</th><th>예약</th><th>비중</th><th>실제 방문</th></tr></thead>
        <tbody>
          <tr><td>🇬🇧 영어권</td><td class="bold">{영어:,}</td><td class="bold">{영어/max(예약고객수,1)*100:.1f}%</td><td>{영어_방문:,}</td></tr>
          <tr><td>🇹🇭 태국어권</td><td class="bold">{태국어:,}</td><td class="bold">{태국어/max(예약고객수,1)*100:.1f}%</td><td>{태국어_방문:,}</td></tr>
          <tr><td>🇨🇳 중국어권</td><td class="bold">{중국어:,}</td><td class="bold">{중국어/max(예약고객수,1)*100:.1f}%</td><td>{중국어_방문:,}</td></tr>
        </tbody>
      </table>
      <div style="margin-top:8px;">
        <h3>방문 고객 유형</h3>
        <table>
          <thead><tr><th>구분</th><th>건수</th><th>비중</th></tr></thead>
          <tbody>
            <tr><td>방한 외국인 (관광)</td><td class="bold">{방한_예약:,}</td><td>{방한_예약/max(예약고객수,1)*100:.1f}%</td></tr>
            <tr><td>재한 외국인 (거주)</td><td class="bold">{재한_예약:,}</td><td>{재한_예약/max(예약고객수,1)*100:.1f}%</td></tr>
          </tbody>
        </table>
        <div class="note">* 4월 1일부 센터 방문 고객 100% 방한 외국인 전환 완료</div>
      </div>
    </div>
    <div class="card">
      <h3>일별 센터 트래픽 추이</h3>
      <div class="chart-container"><canvas id="dailyChart"></canvas></div>
      <div class="chart-container" style="height:120px;margin-top:8px;"><canvas id="weeklyChart"></canvas></div>
    </div>
  </div>
</div>

<!-- 구분선 -->
<div style="border-top:2px solid #FF6A3B;margin:16px 0 12px;opacity:0.3;"></div>

<!-- 서비스 실적: 국적 + 객단가 -->
<div class="section">
  <div class="section-title">언니가이드 서비스 실적 (2026년 3월 정산 기준)</div>
  <div class="kpi-grid">
    <div class="kpi accent"><div class="label">시/수술 완료</div><div class="value">{total_건수:,}건</div><div class="sub">{num_국적}개국 고객</div></div>
    <div class="kpi"><div class="label">총 시/수술 금액</div><div class="value">{format_krw(total_매출)}</div><div class="sub">3월 단월 기준</div></div>
    <div class="kpi"><div class="label">평균 객단가</div><div class="value">{format_krw(avg_객단가)}</div><div class="sub">1인 평균 시/수술 금액</div></div>
  </div>
</div>

<div class="section">
  <div class="row row-2">
    <div class="card">
      <h3>국적별 고객 비중 & 객단가 (3월)</h3>
      <table>
        <thead><tr><th>국적</th><th>건수</th><th>비중</th><th>인당 객단가</th></tr></thead>
        <tbody>
          {nat_rows_html}
        </tbody>
      </table>
    </div>
    <div class="card">
      <h3>국적별 인당 객단가 (2건 이상)</h3>
      <div class="chart-container" style="height:180px;"><canvas id="priceChart"></canvas></div>
    </div>
  </div>
</div>

<!-- 인사이트 -->
<div class="section">
  <div class="insight">
    <strong>Key Insights</strong><br>
    · 오픈 첫 달, <strong>총 {실제방문자:,}명</strong>의 외국인 고객이 센터에 직접 방문하여 일평균 <strong>{일평균방문:.0f}명</strong> 트래픽 확인<br>
    · 방문 예약 수요 <strong>{예약고객수:,}건</strong> — 센터 인지도 빠르게 확산, 오픈 후반부 우상향 추세<br>
    · 언니가이드 서비스를 통한 3월 병원 시/수술 금액 <strong>{format_krw(total_매출)}</strong>, 평균 객단가 <strong>{format_krw(avg_객단가)}</strong><br>
    · <strong>{num_국적}개국</strong> 다국적 고객층 — 태국·대만·미국이 80% 차지, 고단가 고객(미국·중국·말레이시아)도 확인<br>
    · <strong>4월부터 100% 방한 외국인 전환</strong> — 인바운드 관광 타겟 브랜드에 최적 접점<br>
    <br>
    <strong style="color:#FF6A3B;">제휴 브랜드 노출 효과:</strong>
    월 <strong>{실제방문자:,}명+</strong> 외국인이 방문하는 오프라인 공간에서 K-Beauty 브랜드 체험 · 구매 전환까지 원스톱 접점 확보
  </div>
</div>

<div class="footer">
  <p><strong>UNNI GUIDE</strong> | 강남언니 언니가이드 오프라인 센터 · 서울 강남구</p>
  <p>본 리포트는 제휴 브랜드 전용 자료이며 외부 배포를 삼가해 주세요. | {datetime.now().strftime('%Y.%m.%d')}</p>
</div>

</div>
<script>
const D = {chart_data};
const C = ['#FF6A3B','#330C2E','#00B894','#FDCB6E','#0984E3','#E17055','#00CEC9','#A29BFE'];
new Chart(document.getElementById('dailyChart'),{{
  type:'bar',
  data:{{labels:D.labels,datasets:[
    {{label:'방문 예약',data:D.booking,backgroundColor:'#FF6A3B66',borderColor:'#FF6A3B',borderWidth:1,borderRadius:2}},
    {{label:'실제 방문',data:D.visit,type:'line',borderColor:'#330C2E',pointRadius:1,tension:0.3,borderWidth:1.5}}
  ]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'bottom',labels:{{font:{{size:8}}}}}}}},scales:{{x:{{ticks:{{font:{{size:6}},maxRotation:45}},grid:{{display:false}}}},y:{{grid:{{color:'#F1F2F6'}},ticks:{{font:{{size:8}}}}}}}}}}
}});
new Chart(document.getElementById('weeklyChart'),{{
  type:'bar',
  data:{{labels:D.wk_labels,datasets:[
    {{label:'예약',data:D.wk_booking,backgroundColor:'#FF6A3B',borderRadius:4}},
    {{label:'방문',data:D.wk_visit,backgroundColor:'#330C2E',borderRadius:4}}
  ]}},
  options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'bottom',labels:{{font:{{size:8}}}}}}}},scales:{{x:{{grid:{{display:false}},ticks:{{font:{{size:8}}}}}},y:{{grid:{{color:'#F1F2F6'}},ticks:{{font:{{size:8}}}}}}}}}}
}});
new Chart(document.getElementById('priceChart'),{{
  type:'bar',
  data:{{labels:D.price_labels,datasets:[{{data:D.price_values,backgroundColor:C.map(c=>c+'88'),borderColor:C,borderWidth:1,borderRadius:4}}]}},
  options:{{responsive:true,maintainAspectRatio:false,indexAxis:'y',plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{callback:v=>(v/10000).toFixed(0)+'만원',font:{{size:8}}}},grid:{{color:'#F1F2F6'}}}},y:{{grid:{{display:false}},ticks:{{font:{{size:9}}}}}}}}}}
}});
</script>
</body></html>"""

out_path = os.path.join(OUTPUT_DIR, 'unniguide_center_report_202603.html')
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"✅ {out_path}")
print(f"\n=== 센터 ===  방문객: {실제방문자:,} | 예약: {예약고객수:,}")
print(f"=== 서비스 === {total_건수}건 | {format_krw(total_매출)} | 객단가 {format_krw(avg_객단가)} | {num_국적}개국")
