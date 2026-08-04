#!/usr/bin/env python3
"""
언니가이드 인터랙티브 대시보드 (Streamlit) v2
- 원천 2개를 내원일 기준으로 이어붙임: ~2026-06 운영시트 / 2026-07~ 어드민 case-metrics
- 내부리포트/정산은 raw에서 자동 집계
- 취소/노쇼 트래킹 포함
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
import re
import os
import io
import glob
from collections import defaultdict
from datetime import datetime

from admin_source import (
    ADMIN_CUTOFF_MONTH,
    ADMIN_HOSPITAL_ALIASES,
    ADMIN_SHEET_KEYWORD,
    SHEET_HOSPITAL_ALIASES,
    load_admin_case_metrics,
    normalize_admin_frame,
)

# ============================================================
# 페이지 설정
# ============================================================
st.set_page_config(
    page_title="언니가이드 대시보드",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================
# 브랜드 스타일
# ============================================================
BRAND_ORANGE = '#FF6A3B'
BRAND_PLUM = '#330C2E'
BRAND_IVORY = '#FBF9F1'
BRAND_GREEN = '#00B894'
BRAND_RED = '#E74C3C'

CHART_COLORS = [
    '#FF6A3B', '#330C2E', '#00B894', '#FDCB6E', '#0984E3',
    '#E17055', '#00CEC9', '#A29BFE', '#FD79A8', '#55A3E8',
    '#F39C12', '#2ECC71', '#E74C3C', '#9B59B6', '#1ABC9C',
]

st.markdown("""
<style>
    .block-container { padding-top: 1rem; }
    [data-testid="stMetric"] {
        background: white; border: 1px solid #E9ECEF;
        border-radius: 12px; padding: 16px 20px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.04);
    }
    [data-testid="stMetricLabel"] { font-size: 13px !important; }
    [data-testid="stMetricValue"] { font-size: 24px !important; font-weight: 800 !important; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] { padding: 8px 20px; font-weight: 600; }
    div[data-testid="stSidebarContent"] { background: #FBF9F1; }
    h1, h2, h3 { color: #330C2E !important; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# 상수
# ============================================================
NAME_NORMALIZE = {
    '사적인아름다운지유의원': '사적인아름다움지유의원',
    '루호성형외과': '루호성형외과의원',
    '우리성형외과': '우리성형외과의원',
    '테이아 의원': '테이아의원',
    '티에스성형외과의원': '티에스성형외과',
    '톡스앤필-시논현': '톡스앤필의원-신논현점',
    '톡스앤필 - 신논': '톡스앤필의원-신논현점',
    '톡스앤필 - 신논현': '톡스앤필의원-신논현점',
    '제이필 - 홍대': '제이필의원-홍대점',
    '제이필 - 강남': '제이필의원-강남점',
    '플래저성형외과': '플레저성형외과의원',
    '플래너성형외과': '플래너성형외과의원',
    '유픽의원 홍대': '유픽의원-홍대점',
    '유픽의원-홍대': '유픽의원-홍대점',
    '유픽의원-강남': '유픽의원-강남점',
    '디어 청담의원': '청담디어의원',
    '홍대셀레나': '홍대셀레나의원',
    '히트성형외과': '히트성형외과의원',
}

# 어드민 표기(20종) + 운영시트 내부 중복 표기를 같은 정규명으로 흡수.
# 두 원천을 이어붙일 때 같은 병원이 두 줄로 갈라지는 것을 막는다.
NAME_NORMALIZE.update(SHEET_HOSPITAL_ALIASES)
NAME_NORMALIZE.update(ADMIN_HOSPITAL_ALIASES)

COUNTRY_FLAG = {
    '태국': '🇹🇭', '대만': '🇹🇼', '중국': '🇨🇳', '미국': '🇺🇸',
    '호주': '🇦🇺', '일본': '🇯🇵', '홍콩': '🇭🇰', '싱가포르': '🇸🇬',
    '베트남': '🇻🇳', '필리핀': '🇵🇭', '말레이시아': '🇲🇾', '인도네시아': '🇮🇩',
    '영국': '🇬🇧', '프랑스': '🇫🇷', '독일': '🇩🇪', '캐나다': '🇨🇦',
    '인도': '🇮🇳', '러시아': '🇷🇺', '몽골': '🇲🇳', '폴란드': '🇵🇱',
    '캄보디아': '🇰🇭', '아일랜드': '🇮🇪', '키르기스스탄': '🇰🇬',
    '부탄': '🇧🇹', '스페인': '🇪🇸', '뉴질랜드': '🇳🇿',
}


# 시술 키워드 매핑 (자유 텍스트 → 표준 시술 카테고리)
PROCEDURE_KEYWORDS = [
    ('울쎄라', ['울쎄라', 'ulthera', '울쎼라']),
    ('보톡스', ['보톡스', 'botox', '보톡', '더마톡신']),
    ('포텐자', ['포텐자', 'potenza']),
    ('써마지', ['써마지', 'thermage']),
    ('슈링크', ['슈링크', 'shrink']),
    ('리쥬란', ['리쥬란', 'rejuran']),
    ('쥬베룩', ['쥬베룩', 'juvelook']),
    ('필러', ['필러', 'filler']),
    ('리프팅', ['리프팅', 'lifting', '거상', '실리프팅']),
    ('지방분해주사', ['지방분해', '윤곽주사', '커팅주사', 'fat dissolv']),
    ('레이저', ['레이저', 'laser', '피코', 'pico', 'BBL']),
    ('올리지오', ['올리지오', 'oligio']),
    ('온다', ['온다', 'onda']),
    ('물광주사', ['물광', '더마샤인', 'skinbooster']),
    ('스킨보톡스', ['스킨보톡스', '스킨보', 'skinbtx']),
    ('인모드', ['인모드', 'inmode']),
    ('눈수술', ['눈수술', '눈매교정', '쌍꺼풀', '상안검', '눈 수술', '눈재수술']),
    ('코수술', ['코수술', '코끝', '콧대', '코 수술', '코 첫수술']),
    ('소프웨이브', ['소프웨이브', 'sofwave']),
    ('엑소좀', ['엑소좀', 'exosome']),
    ('아쿠아필', ['아쿠아필', 'aquapeel', '아쿠아 필']),
    ('셀르디엠', ['셀르디엠', 'cellrdm']),
    ('수액', ['수액']),
    ('모공치료', ['모공']),
    ('여드름치료', ['여드름']),
]


def extract_procedures(text):
    """자유 텍스트에서 시술 키워드를 추출하여 리스트로 반환"""
    if not text or str(text).strip() in ('', 'nan'):
        return []
    text_lower = str(text).lower()
    found = []
    for label, keywords in PROCEDURE_KEYWORDS:
        for kw in keywords:
            if kw.lower() in text_lower:
                found.append(label)
                break
    return found if found else [str(text).strip()[:30]]


def normalize_hospital(name):
    if not name or str(name).strip() in ('', 'nan', 'None'):
        return None
    name = str(name).strip()
    return NAME_NORMALIZE.get(name, name)


def format_krw(amount):
    if pd.isna(amount) or amount == 0:
        return '0원'
    amount = float(amount)
    if amount >= 100_000_000:
        return f"{amount / 100_000_000:.1f}억원"
    elif amount >= 10_000:
        return f"{amount / 10_000:,.0f}만원"
    else:
        return f"{amount:,.0f}원"


# ============================================================
# 데이터 로딩
# ============================================================
def finalize_reservations(df_res, source):
    """예약 raw(운영시트/어드민 공통)에 파생 컬럼을 붙인다.

    두 원천이 같은 타입·같은 표기로 정규화되어야 이어붙였을 때 집계가 깨지지 않는다.
    """
    df_res = df_res.copy()
    df_res['병원명'] = df_res['예약클리닉'].apply(normalize_hospital)
    df_res['내원일'] = pd.to_datetime(df_res['내원일'], errors='coerce')
    df_res['월'] = df_res['내원일'].dt.to_period('M').astype(str)
    df_res['실제금액'] = pd.to_numeric(df_res['실제금액'], errors='coerce').fillna(0)
    for col in ('고객국적', '종류', '시술수술명', '예약상태'):
        df_res[col] = df_res[col].apply(lambda x: str(x).strip() if pd.notna(x) else '')
    if '_source' not in df_res.columns:
        df_res['_source'] = source
    return df_res


@st.cache_data(show_spinner=False, ttl=600)
def load_operation_excel(file_path):
    """운영 트렌드 Excel → 예약완료 df + 정산 df + 전체예약 df"""

    # URL인 경우 다운로드
    if str(file_path).startswith('http'):
        import urllib.request
        tmp = '/tmp/unniguide_gsheet_op.xlsx'
        urllib.request.urlretrieve(file_path, tmp)
        file_path = tmp

    # 예약 시트
    # 시트 이름 호환: 로컬 Excel vs Google Sheets
    xls = pd.ExcelFile(file_path)
    res_sheet = None
    for name in xls.sheet_names:
        if '예약확정' in name:
            res_sheet = name
            break
    if res_sheet is None:
        raise ValueError(f"예약확정 시트를 찾을 수 없습니다. 시트 목록: {xls.sheet_names}")
    # 헤더 행 자동 탐지 + 이름 기반 컬럼 매핑 (시트 컬럼 순서/헤더 위치 변동에 robust)
    def _norm_header(c):
        return str(c).replace(' ', '').replace('\n', '').strip()

    raw_head = pd.read_excel(file_path, sheet_name=res_sheet, header=None, nrows=5)
    header_idx = None
    for i in range(len(raw_head)):
        vals = [_norm_header(v) for v in raw_head.iloc[i].tolist()]
        if any('예약상태' in v for v in vals) and any('고객명' in v for v in vals):
            header_idx = i
            break
    if header_idx is None:
        header_idx = 1  # 과거 포맷 fallback

    df_res = pd.read_excel(file_path, sheet_name=res_sheet, header=header_idx)
    name_map = [  # (표준명, 매칭 키워드) — 순서 중요: 구체적인 것 먼저
        ('시술수술명', '시술/수술명'), ('시술수술명', '시술수술명'),
        ('예약클리닉', '예약클리닉'), ('추천클리닉', '추천클리닉'),
        ('예약상태', '예약상태'), ('고객국적', '고객국적'),
        ('실제금액', '실제금액'), ('예상금액', '예상금액'),
        ('내원일', '내원일'), ('예약확정일', '예약확정일'),
        ('채팅접수일자', '채팅접수일자'), ('사용언어', '사용언어'),
        ('고객명', '고객명'), ('종류', '종류'), ('시간', '시간'),
    ]
    rename, used = {}, set()
    for col in df_res.columns:
        n = _norm_header(col)
        for std, key in name_map:
            if std not in used and key in n:
                rename[col] = std
                used.add(std)
                break
    df_res = df_res.rename(columns=rename)

    required = ['예약상태', '예약클리닉', '고객국적', '내원일', '실제금액', '종류', '시술수술명']
    missing = [c for c in required if c not in df_res.columns]
    if missing:
        raise ValueError(f"예약확정 시트에서 필수 컬럼을 찾을 수 없습니다: {missing} (헤더 행 {header_idx + 1}행 기준)")
    df_res = finalize_reservations(df_res, source='운영시트')

    df_completed = df_res[df_res['예약상태'] == '시/수술 완료'].copy()
    df_all = df_res.copy()  # 전체 (취소/노쇼 포함)

    # 정산 시트 (openpyxl)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    # 시트 이름 호환: 로컬 Excel vs Google Sheets
    settle_sheet = None
    for name in wb.sheetnames:
        if '정산' in name:
            settle_sheet = name
            break
    if settle_sheet is None:
        raise ValueError(f"정산 시트를 찾을 수 없습니다. 시트 목록: {wb.sheetnames}")
    ws2 = wb[settle_sheet]
    settlement_records = []
    current_month = None
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=False):
        vals = [c.value for c in row]
        a_val = str(vals[0]).strip() if vals[0] else ''
        if '정산 내역' in a_val:
            parts = a_val.replace('년', '-').replace('월', '').replace('정산 내역', '').strip()
            try:
                year, month = parts.split('-')[:2]
                current_month = f"{year.strip()}-{int(month.strip()):02d}"
            except Exception:
                pass
            continue
        if a_val in ('NO', 'NO.', '', 'None', '재무팀 정산 요청 내역') or '정산 요청일' in a_val:
            continue
        if current_month and vals[1]:
            hospital = str(vals[1]).strip()
            if hospital in ('병원명', ''):
                continue

            def _to_num(v):
                if v is None or v == '':
                    return None
                if isinstance(v, (int, float)):
                    return float(v)
                try:
                    return float(str(v).replace('₩', '').replace(',', '').strip())
                except ValueError:
                    return None

            # 구분(시술/수술 텍스트)과 시술금액(숫자) 칸이 일부 월에서 바뀌어 입력됨
            # → 숫자인 칸 = 금액, 텍스트인 칸 = 구분 으로 자동 판별
            c6, c7 = vals[6], vals[7]
            n6, n7 = _to_num(c6), _to_num(c7)
            if n6 is not None and n7 is None:
                amount, kind = n6, str(c7).strip() if c7 else ''
            elif n7 is not None and n6 is None:
                amount, kind = n7, str(c6).strip() if c6 else ''
            else:
                # 판별 불가 시 헤더 순서(구분=6, 금액=7) 사용
                amount = n7 if n7 is not None else 0
                kind = str(c6).strip() if c6 else ''
            settlement_records.append({
                '정산월': current_month,
                '병원명': normalize_hospital(hospital),
                '고객명': str(vals[2]).strip() if vals[2] else '',
                '국적': str(vals[4]).strip() if vals[4] else '',
                '구분': kind,
                '시술금액': amount,
                '수수료금액': _to_num(vals[8]) or 0,
            })
    wb.close()
    df_settle = pd.DataFrame(settlement_records)

    return df_completed, df_settle, df_all


def _agg_month_series(df):
    """집계용 월(YYYY-MM): 내원일 기준, 없으면 예약확정일로 보완."""
    m = pd.to_datetime(df['내원일'], errors='coerce')
    if '예약확정일' in df.columns:
        m = m.fillna(pd.to_datetime(df['예약확정일'], errors='coerce'))
    return m.dt.to_period('M').astype(str)


@st.cache_data(show_spinner=False, ttl=600)
def load_admin_reservations(path):
    """어드민 case-metrics CSV 파일 → 예약완료 df + 전체예약 df (운영시트와 동일 스키마)"""
    return _split_admin(load_admin_case_metrics(path))


@st.cache_data(show_spinner=False, ttl=600)
def load_admin_from_gsheet(file_path):
    """운영 구글시트의 어드민 raw 탭 → 예약완료 df + 전체예약 df.

    고객 실명·금액이 담긴 원천이라 공개 repo에 두지 않고, 기존 운영시트와
    같은 접근통제 아래(같은 문서의 별도 탭) 관리한다.
    탭이 없으면 (None, None)을 돌려 호출부가 다른 경로로 넘어가게 한다.
    """
    if str(file_path).startswith('http'):
        import urllib.request
        tmp = '/tmp/unniguide_gsheet_op.xlsx'
        urllib.request.urlretrieve(file_path, tmp)
        file_path = tmp

    xls = pd.ExcelFile(file_path)
    tab = next((n for n in xls.sheet_names if ADMIN_SHEET_KEYWORD in n), None)
    if tab is None:
        return None, None
    df = pd.read_excel(file_path, sheet_name=tab, dtype=str)
    return _split_admin(normalize_admin_frame(df))


def _split_admin(df_admin):
    df_res = finalize_reservations(df_admin, source='어드민')
    df_completed = df_res[df_res['예약상태'] == '시/수술 완료'].copy()
    return df_completed, df_res


def merge_sources(op_completed, op_all, op_settle, ad_completed, ad_all,
                  cutoff=ADMIN_CUTOFF_MONTH):
    """운영시트(컷오프 전) + 어드민(컷오프 후)을 내원일 기준으로 이어붙인다.

    2026년 7월 운영이 어드민으로 이관되면서 운영시트 입력이 2026-07-07에 멈췄다.
    같은 달을 두 원천에서 세면 중복되므로, 내원 월을 기준으로 한쪽만 채택한다.
      - 집계월 < cutoff  → 운영시트 (어드민은 4월 14% / 5월 87% / 6월 94%만 커버)
      - 집계월 >= cutoff → 어드민   (운영시트는 7월 상순만 존재)
    반환: (df_completed, df_settle, df_all, seam_info)
    """
    if ad_all is None or len(ad_all) == 0:
        return op_completed, op_settle, op_all, None

    def _split(df, keep_before):
        m = _agg_month_series(df)
        return df[(m < cutoff) if keep_before else (m >= cutoff)].copy()

    op_c, op_a = _split(op_completed, True), _split(op_all, True)
    ad_c, ad_a = _split(ad_completed, False), _split(ad_all, False)

    df_completed = pd.concat([op_c, ad_c], ignore_index=True, sort=False)
    df_all = pd.concat([op_a, ad_a], ignore_index=True, sort=False)

    # 정산: 컷오프 이후는 정산 시트에 아직 없으므로 어드민 완료건에서 직접 산출
    #       (수수료 = 시술 10% / 수술 20% — 정산 시트와 동일 규칙)
    settle_parts = []
    if op_settle is not None and len(op_settle) > 0:
        settle_parts.append(op_settle[op_settle['정산월'] < cutoff].copy())
    if len(ad_c) > 0:
        ad_settle = pd.DataFrame({
            '정산월': _agg_month_series(ad_c),
            '병원명': ad_c['병원명'],
            '고객명': ad_c['고객명'],
            '국적': ad_c['고객국적'],
            '구분': ad_c['종류'],
            '시술금액': ad_c['실제금액'],
        })
        ad_settle['수수료금액'] = ad_settle.apply(
            lambda r: _commission(r['시술금액'], r['구분']), axis=1)
        settle_parts.append(ad_settle[ad_settle['시술금액'] > 0])
    df_settle = (pd.concat(settle_parts, ignore_index=True, sort=False)
                 if settle_parts else pd.DataFrame())

    seam_info = {
        'cutoff': cutoff,
        'sheet_months': sorted(m for m in op_a['월'].dropna().unique() if m != 'NaT'),
        'admin_months': sorted(m for m in ad_a['월'].dropna().unique() if m != 'NaT'),
        'sheet_completed': len(op_c),
        'admin_completed': len(ad_c),
        'admin_settle_generated': len(ad_c),
    }
    return df_completed, df_settle, df_all, seam_info


def _commission(amount, kind):
    """수수료 = 시술 10% / 수술 20% (그 외 0). fill_settlement.py 규칙과 동일."""
    kind = str(kind).strip()
    if kind == '시술':
        return float(amount) * 0.10
    if kind == '수술':
        return float(amount) * 0.20
    return 0.0


def load_internal_report(df_completed, df_all, df_settle):
    """예약확정 raw에서 월별트렌드/병원별성과/취소노쇼 요약·상세를 직접 계산.
    (기존: 내부리포트 시트를 읽음 → 현재: raw에서 자동 집계, 별도 시트 불필요)"""

    comp = df_completed.copy()
    comp['_월'] = _agg_month_series(comp)
    comp['_수수료'] = comp.apply(lambda r: _commission(r['실제금액'], r['종류']), axis=1)
    comp['_시술'] = (comp['종류'] == '시술').astype(int)
    comp['_수술'] = (comp['종류'] == '수술').astype(int)
    comp_v = comp[comp['_월'] != 'NaT'].copy()

    # 취소/노쇼 마스크 (전체 예약 기준)
    allr = df_all.copy()
    allr['_월'] = _agg_month_series(allr)
    status = allr['예약상태'].fillna('').astype(str).str.strip()
    is_cancel = status == '예약 취소'
    is_noshow = status.str.lower().str.contains('no-show|no show|noshow', na=False, regex=True)
    is_completed = status == '시/수술 완료'

    # ---------- 1) 월별 트렌드 ----------
    if len(comp_v) > 0:
        g = comp_v.groupby('_월')
        df_monthly = pd.DataFrame({
            '월': list(g.groups.keys()),
            '완료건수': g.size().values,
            '시술건수': g['_시술'].sum().values,
            '수술건수': g['_수술'].sum().values,
            '시수술금액': g['실제금액'].sum().values,
            '수수료매출': g['_수수료'].sum().values,
        })
    else:
        df_monthly = pd.DataFrame(columns=['월', '완료건수', '시술건수', '수술건수', '시수술금액', '수수료매출'])
    # 월별 취소+노쇼
    cn = allr[(is_cancel | is_noshow) & (allr['_월'] != 'NaT')]
    cn_by_month = cn.groupby('_월').size() if len(cn) > 0 else pd.Series(dtype=int)
    df_monthly['취소노쇼'] = df_monthly['월'].map(cn_by_month).fillna(0)
    df_monthly['평균객단가'] = (df_monthly['시수술금액'] / df_monthly['완료건수'].replace(0, pd.NA)).fillna(0)
    df_monthly = df_monthly.sort_values('월').reset_index(drop=True)
    for col in ['완료건수', '시술건수', '수술건수', '시수술금액', '수수료매출', '평균객단가', '취소노쇼']:
        df_monthly[col] = pd.to_numeric(df_monthly[col], errors='coerce').fillna(0)

    # ---------- 2) 병원별 성과 (누적) ----------
    comp_h = comp[comp['병원명'].notna()].copy()
    if len(comp_h) > 0:
        hg = comp_h.groupby('병원명')
        df_hosp = pd.DataFrame({
            '병원명': list(hg.groups.keys()),
            '누적건수': hg.size().values,
            '누적시수술금액': hg['실제금액'].sum().values,
            '누적수수료': hg['_수수료'].sum().values,
        })
        # 최신월 (집계 가능한 월 중 최댓값)
        valid_months = comp_h[comp_h['_월'] != 'NaT']['_월']
        latest = valid_months.max() if len(valid_months) > 0 else None
        if latest:
            lm = comp_h[comp_h['_월'] == latest].groupby('병원명').agg(
                최신월건수=('실제금액', 'size'), 최신월금액=('실제금액', 'sum')).reset_index()
            df_hosp = df_hosp.merge(lm, on='병원명', how='left')
        else:
            df_hosp['최신월건수'] = 0
            df_hosp['최신월금액'] = 0
        df_hosp[['최신월건수', '최신월금액']] = df_hosp[['최신월건수', '최신월금액']].fillna(0)
        df_hosp['전월대비'] = 0.0
        df_hosp = df_hosp.sort_values('누적시수술금액', ascending=False).reset_index(drop=True)
        df_hosp['순위'] = range(1, len(df_hosp) + 1)
        df_hosp = df_hosp[['순위', '병원명', '누적건수', '누적시수술금액', '누적수수료', '최신월건수', '최신월금액', '전월대비']]
    else:
        df_hosp = pd.DataFrame(columns=['순위', '병원명', '누적건수', '누적시수술금액', '누적수수료', '최신월건수', '최신월금액', '전월대비'])

    # ---------- 3) 취소/노쇼 요약 ----------
    total_cancel = int(is_cancel.sum())
    total_noshow = int(is_noshow.sum())
    denom = int((is_completed | is_cancel | is_noshow).sum())
    rate = ((total_cancel + total_noshow) / denom * 100) if denom > 0 else 0
    cancel_summary = {
        'total_cancel': total_cancel,
        'total_noshow': total_noshow,
        'cancel_rate': f"{rate:.1f}%",
    }

    # ---------- 4) 병원별 취소/노쇼 ----------
    allr['_취소'] = is_cancel.astype(int)
    allr['_노쇼'] = is_noshow.astype(int)
    allr['_유효'] = (is_completed | is_cancel | is_noshow).astype(int)
    ch = allr[allr['병원명'].notna()].groupby('병원명').agg(
        전체예약=('_유효', 'sum'), 취소=('_취소', 'sum'), **{'No-show': ('_노쇼', 'sum')}).reset_index()
    ch = ch[(ch['취소'] + ch['No-show']) > 0].copy()
    ch['취소노쇼율'] = ((ch['취소'] + ch['No-show']) / ch['전체예약'].replace(0, pd.NA)).fillna(0)
    df_cancel_hospital = ch[['병원명', '전체예약', '취소', 'No-show', '취소노쇼율']].reset_index(drop=True)

    # ---------- 5) 취소/노쇼 상세 ----------
    det = allr[is_cancel | is_noshow].copy()
    df_cancel_detail = pd.DataFrame({
        '월': det['_월'].values,
        '상태': status[is_cancel | is_noshow].values,
        '병원명': det['병원명'].values,
        '국적': det['고객국적'].values if '고객국적' in det.columns else '',
        '고객명': det['고객명'].values if '고객명' in det.columns else '',
        '종류': det['종류'].values if '종류' in det.columns else '',
        '시술수술명': det['시술수술명'].values if '시술수술명' in det.columns else '',
    })
    df_cancel_detail['병원명'] = df_cancel_detail['병원명'].fillna('')

    return df_monthly, df_hosp, cancel_summary, df_cancel_hospital, df_cancel_detail


# ============================================================
# Google Sheets 설정
# ============================================================
# 통합 운영 시트 (예약확정 리스트 + 정산 요청 리스트). 내부리포트는 이 raw에서 자동 계산됨.
GSHEET_ID_MAIN = "1MNG3lIL2P-DIydra7vatnaKUq8MLKI0xm9YSiD8uYFk"
GID_RESERVATION = 123775075   # 예약확정 시트
GID_SETTLEMENT = 622724794    # 내부리포트 (정산 포함)
GID_OFFLINE = 1126075757      # 오프라인 데일리
GID_DASHBOARD = 1029704191    # 가공 대시보드


def gsheet_xlsx_url(sheet_id):
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


def gsheet_csv_url(sheet_id, gid):
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


# ============================================================
# 어드민 원천 (2026-07 이후)
# ============================================================
# ⚠️ 이 repo는 공개다. 어드민 원본에는 고객 실명·나이·성별이 행 단위로 있어
#    그대로 커밋할 수 없다. repo에 두는 파일은 deidentify_admin.py로
#    개인식별정보를 제거한 버전이며, 고객명이 필요한 화면(취소/노쇼 상세, 정산 명세)은
#    비게 된다. 실명까지 필요하면 운영시트에 '어드민 케이스 raw' 탭을 만들면
#    그쪽을 우선 사용한다(ADMIN_SOURCE.md 참고).
ADMIN_CSV_LOCAL_GLOB = '~/Downloads/case-metrics-*.csv'
ADMIN_CSV_REPO = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              'data', 'admin_case_metrics.csv')


def resolve_admin_csv():
    """어드민 CSV 경로: 로컬 Downloads 최신 export(실명 포함) → repo 비식별본."""
    local = sorted(glob.glob(os.path.expanduser(ADMIN_CSV_LOCAL_GLOB)),
                   key=os.path.getmtime, reverse=True)
    if local:
        return local[0]
    return ADMIN_CSV_REPO if os.path.exists(ADMIN_CSV_REPO) else None


# ============================================================
# 사이드바
# ============================================================
with st.sidebar:
    st.markdown(f"""
    <div style="text-align:center; padding: 12px 0 20px;">
        <span style="font-size:28px; font-weight:800; color:{BRAND_PLUM};">UNNI</span>
        <span style="font-size:28px; font-weight:400; color:{BRAND_PLUM};"> GUIDE</span>
        <br><span style="font-size:13px; color:{BRAND_ORANGE}; font-weight:600;">운영 대시보드</span>
    </div>
    """, unsafe_allow_html=True)
    st.divider()

    data_source = st.radio("데이터 소스", ["Google Sheets (자동)", "로컬 파일"], index=0)

    df_completed = df_settle = df_all = None
    df_monthly = df_hosp_perf = cancel_summary = df_cancel_hospital = df_cancel_detail = None
    seam_info = None

    if data_source == "Google Sheets (자동)":
        st.caption("Google Sheets에서 자동으로 데이터를 불러옵니다.")
        if st.button("데이터 새로고침", type="primary"):
            st.cache_data.clear()

        try:
            with st.spinner("운영 데이터 로딩 중..."):
                xlsx_url = gsheet_xlsx_url(GSHEET_ID_MAIN)
                df_completed, df_settle, df_all = load_operation_excel(xlsx_url)
            st.success("운영 데이터 로드 완료")
        except Exception as e:
            st.error(f"운영 데이터 로드 실패: {str(e)[:50]}")

        # 2026-07 이후 내원분은 어드민 원천으로 교체
        # 1순위: 운영시트의 어드민 raw 탭 (배포본) → 2순위: 로컬 Downloads CSV
        if df_completed is not None:
            try:
                with st.spinner("어드민 데이터 병합 중..."):
                    ad_completed, ad_all = load_admin_from_gsheet(xlsx_url)
                    origin = f"시트 '{ADMIN_SHEET_KEYWORD}' 탭"
                    if ad_all is None:
                        admin_csv = resolve_admin_csv()
                        if admin_csv:
                            ad_completed, ad_all = load_admin_reservations(admin_csv)
                            origin = os.path.basename(admin_csv)
                    if ad_all is not None:
                        df_completed, df_settle, df_all, seam_info = merge_sources(
                            df_completed, df_all, df_settle, ad_completed, ad_all)
                if ad_all is not None:
                    st.success(f"어드민 병합 완료 ({origin})")
                else:
                    st.warning(
                        f"어드민 원천을 찾을 수 없어 {ADMIN_CUTOFF_MONTH} 이후가 비어 있습니다. "
                        f"운영시트에 '{ADMIN_SHEET_KEYWORD} raw' 탭을 만들어 "
                        f"어드민 case-metrics CSV를 붙여넣어 주세요.")
            except Exception as e:
                st.warning(f"어드민 병합 실패 (운영시트만 사용): {str(e)[:60]}")

        try:
            if df_completed is not None:
                with st.spinner("내부리포트 집계 중..."):
                    df_monthly, df_hosp_perf, cancel_summary, df_cancel_hospital, df_cancel_detail = load_internal_report(df_completed, df_all, df_settle)
                st.success("내부리포트 집계 완료")
        except Exception as e:
            st.warning(f"내부리포트 집계 실패: {str(e)[:50]}")

    else:
        # 로컬 파일 모드 (기존 방식)
        st.markdown("**1. 운영 트렌드 데이터**")
        pattern1 = os.path.expanduser('~/Downloads/언니가이드 운영 트렌드 데이터_*.xlsx')
        pattern1b = os.path.expanduser('~/Desktop/언니가이드_리포트/언니가이드 운영 트렌드 데이터_*.xlsx')
        op_candidates = sorted(glob.glob(pattern1) + glob.glob(pattern1b), key=os.path.getmtime, reverse=True)

        if op_candidates:
            op_file = st.selectbox("운영 Excel", op_candidates, format_func=os.path.basename, key="op")
            with st.spinner("운영 데이터 로딩..."):
                df_completed, df_settle, df_all = load_operation_excel(op_file)
        else:
            op_upload = st.file_uploader("운영 Excel 업로드", type=['xlsx'], key="op_up")
            if op_upload:
                tmp = "/tmp/unniguide_op.xlsx"
                with open(tmp, "wb") as f:
                    f.write(op_upload.getvalue())
                with st.spinner("운영 데이터 로딩..."):
                    df_completed, df_settle, df_all = load_operation_excel(tmp)

        st.markdown(f"**2. 어드민 데이터 ({ADMIN_CUTOFF_MONTH}~)**")
        ad_candidates = sorted(glob.glob(os.path.expanduser(ADMIN_CSV_LOCAL_GLOB)),
                               key=os.path.getmtime, reverse=True)
        admin_csv = None
        if ad_candidates:
            admin_csv = st.selectbox("어드민 CSV", ad_candidates,
                                     format_func=os.path.basename, key="ad")
        else:
            ad_upload = st.file_uploader("어드민 CSV 업로드", type=['csv'], key="ad_up")
            if ad_upload:
                admin_csv = "/tmp/unniguide_admin.csv"
                with open(admin_csv, "wb") as f:
                    f.write(ad_upload.getvalue())

        if df_completed is not None and admin_csv:
            with st.spinner("어드민 병합..."):
                ad_completed, ad_all = load_admin_reservations(admin_csv)
                df_completed, df_settle, df_all, seam_info = merge_sources(
                    df_completed, df_all, df_settle, ad_completed, ad_all)

        # 내부 리포트: 운영 데이터(raw)에서 자동 계산 (별도 파일 불필요)
        if df_completed is not None:
            with st.spinner("내부리포트 집계..."):
                df_monthly, df_hosp_perf, cancel_summary, df_cancel_hospital, df_cancel_detail = load_internal_report(df_completed, df_all, df_settle)

# ============================================================
# 데이터 없으면 안내
# ============================================================
if df_completed is None:
    st.markdown(f"""
    <div style="text-align:center; padding:80px 0;">
        <div style="font-size:48px; margin-bottom:16px;">📊</div>
        <h2>언니가이드 운영 대시보드</h2>
        <p style="color:#636E72; font-size:16px; margin-top:8px;">
            왼쪽 사이드바에서 데이터 파일을 선택하거나 업로드해주세요.
        </p>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ============================================================
# 필터
# ============================================================
def _valid_months(values):
    """'월'은 Period→문자열이라 내원일이 빈 행이 'NaT' 문자열로 남는다.
    dropna()로는 걸러지지 않아 월 선택 옵션에 섞여 들어가므로 명시적으로 제외한다."""
    return sorted({str(v) for v in values
                   if pd.notna(v) and re.fullmatch(r'\d{4}-\d{2}', str(v))})


all_months_res = _valid_months(df_completed['월'])
all_months_settle = _valid_months(df_settle['정산월']) if len(df_settle) > 0 else []
all_months = sorted(set(all_months_res + all_months_settle))

if not all_months:
    st.error("데이터에 유효한 월 정보가 없습니다.")
    st.stop()

with st.sidebar:
    st.divider()
    st.subheader("필터")
    if len(all_months) >= 2:
        # 기본 끝점 = 지난달(마지막으로 마감된 달).
        # 어드민에는 미래 내원일 예약이 들어있어 최신월을 그대로 쓰면
        # 진행 중/미래 달이 기본값이 되어 실적이 급락한 것처럼 보인다.
        last_closed = str((pd.Timestamp.today().to_period('M') - 1))
        past = [m for m in all_months if m <= last_closed]
        default_end = past[-1] if past else all_months[-1]
        month_range = st.select_slider("기간 선택", options=all_months, value=(all_months[0], default_end))
        selected_months = [m for m in all_months if month_range[0] <= m <= month_range[1]]
    else:
        selected_months = all_months
        month_range = (all_months[0], all_months[0])

    all_nationalities = sorted([n for n in df_completed['고객국적'].unique() if n and n != 'nan'])
    selected_nationalities = st.multiselect("국적", all_nationalities, default=[], placeholder="전체 국적")

    all_hospitals = sorted(df_completed['병원명'].dropna().unique())
    selected_hospitals = st.multiselect("병원", all_hospitals, default=[], placeholder="전체 병원")
    st.divider()
    st.caption(f"예약완료 {len(df_completed):,}건 | 정산 {len(df_settle):,}건")

    # 원천 구간 안내 — 어느 달을 어느 데이터로 세고 있는지 항상 보이게 둔다
    if seam_info:
        sm, am = seam_info['sheet_months'], seam_info['admin_months']
        st.markdown(f"""
        <div style="background:#F8F5FA; border-left:3px solid {BRAND_PLUM};
             padding:10px 12px; border-radius:4px; font-size:11.5px; line-height:1.7;">
            <b>데이터 원천</b><br>
            📗 운영시트 · {sm[0] if sm else '-'} ~ {sm[-1] if sm else '-'}
            <span style="color:#888;">({seam_info['sheet_completed']:,}건 완료)</span><br>
            📘 어드민 · {am[0] if am else '-'} ~ {am[-1] if am else '-'}
            <span style="color:#888;">({seam_info['admin_completed']:,}건 완료)</span><br>
            <span style="color:#888;">{seam_info['cutoff']} 내원분부터 어드민 기준.
            해당 월 정산은 실제금액 × 시술10%/수술20%로 자동 산출.</span>
        </div>
        """, unsafe_allow_html=True)

# 필터 적용
mask_res = df_completed['월'].isin(selected_months)
mask_set = df_settle['정산월'].isin(selected_months) if len(df_settle) > 0 else pd.Series(dtype=bool)
if selected_nationalities:
    mask_res = mask_res & df_completed['고객국적'].isin(selected_nationalities)
    if len(df_settle) > 0:
        mask_set = mask_set & df_settle['국적'].isin(selected_nationalities)
if selected_hospitals:
    mask_res = mask_res & df_completed['병원명'].isin(selected_hospitals)
    if len(df_settle) > 0:
        mask_set = mask_set & df_settle['병원명'].isin(selected_hospitals)

filtered_res = df_completed[mask_res].copy()
filtered_set = df_settle[mask_set].copy() if len(df_settle) > 0 else pd.DataFrame()

# ============================================================
# 헤더
# ============================================================
period_label = f"{month_range[0]} ~ {month_range[1]}" if month_range[0] != month_range[1] else month_range[0]
st.markdown(f"""
<div style="background: linear-gradient(135deg, {BRAND_ORANGE} 0%, #E8551F 100%);
     color: white; padding: 28px 32px; border-radius: 0 0 20px 20px; margin: -1rem -1rem 24px -1rem;">
    <div style="display:flex; justify-content:space-between; align-items:center;">
        <div>
            <div style="font-size:14px; opacity:0.85; margin-bottom:4px;">UNNI GUIDE 운영 대시보드</div>
            <div style="font-size:24px; font-weight:800;">{period_label} 데이터</div>
        </div>
        <div style="font-size:13px; opacity:0.75;">예약 {len(filtered_res):,}건 · 정산 {len(filtered_set):,}건</div>
    </div>
</div>
""", unsafe_allow_html=True)

# ============================================================
# 탭
# ============================================================
tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "📊 Overview", "🌍 국적 분석", "🏥 병원 분석", "💉 시술 트렌드", "⚠️ 취소/No-show", "📋 원본 데이터", "📦 리포트 생성"
])

# 공통: 최신월/전월
latest_month = selected_months[-1] if selected_months else None
prev_candidates = [m for m in all_months if m < latest_month] if latest_month else []
prev_month = prev_candidates[-1] if prev_candidates else None

# ============================================================
# Tab 1: Overview (필터 반응 단일 구조)
# ============================================================
with tab1:
    def pct_delta(cur, prev):
        return f"{(cur - prev) / max(prev, 1) * 100:+.1f}%" if prev > 0 else None

    # 현재 필터 표시
    filter_desc_parts = []
    if month_range[0] != month_range[1]:
        filter_desc_parts.append(f"기간: {month_range[0]} ~ {month_range[1]}")
    else:
        filter_desc_parts.append(f"기간: {month_range[0]}")
    if selected_nationalities:
        filter_desc_parts.append(f"국적: {', '.join(selected_nationalities[:3])}{'...' if len(selected_nationalities) > 3 else ''}")
    if selected_hospitals:
        filter_desc_parts.append(f"병원: {', '.join(selected_hospitals[:2])}{'...' if len(selected_hospitals) > 2 else ''}")
    filter_label = " · ".join(filter_desc_parts)

    st.subheader("📊 성과 Overview")
    st.caption(f"{filter_label}")

    # 필터된 데이터 KPI
    f_cnt = len(filtered_set) if len(filtered_set) > 0 else len(filtered_res)
    f_rev = filtered_set['시술금액'].sum() if len(filtered_set) > 0 else filtered_res['실제금액'].sum()
    f_comm = filtered_set['수수료금액'].sum() if len(filtered_set) > 0 else 0
    f_avg = f_rev / f_cnt if f_cnt > 0 else 0
    f_nat = filtered_set['국적'].nunique() if len(filtered_set) > 0 else filtered_res['고객국적'].nunique()
    f_hosp = filtered_set['병원명'].nunique() if len(filtered_set) > 0 else filtered_res['병원명'].nunique()

    fc1, fc2, fc3, fc4, fc5, fc6 = st.columns(6)
    fc1.metric("완료 건수", f"{f_cnt:,}건")
    fc2.metric("정산 매출", format_krw(f_rev))
    fc3.metric("수수료", format_krw(f_comm))
    fc4.metric("객단가", format_krw(f_avg))
    fc5.metric("국적 수", f"{f_nat}개국")
    fc6.metric("병원 수", f"{f_hosp}개")

    st.markdown("")

    # --- 월별 트렌드 차트 (필터 반응) ---
    st.subheader("월별 트렌드")
    st.caption("선택 필터가 적용된 월별 추이입니다.")

    # 필터된 데이터로 월별 집계
    if len(filtered_set) > 0:
        monthly_filtered = filtered_set.groupby('정산월').agg(
            건수=('시술금액', 'count'), 매출=('시술금액', 'sum'), 수수료=('수수료금액', 'sum'),
        ).reset_index().sort_values('정산월')
        monthly_filtered['누적건수'] = monthly_filtered['건수'].cumsum()
        m_labels = monthly_filtered['정산월'].tolist()
        m_counts = monthly_filtered['건수'].tolist()
        m_revenues = monthly_filtered['매출'].tolist()
        m_commissions = monthly_filtered['수수료'].tolist()
        m_cum = monthly_filtered['누적건수'].tolist()
    else:
        monthly_filtered_r = filtered_res.groupby('월').agg(건수=('월', 'count'), 매출=('실제금액', 'sum')).reset_index().sort_values('월')
        monthly_filtered_r['누적건수'] = monthly_filtered_r['건수'].cumsum()
        m_labels = monthly_filtered_r['월'].tolist()
        m_counts = monthly_filtered_r['건수'].tolist()
        m_revenues = monthly_filtered_r['매출'].tolist()
        m_commissions = [0] * len(m_labels)
        m_cum = monthly_filtered_r['누적건수'].tolist()

    col_c1, col_c2 = st.columns(2)
    with col_c1:
        fig1 = go.Figure()
        fig1.add_trace(go.Bar(x=m_labels, y=m_counts, name='완료 건수', marker_color=BRAND_ORANGE, opacity=0.8, text=m_counts, textposition='outside'))
        fig1.add_trace(go.Scatter(x=m_labels, y=m_cum, name='누적 건수', line=dict(color=BRAND_PLUM, width=2.5), mode='lines+markers', yaxis='y2'))
        fig1.update_layout(
            title="월별 완료 건수 & 누적",
            yaxis=dict(title="건수"), yaxis2=dict(title="누적", overlaying='y', side='right'),
            legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5),
            height=420, margin=dict(t=50, b=80),
        )
        st.plotly_chart(fig1, use_container_width=True)

    with col_c2:
        fig2 = go.Figure()
        fig2.add_trace(go.Bar(x=m_labels, y=m_revenues, name='정산 매출', marker_color=BRAND_ORANGE, opacity=0.8, text=[format_krw(v) for v in m_revenues], textposition='outside'))
        if any(c > 0 for c in m_commissions):
            fig2.add_trace(go.Bar(x=m_labels, y=m_commissions, name='수수료', marker_color=BRAND_GREEN, opacity=0.7))
        fig2.update_layout(
            title="월별 매출 & 수수료",
            yaxis=dict(title="금액 (원)"), barmode='group',
            legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5),
            height=420, margin=dict(t=50, b=80),
        )
        st.plotly_chart(fig2, use_container_width=True)

    # --- MoM 테이블 (필터 반응) ---
    if len(m_labels) > 0:
        st.subheader("월별 성장률 (MoM)")
        st.caption("활성 병원 수 = 그 달에 시술/수술 완료를 1건 이상 보낸 병원 수 (예약이 여러 병원에 고루 분산되는지 확인)")
        # 월별 활성 병원 수 (예약완료 기준 고유 병원 수)
        active_by_month = (
            filtered_res[filtered_res['병원명'].notna()].groupby('월')['병원명'].nunique()
            if len(filtered_res) > 0 else pd.Series(dtype=int)
        )
        mom_data = []
        for i in range(len(m_labels)):
            row = {'월': m_labels[i], '완료건수': m_counts[i], '정산매출': m_revenues[i], '수수료': m_commissions[i]}
            row['객단가'] = m_revenues[i] / m_counts[i] if m_counts[i] > 0 else 0
            active = int(active_by_month.get(m_labels[i], 0))
            row['활성병원수'] = active
            if i > 0:
                prev_active = int(active_by_month.get(m_labels[i - 1], 0))
                row['활성병원MoM'] = f"{active - prev_active:+d}"
            else:
                row['활성병원MoM'] = '-'
            if i > 0 and m_counts[i - 1] > 0:
                row['건수MoM'] = f"{(m_counts[i] - m_counts[i-1]) / m_counts[i-1] * 100:+.1f}%"
            else:
                row['건수MoM'] = '-'
            if i > 0 and m_revenues[i - 1] > 0:
                row['매출MoM'] = f"{(m_revenues[i] - m_revenues[i-1]) / m_revenues[i-1] * 100:+.1f}%"
            else:
                row['매출MoM'] = '-'
            mom_data.append(row)
        df_mom = pd.DataFrame(mom_data)
        df_mom_disp = df_mom.copy()
        df_mom_disp['정산매출'] = df_mom_disp['정산매출'].apply(format_krw)
        df_mom_disp['수수료'] = df_mom_disp['수수료'].apply(format_krw)
        df_mom_disp['객단가'] = df_mom_disp['객단가'].apply(format_krw)
        st.dataframe(
            df_mom_disp[['월', '완료건수', '활성병원수', '활성병원MoM', '정산매출', '수수료', '객단가', '건수MoM', '매출MoM']].rename(columns={
                '활성병원수': '활성 병원 수', '활성병원MoM': '활성병원 증감', '건수MoM': '건수 MoM', '매출MoM': '매출 MoM',
            }),
            use_container_width=True, hide_index=True,
        )


# ============================================================
# Tab 2: 국적 분석
# ============================================================
with tab2:
    st.subheader("국적별 고객 분석")
    st.caption(f"데이터 기간: {period_label} | 정산 데이터 기준")

    if len(filtered_set) > 0:
        nat_data = filtered_set.groupby('국적').agg(
            건수=('시술금액', 'count'), 매출=('시술금액', 'sum'), 수수료=('수수료금액', 'sum'),
        ).sort_values('매출', ascending=False).reset_index()
        nat_data['비중'] = (nat_data['건수'] / nat_data['건수'].sum() * 100).round(1)
        nat_data['객단가'] = (nat_data['매출'] / nat_data['건수']).round(0)
        nat_data['국기'] = nat_data['국적'].map(COUNTRY_FLAG).fillna('🌍')
        nat_data['국적표시'] = nat_data['국기'] + ' ' + nat_data['국적']
        label_col = '국적표시'
    else:
        nat_data = filtered_res.groupby('고객국적').agg(
            건수=('고객국적', 'count'), 매출=('실제금액', 'sum'),
        ).sort_values('건수', ascending=False).reset_index()
        nat_data['비중'] = (nat_data['건수'] / nat_data['건수'].sum() * 100).round(1)
        nat_data['객단가'] = (nat_data['매출'] / nat_data['건수']).round(0)
        nat_data['국적표시'] = nat_data['고객국적']
        label_col = '국적표시'

    # 도넛 차트용: 상위 8개 + 기타
    TOP_N = 8
    pie_data = nat_data.head(TOP_N).copy()
    if len(nat_data) > TOP_N:
        etc = nat_data.iloc[TOP_N:]
        etc_row = pd.DataFrame([{
            label_col: '🌐 기타 ' + str(len(etc)) + '개국',
            '건수': etc['건수'].sum(),
            '매출': etc['매출'].sum(),
        }])
        pie_data = pd.concat([pie_data, etc_row], ignore_index=True)

    col1, col2 = st.columns(2)
    with col1:
        fig_d = px.pie(pie_data, values='건수', names=label_col, color_discrete_sequence=CHART_COLORS, hole=0.45)
        fig_d.update_layout(
            title="국적별 예약 비중", height=450,
            legend=dict(font=dict(size=13), orientation="v", yanchor="middle", y=0.5, xanchor="left", x=1.05),
            margin=dict(r=160),
        )
        fig_d.update_traces(textinfo='percent', textfont_size=13, insidetextorientation='horizontal')
        st.plotly_chart(fig_d, use_container_width=True)
    with col2:
        top_n = nat_data.head(10).sort_values('객단가')
        fig_b = px.bar(top_n, x='객단가', y=label_col, orientation='h', color_discrete_sequence=[BRAND_ORANGE],
                       text=top_n['객단가'].apply(format_krw))
        fig_b.update_layout(title="국적별 인당 객단가", height=420)
        fig_b.update_traces(textposition='outside')
        st.plotly_chart(fig_b, use_container_width=True)

    # 상세 테이블
    st.subheader("국적별 상세 데이터")
    disp = nat_data[[label_col, '건수', '비중', '매출', '객단가']].copy()
    if '수수료' in nat_data.columns:
        disp['수수료'] = nat_data['수수료'].apply(format_krw)
    disp.columns = ['국적', '건수', '비중(%)', '매출', '객단가'] + (['수수료'] if '수수료' in disp.columns else [])
    disp['매출'] = nat_data['매출'].apply(format_krw)
    disp['객단가'] = nat_data['객단가'].apply(format_krw)
    st.dataframe(disp, use_container_width=True, hide_index=True)

    # 월별 국적 추이
    st.subheader("월별 국적 인입 트렌드")
    nat_col = '국적' if len(filtered_set) > 0 else '고객국적'
    month_col = '정산월' if len(filtered_set) > 0 else '월'
    src = filtered_set if len(filtered_set) > 0 else filtered_res
    top5 = nat_data.head(5)[nat_data.columns[0]].tolist() if '국적' not in nat_data.columns else nat_data.head(5)['국적'].tolist()
    if len(src) > 0:
        trend = src[src[nat_col].isin(top5)].groupby([month_col, nat_col]).size().reset_index(name='건수')
        if len(trend) > 0:
            fig_t = px.line(trend, x=month_col, y='건수', color=nat_col, color_discrete_sequence=CHART_COLORS, markers=True)
            fig_t.update_layout(height=420, margin=dict(t=20, b=80), legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5, font=dict(size=12)))
            st.plotly_chart(fig_t, use_container_width=True)


# ============================================================
# Tab 3: 병원 분석
# ============================================================
with tab3:
    st.subheader("병원별 성과")
    st.caption(f"데이터 기간: {period_label} | 정산 데이터 기준")

    # 내부리포트 병원별 성과가 있으면 우선 사용
    if df_hosp_perf is not None and len(df_hosp_perf) > 0:
        st.markdown("*정산 기준 누적 데이터*")
        # 정규화된 이름 기준으로 합산 (같은 병원 다른 표기 합치기)
        hosp_agg = df_hosp_perf.groupby('병원명').agg(
            누적건수=('누적건수', 'sum'), 누적시수술금액=('누적시수술금액', 'sum'),
            누적수수료=('누적수수료', 'sum'), 최신월건수=('최신월건수', 'sum'), 최신월금액=('최신월금액', 'sum'),
        ).sort_values('누적시수술금액', ascending=False).reset_index()
        hosp_agg['순위'] = range(1, len(hosp_agg) + 1)

        fig_h = px.bar(
            hosp_agg.head(15).sort_values('누적시수술금액'), x='누적시수술금액', y='병원명',
            orientation='h', color_discrete_sequence=[BRAND_ORANGE],
            text=[format_krw(v) for v in hosp_agg.head(15).sort_values('누적시수술금액')['누적시수술금액']],
        )
        fig_h.update_layout(title=f"병원 누적 매출 순위 TOP 15", height=max(400, 15 * 35), margin=dict(l=180, t=50))
        fig_h.update_traces(textposition='outside')
        st.plotly_chart(fig_h, use_container_width=True)

        # 최신월 = 정산 데이터 기준 최신월 (라벨/건수/매출/수수료 모두 이 달로 통일)
        latest_settle_month = all_months_settle[-1] if all_months_settle else None
        # 최신월 건수/매출: 같은 달의 예약완료 기준으로 재계산 (df_hosp의 '최신월'은
        # 예약완료 중 가장 늦은 달을 잡아 라벨과 어긋나므로 여기서 덮어씀)
        if latest_settle_month:
            lm_res = df_completed[df_completed['월'] == latest_settle_month].groupby('병원명').agg(
                _건수=('실제금액', 'size'), _금액=('실제금액', 'sum')).reset_index()
            hosp_agg = hosp_agg.drop(columns=['최신월건수', '최신월금액']).merge(
                lm_res.rename(columns={'_건수': '최신월건수', '_금액': '최신월금액'}), on='병원명', how='left')
            hosp_agg[['최신월건수', '최신월금액']] = hosp_agg[['최신월건수', '최신월금액']].fillna(0)
            hosp_agg['최신월건수'] = hosp_agg['최신월건수'].astype(int)
        if latest_settle_month and len(df_settle) > 0:
            latest_comm = df_settle[df_settle['정산월'] == latest_settle_month].groupby('병원명')['수수료금액'].sum().reset_index()
            latest_comm.columns = ['병원명', '최신월수수료']
            hosp_agg = hosp_agg.merge(latest_comm, on='병원명', how='left')
            hosp_agg['최신월수수료'] = hosp_agg['최신월수수료'].fillna(0)
        else:
            hosp_agg['최신월수수료'] = 0

        disp_h = hosp_agg[['순위', '병원명', '누적건수', '누적시수술금액', '누적수수료', '최신월건수', '최신월금액', '최신월수수료']].copy()
        disp_h['누적시수술금액'] = disp_h['누적시수술금액'].apply(format_krw)
        disp_h['누적수수료'] = disp_h['누적수수료'].apply(format_krw)
        disp_h['최신월금액'] = disp_h['최신월금액'].apply(format_krw)
        disp_h['최신월수수료'] = disp_h['최신월수수료'].apply(format_krw)
        month_label = latest_settle_month if latest_settle_month else '최신월'
        disp_h.columns = ['순위', '병원명', '누적건수', '누적매출', '누적수수료', f'{month_label} 건수', f'{month_label} 매출', f'{month_label} 수수료']
        st.dataframe(disp_h, use_container_width=True, hide_index=True)

    elif len(filtered_set) > 0:
        hosp_s = filtered_set.groupby('병원명').agg(
            건수=('시술금액', 'count'), 매출=('시술금액', 'sum'), 수수료=('수수료금액', 'sum'),
        ).sort_values('매출', ascending=False).reset_index()
        hosp_s['객단가'] = (hosp_s['매출'] / hosp_s['건수']).round(0)
        hosp_s['순위'] = range(1, len(hosp_s) + 1)

        fig_h = px.bar(hosp_s.head(15).sort_values('매출'), x='매출', y='병원명', orientation='h',
                       color_discrete_sequence=[BRAND_ORANGE], text=[format_krw(v) for v in hosp_s.head(15).sort_values('매출')['매출']])
        fig_h.update_layout(title="병원 매출 순위 TOP 15", height=max(400, 15*35), margin=dict(l=180, t=50))
        fig_h.update_traces(textposition='outside')
        st.plotly_chart(fig_h, use_container_width=True)

        disp_h = hosp_s[['순위', '병원명', '건수', '매출', '수수료', '객단가']].copy()
        disp_h['매출'] = disp_h['매출'].apply(format_krw)
        disp_h['수수료'] = disp_h['수수료'].apply(format_krw)
        disp_h['객단가'] = disp_h['객단가'].apply(format_krw)
        st.dataframe(disp_h, use_container_width=True, hide_index=True)

    # 병원별 월별 추이
    st.subheader("병원별 월별 매출 추이")
    if len(df_settle) > 0:
        top_h = list(filtered_set.groupby('병원명')['시술금액'].sum().sort_values(ascending=False).head(10).index) if len(filtered_set) > 0 else []
        sel_h = st.multiselect("병원 선택", options=top_h + [h for h in all_hospitals if h not in top_h], default=top_h[:5], max_selections=10, key="hosp_trend")
        if sel_h:
            hm = df_settle[df_settle['병원명'].isin(sel_h)].groupby(['정산월', '병원명'])['시술금액'].sum().reset_index()
            fig_ht = px.line(hm, x='정산월', y='시술금액', color='병원명', color_discrete_sequence=CHART_COLORS, markers=True)
            fig_ht.update_layout(title="선택 병원 월별 매출", yaxis_title="정산 매출", height=480, margin=dict(t=40, b=100), legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5, font=dict(size=11)))
            st.plotly_chart(fig_ht, use_container_width=True)

    # 월별 1위
    st.subheader("월별 매출 1위 병원")
    if len(df_settle) > 0:
        top1 = []
        for m in sorted(all_months_settle):
            md = df_settle[df_settle['정산월'] == m]
            if len(md) > 0:
                hr = md.groupby('병원명')['시술금액'].sum().sort_values(ascending=False)
                top1.append({'월': m, '1위 병원': hr.index[0], '매출': format_krw(hr.values[0]), '건수': len(md[md['병원명'] == hr.index[0]])})
        if top1:
            st.dataframe(pd.DataFrame(top1), use_container_width=True, hide_index=True)

    # ============================================================
    # 병원 개별 리포트 (병원별 트렌드 리포트와 동일한 내용)
    # ============================================================
    st.divider()
    st.subheader("🏥 병원 개별 리포트")
    st.caption("병원을 선택하면 해당 병원에 제공하는 트렌드 리포트와 동일한 내용을 한 번에 볼 수 있어요.")

    # 누적 매출 순으로 병원 정렬 → 선택 박스 (상위 병원이 먼저)
    _rev_rank = df_completed[df_completed['병원명'].notna()].groupby('병원명')['실제금액'].sum().sort_values(ascending=False)
    _hosp_options = list(_rev_rank.index) + [h for h in all_hospitals if h not in _rev_rank.index]
    if _hosp_options:
        rep_hosp = st.selectbox("병원 선택", _hosp_options, key="hosp_report_pick")

        hc = df_completed[df_completed['병원명'] == rep_hosp].copy()
        hs = df_settle[df_settle['병원명'] == rep_hosp].copy() if len(df_settle) > 0 else pd.DataFrame(columns=['정산월', '시술금액', '수수료금액'])
        hc_months = sorted([m for m in hc['월'].unique() if m and m != 'NaT'])
        hs_months = sorted([m for m in hs['정산월'].unique()]) if len(hs) > 0 else []

        if len(hc) == 0 and len(hs) == 0:
            st.info("선택한 병원의 데이터가 없습니다.")
        else:
            ref_month = hc_months[-1] if hc_months else (hs_months[-1] if hs_months else None)

            # --- 헤더: 순위 + 누적 요약 ---
            total_h = len(_rev_rank)
            rank = list(_rev_rank.index).index(rep_hosp) + 1 if rep_hosp in _rev_rank.index else None
            cum_cnt, cum_rev = len(hc), hc['실제금액'].sum()
            cum_comm = hs['수수료금액'].sum() if len(hs) > 0 else 0
            rank_txt = f" · 전체 {total_h}개 병원 중 누적 매출 **{rank}위**" if rank else ""
            if rank and rank <= max(1, int(total_h * 0.2)):
                rank_txt += " 🏆 **TOP 20%**"
            st.markdown(f"### {rep_hosp}{rank_txt}")
            s1, s2, s3 = st.columns(3)
            s1.metric("누적 완료 건수", f"{cum_cnt:,}건")
            s2.metric("누적 매출", format_krw(cum_rev))
            s3.metric("누적 수수료", format_krw(cum_comm))

            # --- 기준월 성과 (MoM) ---
            if ref_month:
                st.markdown(f"#### 📊 {ref_month} 성과")
                cur = hc[hc['월'] == ref_month]
                cur_cnt, cur_rev = len(cur), cur['실제금액'].sum()
                cur_comm = hs[hs['정산월'] == ref_month]['수수료금액'].sum() if len(hs) > 0 else 0
                aov = cur_rev / cur_cnt if cur_cnt else 0
                # 전월
                idx = hc_months.index(ref_month) if ref_month in hc_months else -1
                prev_month = hc_months[idx - 1] if idx > 0 else None
                d_cnt = d_rev = None
                if prev_month:
                    p = hc[hc['월'] == prev_month]
                    pc, pr = len(p), p['실제금액'].sum()
                    if pc > 0:
                        d_cnt = f"{(cur_cnt - pc) / pc * 100:+.1f}% vs 전월"
                    if pr > 0:
                        d_rev = f"{(cur_rev - pr) / pr * 100:+.1f}% vs 전월"
                k1, k2, k3, k4 = st.columns(4)
                k1.metric("완료 건수", f"{cur_cnt}건", delta=d_cnt)
                k2.metric("매출", format_krw(cur_rev), delta=d_rev)
                k3.metric("수수료", format_krw(cur_comm))
                k4.metric("객단가", format_krw(aov))

            # --- 월별 추이 ---
            st.markdown("#### 📈 월별 추이")
            mc = hc[hc['월'] != 'NaT'].groupby('월').agg(완료건수=('실제금액', 'size'), 매출=('실제금액', 'sum'))
            ms = hs.groupby('정산월').agg(수수료=('수수료금액', 'sum')) if len(hs) > 0 else pd.DataFrame()
            all_m = sorted(set(mc.index) | set(ms.index))
            trows, cumc = [], 0
            for m in all_m:
                c = int(mc.loc[m, '완료건수']) if m in mc.index else 0
                r = float(mc.loc[m, '매출']) if m in mc.index else 0.0
                comm = float(ms.loc[m, '수수료']) if (len(ms) > 0 and m in ms.index) else 0.0
                cumc += c
                trows.append({'월': m, '완료건수': c, '매출': r, '수수료': comm, '누적건수': cumc})
            tdf = pd.DataFrame(trows)
            if len(tdf) > 0:
                fig_ph = go.Figure()
                fig_ph.add_trace(go.Bar(x=tdf['월'], y=tdf['완료건수'], name='완료 건수', marker_color=BRAND_ORANGE, opacity=0.85, text=tdf['완료건수'], textposition='outside'))
                fig_ph.add_trace(go.Scatter(x=tdf['월'], y=tdf['매출'], name='매출', yaxis='y2', line=dict(color=BRAND_PLUM, width=2.5), mode='lines+markers'))
                fig_ph.update_layout(
                    yaxis=dict(title="완료 건수"), yaxis2=dict(title="매출(원)", overlaying='y', side='right'),
                    height=380, margin=dict(t=30, b=80), legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5),
                )
                st.plotly_chart(fig_ph, use_container_width=True)
                # 표 + MoM
                disp_t = tdf.copy()
                disp_t['건수MoM'] = ['-'] + [f"{(tdf['완료건수'][i] - tdf['완료건수'][i-1]) / tdf['완료건수'][i-1] * 100:+.1f}%" if tdf['완료건수'][i-1] > 0 else '-' for i in range(1, len(tdf))]
                disp_t['매출MoM'] = ['-'] + [f"{(tdf['매출'][i] - tdf['매출'][i-1]) / tdf['매출'][i-1] * 100:+.1f}%" if tdf['매출'][i-1] > 0 else '-' for i in range(1, len(tdf))]
                disp_t['매출'] = disp_t['매출'].apply(format_krw)
                disp_t['수수료'] = disp_t['수수료'].apply(format_krw)
                disp_t = disp_t[['월', '완료건수', '매출', '수수료', '누적건수', '건수MoM', '매출MoM']].rename(columns={'건수MoM': '건수 MoM', '매출MoM': '매출 MoM'})
                st.dataframe(disp_t, use_container_width=True, hide_index=True)

            # --- 국적 분포 (기준월, 없으면 누적) ---
            nat_src = hc[hc['월'] == ref_month] if (ref_month and len(hc[hc['월'] == ref_month]) > 0) else hc
            nat_label = ref_month if (ref_month and len(hc[hc['월'] == ref_month]) > 0) else '누적'
            natg = nat_src[nat_src['고객국적'].astype(str).str.strip().replace('nan', '') != ''].groupby('고객국적').agg(
                건수=('실제금액', 'size'), 매출=('실제금액', 'sum')).sort_values('매출', ascending=False).reset_index()
            if len(natg) > 0:
                st.markdown(f"#### 🌏 고객 국적 분포 ({nat_label})")
                natg['객단가'] = (natg['매출'] / natg['건수']).round(0)
                cc1, cc2 = st.columns([1, 1])
                with cc1:
                    fig_n = px.bar(natg.head(8).sort_values('건수'), x='건수', y='고객국적', orientation='h', color_discrete_sequence=[BRAND_PLUM], text='건수')
                    fig_n.update_layout(height=max(280, len(natg.head(8)) * 36), margin=dict(l=80, t=20, b=20), yaxis_title="")
                    fig_n.update_traces(textposition='outside')
                    st.plotly_chart(fig_n, use_container_width=True)
                with cc2:
                    nd = natg.head(10).copy()
                    nd['매출'] = nd['매출'].apply(format_krw)
                    nd['객단가'] = nd['객단가'].apply(format_krw)
                    st.dataframe(nd.rename(columns={'고객국적': '국적'}), use_container_width=True, hide_index=True, height=360)

            # --- 인기 시술 (누적) + 성장 기회 ---
            cg1, cg2 = st.columns([1, 1])
            with cg1:
                st.markdown("#### 💉 인기 시술 (누적)")
                proc_cnt = {}
                for txt in hc['시술수술명']:
                    for p in extract_procedures(txt):
                        proc_cnt[p] = proc_cnt.get(p, 0) + 1
                if proc_cnt:
                    pdf = pd.DataFrame(sorted(proc_cnt.items(), key=lambda x: -x[1])[:10], columns=['시술', '건수'])
                    st.dataframe(pdf, use_container_width=True, hide_index=True, height=360)
                else:
                    st.caption("시술 데이터 없음")
            with cg2:
                st.markdown("#### 🚀 성장 기회")
                plat = {}
                for txt in df_completed['시술수술명']:
                    for p in extract_procedures(txt):
                        plat[p] = plat.get(p, 0) + 1
                plat_top = [p for p, _ in sorted(plat.items(), key=lambda x: -x[1])[:12]]
                hosp_proc_set = set(proc_cnt.keys())
                opps = [p for p in plat_top if p not in hosp_proc_set][:5]
                if opps:
                    st.info("**언니가이드 전체 인기 시술 중 귀원이 아직 적극적으로 하지 않는 시술:**\n\n" + "  ·  ".join(opps) + "\n\n→ 이 시술들을 강화하면 유입 확대 여지가 있어요.")
                else:
                    st.success("전체 인기 시술을 대부분 보유하고 있어요 👍")
            st.caption("※ 카톡/이메일로 보낼 공식 HTML 리포트는 '리포트 생성' 탭에서 받을 수 있어요.")


# ============================================================
# Tab 4: 시술 트렌드
# ============================================================
with tab4:
    st.subheader("시술 트렌드")
    st.caption(f"데이터 기간: {period_label} | 전체 예약 데이터 기준 (시/수술 완료 + 예약확정 포함)")

    # 시술 데이터: 완료건뿐 아니라 전체 예약 중 시술명이 있는 건 활용
    if df_all is not None:
        proc_source = df_all[df_all['월'].isin(selected_months)].copy()
        if selected_nationalities:
            proc_source = proc_source[proc_source['고객국적'].isin(selected_nationalities)]
        if selected_hospitals:
            proc_source = proc_source[proc_source['병원명'].isin(selected_hospitals)]
    else:
        proc_source = filtered_res.copy()

    proc_source_valid = proc_source[proc_source['시술수술명'].str.strip() != ''].copy()

    # ---------- 시술명 입력 커버율 ----------
    # 어드민 이관 중 treatments가 후순위로 입력돼 월별 충족률이 크게 다르다.
    # 커버율을 숨기면 "울쎄라 1위" 같은 순위를 실제보다 확신하게 되므로 항상 노출한다.
    if len(proc_source) > 0:
        cov_rate = len(proc_source_valid) / len(proc_source) * 100
        cov_by_month = (proc_source.assign(_has=proc_source['시술수술명'].str.strip() != '')
                        .groupby('월')['_has'].agg(['sum', 'size']))
        cov_by_month = cov_by_month[cov_by_month.index != 'NaT']
        weak = [f"{m} {r['sum'] / r['size'] * 100:.0f}%"
                for m, r in cov_by_month.iterrows() if r['size'] and r['sum'] / r['size'] < 0.8]
        tone = ('#D63031', '#FFF5F5') if cov_rate < 70 else (
            ('#E17055', '#FFF9F3') if cov_rate < 90 else (BRAND_GREEN, '#F4FBF6'))
        st.markdown(f"""
        <div style="background:{tone[1]}; border-left:3px solid {tone[0]};
             padding:10px 14px; border-radius:4px; margin-bottom:14px; font-size:12.5px; line-height:1.7;">
            <b style="color:{tone[0]};">시술명 입력 커버율 {cov_rate:.1f}%</b>
            <span style="color:#666;">— 전체 {len(proc_source):,}건 중 {len(proc_source_valid):,}건에만 시술명이 있습니다.
            아래 순위·비중은 <b>입력된 건만</b>의 분포입니다.</span>
            {f'<br><span style="color:#888;">커버율 낮은 월: {" · ".join(weak)}</span>' if weak else ''}
        </div>
        """, unsafe_allow_html=True)

    col1, col2 = st.columns(2)

    with col1:
        # 시술/수술 비중 - 내부리포트 or 정산 기준
        if df_monthly is not None:
            m_filtered = df_monthly[df_monthly['월'].isin(selected_months)] if len(df_monthly) > 0 else df_monthly
            total_시술 = m_filtered['시술건수'].sum()
            total_수술 = m_filtered['수술건수'].sum()
            type_df = pd.DataFrame({'구분': ['시술', '수술'], '건수': [int(total_시술), int(total_수술)]})
        elif len(filtered_set) > 0:
            tc = filtered_set['구분'].value_counts().reset_index()
            tc.columns = ['구분', '건수']
            type_df = tc[tc['구분'].str.strip() != '']
        else:
            tc = proc_source['종류'].value_counts().reset_index()
            tc.columns = ['구분', '건수']
            type_df = tc[tc['구분'].str.strip() != '']

        if len(type_df) > 0:
            fig_ty = px.pie(type_df, values='건수', names='구분', color_discrete_sequence=[BRAND_ORANGE, BRAND_PLUM, BRAND_GREEN], hole=0.45)
            fig_ty.update_layout(title="시술 vs 수술 비중 (정산 기준)", height=380)
            fig_ty.update_traces(textinfo='percent+label')
            st.plotly_chart(fig_ty, use_container_width=True)

    with col2:
        # 키워드 기반 시술 카테고리 집계
        all_proc_rows = []
        for _, row in proc_source_valid.iterrows():
            procs = extract_procedures(row['시술수술명'])
            for p in procs:
                all_proc_rows.append({'시술카테고리': p, '매출': row['실제금액']})
        df_proc_kw = pd.DataFrame(all_proc_rows)

        if len(df_proc_kw) > 0:
            top_kw = df_proc_kw.groupby('시술카테고리').agg(
                건수=('시술카테고리', 'count'), 매출=('매출', 'sum'),
            ).sort_values('건수', ascending=False).head(15).reset_index()
            fig_p = px.bar(
                top_kw.sort_values('건수'), x='건수', y='시술카테고리',
                orientation='h', color_discrete_sequence=[BRAND_ORANGE], text='건수',
            )
            fig_p.update_layout(title=f"시술 카테고리 TOP 15 ({period_label})", height=480, margin=dict(l=160, t=50))
            fig_p.update_traces(textposition='outside')
            st.plotly_chart(fig_p, use_container_width=True)

    # 시술 카테고리 상세 테이블
    st.subheader("시술 카테고리별 상세 데이터")
    if len(df_proc_kw) > 0:
        top_kw_full = df_proc_kw.groupby('시술카테고리').agg(
            건수=('시술카테고리', 'count'), 매출=('매출', 'sum'),
        ).sort_values('건수', ascending=False).reset_index()
        top_kw_full['매출표시'] = top_kw_full['매출'].apply(format_krw)
        top_kw_full['객단가'] = (top_kw_full['매출'] / top_kw_full['건수']).apply(format_krw)
        st.dataframe(
            top_kw_full[['시술카테고리', '건수', '매출표시', '객단가']].rename(columns={'시술카테고리': '시술 카테고리', '매출표시': '총 매출'}),
            use_container_width=True, hide_index=True,
        )
        st.caption("* 하나의 예약에 여러 시술이 포함된 경우 각각 카운트됩니다.")

    # 국적별 선호 시술 (키워드 기반)
    st.subheader("국적별 선호 시술 TOP 5")
    top_c = nat_data.head(5)['국적'].tolist() if '국적' in nat_data.columns else nat_data.head(5)['고객국적'].tolist() if '고객국적' in nat_data.columns else []
    if top_c:
        cols_p = st.columns(min(len(top_c), 3))
        for i, country in enumerate(top_c):
            with cols_p[i % 3]:
                flag = COUNTRY_FLAG.get(country, '🌍')
                c_procs = proc_source_valid[proc_source_valid['고객국적'] == country]['시술수술명']
                kw_list = []
                for txt in c_procs:
                    kw_list.extend(extract_procedures(txt))
                cp = pd.Series(kw_list).value_counts().head(5)
                if len(cp) > 0:
                    st.markdown(f"**{flag} {country}**")
                    for proc, cnt in cp.items():
                        st.markdown(f"- {proc[:35]} ({cnt}건)")
                    st.markdown("")


# ============================================================
# Tab 5: 취소/No-show
# ============================================================
with tab5:
    st.subheader("취소 / No-show 트래킹")
    st.caption("데이터 기간: 전체 누적 | 내부리포트 기준")

    if cancel_summary is not None:
        # 전체 요약
        c1, c2, c3 = st.columns(3)
        c1.metric("총 취소", f"{cancel_summary['total_cancel']}건")
        c2.metric("총 No-show", f"{cancel_summary['total_noshow']}건")
        c3.metric("취소+노쇼율", cancel_summary['cancel_rate'])

        st.markdown("")

        # 월별 취소+노쇼 추이
        if df_monthly is not None:
            st.subheader("월별 취소+노쇼 추이")
            fig_cn = go.Figure()
            fig_cn.add_trace(go.Bar(
                x=df_monthly['월'], y=df_monthly['취소노쇼'].apply(lambda x: int(x) if pd.notna(x) else 0),
                name='취소+노쇼', marker_color=BRAND_RED, opacity=0.8,
                text=df_monthly['취소노쇼'].apply(lambda x: int(x) if pd.notna(x) else 0), textposition='outside',
            ))
            cancel_rate_monthly = (df_monthly['취소노쇼'] / (df_monthly['완료건수'] + df_monthly['취소노쇼']) * 100).round(1)
            fig_cn.add_trace(go.Scatter(
                x=df_monthly['월'], y=cancel_rate_monthly,
                name='취소+노쇼율(%)', line=dict(color=BRAND_PLUM, width=2.5),
                mode='lines+markers+text', yaxis='y2',
                text=[f"{v}%" for v in cancel_rate_monthly], textposition='top center',
            ))
            fig_cn.update_layout(
                yaxis=dict(title="건수"), yaxis2=dict(title="비율(%)", overlaying='y', side='right'),
                height=360, legend=dict(orientation="h", yanchor="bottom", y=1.08),
            )
            st.plotly_chart(fig_cn, use_container_width=True)

        # 병원별 취소/노쇼 현황
        if df_cancel_hospital is not None and len(df_cancel_hospital) > 0:
            st.subheader("병원별 취소/No-show 현황")

            fig_ch = go.Figure()
            df_ch = df_cancel_hospital.sort_values('취소노쇼율', ascending=True)
            fig_ch.add_trace(go.Bar(x=df_ch['취소'], y=df_ch['병원명'], name='취소', orientation='h', marker_color=BRAND_ORANGE))
            fig_ch.add_trace(go.Bar(x=df_ch['No-show'], y=df_ch['병원명'], name='No-show', orientation='h', marker_color=BRAND_RED))
            fig_ch.update_layout(
                title="병원별 취소 & No-show", barmode='stack',
                height=max(400, len(df_ch) * 28), margin=dict(l=180, t=50),
                legend=dict(orientation="h", yanchor="bottom", y=1.08),
            )
            st.plotly_chart(fig_ch, use_container_width=True)

            # 테이블
            disp_cn = df_cancel_hospital.sort_values('취소노쇼율', ascending=False).copy()
            disp_cn['취소노쇼율'] = (disp_cn['취소노쇼율'] * 100).round(1).astype(str) + '%'
            st.dataframe(disp_cn, use_container_width=True, hide_index=True)

        # 상세 내역
        if df_cancel_detail is not None and len(df_cancel_detail) > 0:
            st.subheader("취소/No-show 상세 내역")
            # 필터
            status_filter = st.multiselect("상태", df_cancel_detail['상태'].unique().tolist(), default=df_cancel_detail['상태'].unique().tolist(), key="cn_status")
            filtered_cn = df_cancel_detail[df_cancel_detail['상태'].isin(status_filter)]
            if selected_hospitals:
                filtered_cn = filtered_cn[filtered_cn['병원명'].isin(selected_hospitals)]
            st.dataframe(filtered_cn.sort_values('월', ascending=False), use_container_width=True, hide_index=True, height=400)

    else:
        st.info("내부리포트 Excel을 업로드하면 취소/No-show 데이터를 볼 수 있습니다.")
        # 운영 데이터에서 기본 취소/노쇼 추출
        if df_all is not None:
            cancel_df = df_all[df_all['예약상태'].isin(['예약 취소'])].copy()
            noshow_df = df_all[df_all['예약상태'].str.lower().str.contains('no-show|no show|noshow', na=False)].copy()
            c1, c2 = st.columns(2)
            c1.metric("예약 취소 (전체)", f"{len(cancel_df)}건")
            c2.metric("No-show (전체)", f"{len(noshow_df)}건")


# ============================================================
# Tab 6: 원본 데이터
# ============================================================
with tab6:
    st.subheader("원본 데이터 조회 및 다운로드")
    data_type = st.radio("데이터 선택", ["예약 완료 데이터", "정산 데이터", "전체 예약 (취소/노쇼 포함)"], horizontal=True)

    if data_type == "예약 완료 데이터":
        cols = ['월', '_source', '병원명', '고객국적', '종류', '시술수술명', '실제금액']
        avail = [c for c in cols if c in filtered_res.columns]
        df_d = filtered_res[avail].rename(columns={'_source': '원천'}).sort_values('월', ascending=False)
        st.dataframe(df_d, use_container_width=True, hide_index=True, height=500)
        st.download_button("CSV 다운로드", df_d.to_csv(index=False).encode('utf-8-sig'), "예약완료_필터.csv", "text/csv")

    elif data_type == "정산 데이터":
        if len(filtered_set) > 0:
            cols = ['정산월', '병원명', '국적', '구분', '시술금액', '수수료금액']
            avail = [c for c in cols if c in filtered_set.columns]
            df_d = filtered_set[avail].sort_values('정산월', ascending=False)
            st.dataframe(df_d, use_container_width=True, hide_index=True, height=500)
            st.download_button("CSV 다운로드", df_d.to_csv(index=False).encode('utf-8-sig'), "정산_필터.csv", "text/csv")
        else:
            st.info("정산 데이터가 없습니다.")

    else:
        if df_all is not None:
            cols = ['월', '병원명', '고객국적', '예약상태', '종류', '시술수술명', '실제금액']
            avail = [c for c in cols if c in df_all.columns]
            mask = df_all['월'].isin(selected_months)
            if selected_nationalities:
                mask = mask & df_all['고객국적'].isin(selected_nationalities)
            if selected_hospitals:
                mask = mask & df_all['병원명'].isin(selected_hospitals)
            df_d = df_all[mask][avail].sort_values('월', ascending=False)
            st.dataframe(df_d, use_container_width=True, hide_index=True, height=500)
            st.download_button("CSV 다운로드", df_d.to_csv(index=False).encode('utf-8-sig'), "전체예약_필터.csv", "text/csv")


# ============================================================
# Tab 7: 리포트 생성 (팀원 누구나 사용 가능)
# ============================================================
with tab7:
    st.subheader("📦 병원용 HTML 리포트 생성")
    st.caption("버튼 클릭 한 번으로 공통 리포트 + 병원별 34개 리포트를 생성하여 ZIP 파일로 다운로드합니다.")

    st.markdown("**생성 기준월 선택**")
    report_month = st.selectbox(
        "리포트 월",
        options=all_months,
        index=len(all_months) - 1 if all_months else 0,
        format_func=lambda x: f"{x} ({datetime.strptime(x + '-01', '%Y-%m-%d').strftime('%Y년 %m월')})",
        key="report_month_sel",
    )

    # generate_report.py는 구글시트만 읽는다 → 컷오프 이후 월은 데이터가 거의 비어 있다.
    if report_month >= ADMIN_CUTOFF_MONTH:
        st.error(f"""
**{report_month}은 아직 일괄 생성을 지원하지 않습니다.**
{ADMIN_CUTOFF_MONTH} 이후 실적은 어드민 원천에 있고, 이 일괄 생성기는 구글시트만 읽습니다.
그대로 생성하면 대부분 빈 리포트가 나옵니다.
→ 그 사이에는 **🏥 병원 분석 탭 하단의 '병원 개별 리포트'**를 쓰세요 (어드민 데이터 반영됨).
        """)

    st.markdown("")
    st.info("""
**생성 프로세스:**
1. 아래 버튼 클릭 → 스크립트 실행 (약 10-30초 소요)
2. ZIP 파일 자동 다운로드
3. 압축 풀면 공통 리포트 + 병원별 34개 HTML 파일
4. 각 병원에 카톡/이메일로 개별 전달
    """)

    if st.button("🚀 리포트 일괄 생성 + ZIP 다운로드", type="primary", use_container_width=True):
        import subprocess
        import zipfile
        import tempfile
        import shutil

        with st.spinner(f"{report_month} 리포트 생성 중..."):
            try:
                # Google Sheets URL에서 임시 xlsx 다운로드
                import urllib.request
                tmp_xlsx = "/tmp/unniguide_report_input.xlsx"
                urllib.request.urlretrieve(
                    f"https://docs.google.com/spreadsheets/d/{GSHEET_ID_MAIN}/export?format=xlsx",
                    tmp_xlsx,
                )

                # generate_report.py 실행
                script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "generate_report.py")
                result = subprocess.run(
                    ["python3", script_path, tmp_xlsx, report_month],
                    capture_output=True, text=True, timeout=300,
                )

                if result.returncode != 0:
                    st.error(f"리포트 생성 실패: {result.stderr[:500]}")
                else:
                    st.success("리포트 생성 완료!")

                    # ZIP 만들기
                    output_dir = os.path.dirname(os.path.abspath(__file__))
                    month_str = report_month.replace("-", "")
                    common_html = os.path.join(output_dir, f"unniguide_report_{month_str}.html")
                    hospital_dir = os.path.join(output_dir, "hospitals")

                    with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as tmp_zip:
                        zip_path = tmp_zip.name

                    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                        if os.path.exists(common_html):
                            zf.write(common_html, f"00_공통_트렌드_리포트_{month_str}.html")
                        if os.path.exists(hospital_dir):
                            for fname in os.listdir(hospital_dir):
                                if fname.endswith(f"_{month_str}.html"):
                                    zf.write(
                                        os.path.join(hospital_dir, fname),
                                        f"병원별/{fname}",
                                    )

                    with open(zip_path, 'rb') as f:
                        zip_bytes = f.read()

                    st.download_button(
                        label=f"📥 언니가이드_리포트_{month_str}.zip 다운로드",
                        data=zip_bytes,
                        file_name=f"언니가이드_리포트_{month_str}.zip",
                        mime="application/zip",
                        use_container_width=True,
                    )

                    # 실행 로그
                    with st.expander("실행 로그 보기"):
                        st.code(result.stdout)

            except Exception as e:
                st.error(f"오류 발생: {str(e)}")

    st.divider()

    st.subheader("📄 제휴사용 센터 리포트 생성")
    st.caption("아모레퍼시픽 등 외부 제휴 브랜드 공유용 원페이지 HTML 리포트 (A4 PDF 인쇄 최적화)")

    if st.button("🏢 제휴사용 리포트 생성", use_container_width=True):
        import subprocess
        with st.spinner("제휴사용 리포트 생성 중..."):
            try:
                script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "generate_partner_report.py")
                result = subprocess.run(
                    ["python3", script_path],
                    capture_output=True, text=True, timeout=120,
                )

                if result.returncode != 0:
                    st.error(f"생성 실패: {result.stderr[:500]}")
                else:
                    output_dir = os.path.dirname(os.path.abspath(__file__))
                    html_path = os.path.join(output_dir, "unniguide_center_report_202603.html")
                    if os.path.exists(html_path):
                        with open(html_path, 'r', encoding='utf-8') as f:
                            html_content = f.read()
                        st.success("제휴사용 리포트 생성 완료!")
                        st.download_button(
                            label="📥 제휴사용_센터_리포트.html 다운로드",
                            data=html_content.encode('utf-8'),
                            file_name="언니가이드_센터_리포트.html",
                            mime="text/html",
                            use_container_width=True,
                        )
                        st.caption("💡 다운로드한 HTML을 브라우저에서 열고 Cmd+P → PDF로 저장하면 제휴사 공유용 PDF가 됩니다.")
                    else:
                        st.error("리포트 파일을 찾을 수 없습니다.")
            except Exception as e:
                st.error(f"오류: {str(e)}")
